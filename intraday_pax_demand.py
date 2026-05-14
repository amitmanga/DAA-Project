import csv
import json
import os
from copy import deepcopy
from datetime import datetime, time
import re


PULSE_FILE = 'intraday_pax_pulse.json'
FLIGHTS_FILE = 'intraday_pax_flights.csv'
FLIGHT_DELAY_EFFECT_FILE = 'flight_delay_effect.xlsx'
TOUCHPOINTS = ('checkin', 'security', 'cbp', 'lounge', 'boarding', 'immigration', 'baggage')
PAX_BY_ICAO = {'A': 20, 'B': 50, 'C': 180, 'D': 260, 'E': 330, 'F': 500}
PAX_DEFAULT = 150
TOUCHPOINT_COLUMN_LABELS = {
    'checkin': 'Checkin',
    'security': 'Security',
    'cbp': 'CBP',
    'lounge': 'Lounge',
    'boarding': 'Boarding',
    'immigration': 'Immigration',
    'baggage': 'Baggage',
}


def _data_path(base_dir, filename):
    return os.path.join(base_dir, 'data', filename)


def _load_json(base_dir, filename):
    with open(_data_path(base_dir, filename), encoding='utf-8') as f:
        return json.load(f)


def _numbers(values):
    out = []
    for value in values or []:
        try:
            out.append(float(value or 0))
        except (TypeError, ValueError):
            out.append(0.0)
    return out


def _peak_idx(values):
    return max(range(len(values)), key=lambda idx: values[idx]) if values else 0


def _active_avg(values):
    active = [v for v in values if v > 0]
    return round(sum(active) / len(active)) if active else 0


def _pct(part, total):
    return round((part / total) * 100) if total else None


def _peak_time(labels, values):
    idx = _peak_idx(values)
    return labels[idx] if idx < len(labels) else '--'


def _insight(series, labels, key, label, accent, secondary_label, secondary_value, metric_label='Peak Pax / min'):
    values = _numbers(series.get(key))
    return {
        'key': key,
        'label': label,
        'accent': accent,
        'peak': round(max(values) if values else 0),
        'metric_label': metric_label,
        'peak_time': _peak_time(labels, values),
        'secondary_label': secondary_label,
        'secondary_value': secondary_value,
    }


def _load_flights(base_dir):
    path = _data_path(base_dir, FLIGHTS_FILE)
    if not os.path.exists(path):
        return []

    flights = []
    with open(path, encoding='utf-8-sig', newline='') as f:
        for row in csv.DictReader(f):
            flight_no = str(row.get('Flight_No') or '').strip()
            if not flight_no:
                continue
            flights.append({
                'flight_no': flight_no,
                'type': str(row.get('Type') or '').strip(),
                'eta': str(row.get('ETA') or '').strip(),
                'etd': str(row.get('ETD') or '').strip(),
                'aircraft_type': str(row.get('Aircraft_Type') or row.get('Aircraft Type') or row.get('Aircraft') or '').strip(),
                'icao_cat': str(row.get('ICAO_Cat') or row.get('ICAO Cat') or row.get('Category') or '').strip(),
                'airline': str(row.get('Airline') or '').strip(),
                'terminal': str(row.get('Terminal') or '').strip(),
                'cbp_required': str(row.get('CBP_Required') or '').strip(),
            })
    return flights


def _pax_for_flight(flight):
    icao = str(flight.get('icao_cat') or '').strip().upper()
    return PAX_BY_ICAO.get(icao, PAX_DEFAULT)


def _terminal_for_flight(flight):
    terminal = str(flight.get('terminal') or '').strip().upper()
    return terminal if terminal in ('T1', 'T2') else None


def _flight_serves_touchpoint(flight, touchpoint):
    flow_type = str(flight.get('type') or '').strip().upper()
    is_departure_flow = flow_type in ('DEP', 'TURN')
    is_arrival_flow = flow_type in ('ARR', 'TURN')

    if touchpoint in ('immigration', 'baggage'):
        return is_arrival_flow
    if touchpoint == 'cbp':
        return is_departure_flow and str(flight.get('cbp_required') or '').strip().lower() == 'true'
    return is_departure_flow


def _flight_touchpoint_share(flights, selected, touchpoint):
    if not selected or not _flight_serves_touchpoint(selected, touchpoint):
        return 0.0

    eligible = [f for f in flights if _flight_serves_touchpoint(f, touchpoint)]
    total_pax = sum(_pax_for_flight(f) for f in eligible)
    if total_pax <= 0:
        return 0.0
    return min(1.0, max(0.0, _pax_for_flight(selected) / total_pax))


def _terminal_weights(flights, touchpoint):
    counts = {'T1': 0.0, 'T2': 0.0}
    for flight in flights:
        terminal = _terminal_for_flight(flight)
        if not terminal:
            continue

        if not _flight_serves_touchpoint(flight, touchpoint):
            continue

        counts[terminal] += _pax_for_flight(flight)

    total = counts['T1'] + counts['T2']
    if total <= 0:
        return {'T1': 0.5, 'T2': 0.5}
    return {'T1': counts['T1'] / total, 'T2': counts['T2'] / total}


def _aggregate_to_15_minutes(labels, values, touchpoint):
    buckets = {minute: [] for minute in range(0, 1440, 15)}
    for label, value in zip(labels or [], _numbers(values)):
        minute = _mins(label)
        bucket = (minute // 15) * 15
        if bucket in buckets:
            buckets[bucket].append(value)

    aggregated = {}
    for minute, bucket_values in buckets.items():
        if not bucket_values:
            aggregated[minute] = 0
        elif touchpoint == 'lounge':
            aggregated[minute] = round(sum(bucket_values) / len(bucket_values))
        else:
            aggregated[minute] = round(sum(bucket_values))
    return aggregated


def _interval_overlaps_windows(interval_start, windows):
    interval_end = interval_start + 15
    for start, end in windows or []:
        if interval_start < int(end) and interval_end > int(start):
            return True
    return False


def _as_float(value, default=0.0):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return default


def _workbook_col_name(terminal, touchpoint):
    return f"{terminal}_{TOUCHPOINT_COLUMN_LABELS[touchpoint]}"


def _parse_relative_t_offset(raw, positive_after_zero=False):
    text = str(raw or '').strip().upper()
    match = re.match(r'^T\s*([+-])?\s*(\d{1,2}):(\d{2})(?::(\d{2}))?$', text)
    if not match:
        return None

    sign_token, hours, minutes, seconds = match.groups()
    offset = int(hours) * 60 + int(minutes) + round(int(seconds or 0) / 60)
    if offset == 0:
        return 0

    # Some Excel exports preserve the visual T+ section as repeated T- text.
    # Once T-00:00:00 has been passed, row order is the reliable direction.
    if positive_after_zero:
        return offset
    return -offset if sign_token != '+' else offset


def _load_flight_delay_overlay_rows(base_dir):
    path = _data_path(base_dir, FLIGHT_DELAY_EFFECT_FILE)
    if not os.path.exists(path):
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
    except Exception:
        return []

    headers = [
        str(value).strip() if value is not None else ''
        for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    overlay_idx = {}
    for touchpoint, label in TOUCHPOINT_COLUMN_LABELS.items():
        target = f"{label} (Overlay)".lower()
        for idx, header in enumerate(headers):
            if header.lower() == target:
                overlay_idx[touchpoint] = idx
                break

    rows = []
    positive_after_zero = False
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not values or values[0] in (None, ''):
            continue
        offset = _parse_relative_t_offset(values[0], positive_after_zero=positive_after_zero)
        if offset is None:
            continue

        rows.append({
            'offset_mins': offset,
            'overlay': {
                touchpoint: _as_float(values[idx] if idx < len(values) else 0)
                for touchpoint, idx in overlay_idx.items()
            },
        })
        if offset == 0:
            positive_after_zero = True

    return rows


def _save_flight_delay_overlay_simulation(base_dir, payload, headers, column_touchpoints):
    """Write a 60-minute departure-delay workbook from flight_delay_effect.xlsx."""
    simulation_meta = payload.get('simulation') or {}
    if simulation_meta.get('scenario') != 'flight_delay':
        return None
    if int(float(simulation_meta.get('delay_min') or 0)) != 60:
        return None

    affected_terminal = simulation_meta.get('affected_terminal') or simulation_meta.get('selected_terminal')
    scheduled_departure = simulation_meta.get('scheduled_departure')
    if affected_terminal not in ('T1', 'T2') or not scheduled_departure:
        return None

    overlay_rows = _load_flight_delay_overlay_rows(base_dir)
    if not overlay_rows:
        return {'saved': False, 'error': f'{FLIGHT_DELAY_EFFECT_FILE} has no readable overlay rows'}

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return {'saved': False, 'error': 'openpyxl is not installed'}

    base_date = _simulation_export_date(base_dir)
    departure_mins = _mins(scheduled_departure)

    overlay_by_minute = {}
    overlay_windows = {}
    for row in overlay_rows:
        exact_minute = departure_mins + int(row['offset_mins'])
        if exact_minute < 0 or exact_minute >= 1440:
            continue
        interval_start = (exact_minute // 15) * 15
        minute_values = overlay_by_minute.setdefault(interval_start, {})

        for touchpoint, delta in row['overlay'].items():
            if abs(delta) < 1e-9:
                continue
            col_name = _workbook_col_name(affected_terminal, touchpoint)
            minute_values[col_name] = round(delta)
            window = [interval_start, min(1440, interval_start + 15)]
            overlay_windows.setdefault(col_name, [])
            if window not in overlay_windows[col_name]:
                overlay_windows[col_name].append(window)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(headers)

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for interval_start in range(0, 1440, 15):
        ts = datetime.combine(base_date, time(hour=interval_start // 60, minute=interval_start % 60))
        row = [ts]
        overlay_values = overlay_by_minute.get(interval_start) or {}
        for terminal, touchpoint in column_touchpoints:
            col_name = _workbook_col_name(terminal, touchpoint)
            row.append(overlay_values.get(col_name, 0))
        ws.append(row)

    meta = wb.create_sheet('_overlay_meta')
    meta.sheet_state = 'hidden'
    meta.append(['Date', 'Column', 'StartMinute', 'EndMinute'])
    for col_name in sorted(overlay_windows):
        for start, end in overlay_windows[col_name]:
            meta.append([base_date.isoformat(), col_name, int(start), int(end)])

    for cell in ws['A'][1:]:
        cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

    output_path = _data_path(base_dir, 'simulated_intraday_PAX.xlsx')
    wb.save(output_path)
    return {
        'saved': True,
        'filename': 'simulated_intraday_PAX.xlsx',
        'path': output_path,
        'source': FLIGHT_DELAY_EFFECT_FILE,
        'mode': 'flight_delay_overlay_60',
    }


def _simulation_export_date(base_dir):
    path = _data_path(base_dir, 'short term PAX.xlsx')
    if os.path.exists(path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            first_ts = ws.cell(2, 1).value
            if isinstance(first_ts, datetime):
                return first_ts.date()
        except Exception:
            pass
    return datetime.now().date()


def save_intraday_pax_simulation(base_dir, payload):
    """Write simulated intraday PAX demand in the short term PAX workbook shape."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return {'saved': False, 'error': 'openpyxl is not installed'}

    series = payload.get('series') or {}
    labels = series.get('labels') or []
    if not labels:
        return {'saved': False, 'error': 'Simulation payload has no time labels'}

    headers = [
        'Timestamp',
        'T1_Checkin', 'T1_Security', 'T1_CBP', 'T1_Lounge', 'T1_Boarding', 'T1_Immigration', 'T1_Baggage',
        'T2_Checkin', 'T2_Security', 'T2_CBP', 'T2_Lounge', 'T2_Boarding', 'T2_Immigration', 'T2_Baggage',
    ]
    column_touchpoints = [
        ('T1', 'checkin'), ('T1', 'security'), ('T1', 'cbp'), ('T1', 'lounge'), ('T1', 'boarding'), ('T1', 'immigration'), ('T1', 'baggage'),
        ('T2', 'checkin'), ('T2', 'security'), ('T2', 'cbp'), ('T2', 'lounge'), ('T2', 'boarding'), ('T2', 'immigration'), ('T2', 'baggage'),
    ]

    flight_delay_export = _save_flight_delay_overlay_simulation(base_dir, payload, headers, column_touchpoints)
    if flight_delay_export is not None:
        return flight_delay_export

    simulation_meta = payload.get('simulation') or {}
    affected_terminal = simulation_meta.get('affected_terminal')  # e.g. 'T1', 'T2', or None
    affected_touchpoints = set(simulation_meta.get('affected_touchpoints') or [])
    affected_windows = simulation_meta.get('affected_windows') or {}

    flights = _load_flights(base_dir)
    weights = {touchpoint: _terminal_weights(flights, touchpoint) for touchpoint in TOUCHPOINTS}
    aggregated = {
        touchpoint: _aggregate_to_15_minutes(labels, series.get(touchpoint), touchpoint)
        for touchpoint in TOUCHPOINTS
    }

    # Load baseline (unshifted) aggregate so we can isolate changes to affected terminal only
    baseline_aggregated = None
    if affected_terminal in ('T1', 'T2'):
        try:
            pulse = _load_json(base_dir, PULSE_FILE)
            base_labels = pulse.get('labels', [])
            baseline_aggregated = {
                tp: _aggregate_to_15_minutes(base_labels, pulse.get(tp), tp)
                for tp in TOUCHPOINTS
            }
        except Exception:
            baseline_aggregated = None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(headers)

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    base_date = _simulation_export_date(base_dir)
    for interval_start in range(0, 1440, 15):
        ts = datetime.combine(base_date, time(hour=interval_start // 60, minute=interval_start % 60))
        row = [ts]
        for terminal, touchpoint in column_touchpoints:
            w = weights[touchpoint][terminal]
            sim_val = aggregated[touchpoint].get(interval_start, 0)

            if baseline_aggregated and affected_terminal:
                in_affected_window = _interval_overlaps_windows(
                    interval_start,
                    affected_windows.get(touchpoint) or [],
                )
                if terminal == affected_terminal and touchpoint in affected_touchpoints and in_affected_window:
                    base_val = baseline_aggregated[touchpoint].get(interval_start, 0)
                    delta = sim_val - base_val
                    value = max(0, round(base_val * w + delta))
                else:
                    # A terminal-specific delay overlay writes only the affected
                    # terminal/touchpoints. Other terminal columns remain zero in
                    # the workbook and are ignored by overlay metadata at load time.
                    value = 0
            else:
                value = max(0, round(sim_val * w))

            row.append(value)
        ws.append(row)

    if baseline_aggregated and affected_terminal:
        meta = wb.create_sheet('_overlay_meta')
        meta.sheet_state = 'hidden'
        meta.append(['Date', 'Column', 'StartMinute', 'EndMinute'])
        header_by_touchpoint = {
            touchpoint: f"{affected_terminal}_{touchpoint.title() if touchpoint != 'cbp' else 'CBP'}"
            for touchpoint in TOUCHPOINTS
        }
        for touchpoint in sorted(affected_touchpoints):
            col_name = header_by_touchpoint.get(touchpoint)
            if not col_name:
                continue
            for start, end in affected_windows.get(touchpoint) or []:
                meta.append([base_date.isoformat(), col_name, int(start), int(end)])

    for cell in ws['A'][1:]:
        cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

    output_path = _data_path(base_dir, 'simulated_intraday_PAX.xlsx')
    wb.save(output_path)
    return {'saved': True, 'filename': 'simulated_intraday_PAX.xlsx', 'path': output_path}


def _table_from_series(series):
    labels = series.get('labels', [])
    rows = []
    for start in range(30, len(labels), 15):
        row = {'time': labels[start]}
        end = min(start + 15, len(labels))
        for key in TOUCHPOINTS:
            vals = _numbers(series.get(key))[start:end]
            if key == 'lounge':
                row[key] = round(sum(vals) / len(vals)) if vals else 0
            else:
                row[key] = round(sum(vals))
        rows.append(row)
    return rows


def _build_insights(series):
    labels = series.get('labels', [])
    max_checkin = max(series['checkin'] or [0])
    max_security = max(series['security'] or [0])
    max_cbp = max(series['cbp'] or [0])
    max_lounge = max(series['lounge'] or [0])
    max_boarding = max(series['boarding'] or [0])
    max_immigration = max(series['immigration'] or [0])
    max_baggage = max(series['baggage'] or [0])
    avg_security = _active_avg(series['security'])

    return [
        _insight(series, labels, 'checkin', 'Check-In', 'teal', 'Avg (active)', f"{_active_avg(series['checkin']):,} pax"),
        _insight(series, labels, 'security', 'Security', 'purple', 'Pressure Index', f"{(max_security / avg_security):.1f}x avg" if avg_security else '--'),
        _insight(series, labels, 'cbp', 'CBP', 'crit', '% of Security', f"{_pct(max_cbp, max_security) or '--'}% of sec peak"),
        _insight(series, labels, 'lounge', 'Lounge', 'warn', 'Avg Occupancy', f"{_active_avg(series['lounge']):,} pax", 'Peak Concurrent'),
        _insight(series, labels, 'boarding', 'Boarding', 'info', 'Gate Pressure', f"{_pct(max_boarding, max_checkin) or '--'}% of check-in"),
        _insight(series, labels, 'immigration', 'Immigration', 'ok', 'Intl. Arrival Share', f"{_pct(max_immigration, max_baggage) or '--'}% of arrivals"),
    ]


def _build_impact(baseline, simulated):
    impact = {}
    for key in TOUCHPOINTS:
        base_vals = _numbers(baseline.get(key))
        sim_vals = _numbers(simulated.get(key))
        base_peak = max(base_vals) if base_vals else 0
        sim_peak = max(sim_vals) if sim_vals else 0
        delta = sim_peak - base_peak
        delta_pct = round((delta / base_peak) * 100, 1) if base_peak > 0 else 0
        if abs(delta_pct) < 0.5 and base_peak > 0 and base_vals and sim_vals:
            paired = zip(base_vals, sim_vals)
            local_delta = max((s - b for b, s in paired), key=lambda v: abs(v), default=0)
            if abs(local_delta) > 0:
                delta = local_delta
                delta_pct = round((local_delta / base_peak) * 100, 1)
        impact[key] = {
            'baseline_peak': round(base_peak, 1),
            'simulated_peak': round(sim_peak, 1),
            'delta': round(delta, 1),
            'delta_pct': delta_pct,
        }
    return impact


def _base_payload(base_dir):
    pulse = _load_json(base_dir, PULSE_FILE)
    labels = pulse.get('labels', [])
    series = {'labels': labels}
    for key in TOUCHPOINTS:
        series[key] = _numbers(pulse.get(key))
    return pulse, series


def build_intraday_pax_pulse(base_dir):
    pulse, series = _base_payload(base_dir)

    table_rows = []
    table_labels = pulse.get('table_labels', [])
    table_data = pulse.get('table_data', {})
    if table_labels and table_data:
        for idx, label in enumerate(table_labels):
            table_rows.append({
                'time': label,
                **{key: round(_numbers(table_data.get(key))[idx]) if idx < len(_numbers(table_data.get(key))) else 0 for key in TOUCHPOINTS},
            })
    else:
        table_rows = _table_from_series(series)

    max_lounge = max(series['lounge'] or [0])
    return {
        'summary': {
            'title': 'Intra-Day Tactical Pulse',
            'subtitle': 'PAX flow aligned to flights between 7 AM to 8:30 AM based on touchpoint-wise probability distributions',
            'window': f"{series['labels'][0]}-{series['labels'][-1]}" if series.get('labels') else '',
            'lounge_peak': round(max_lounge),
        },
        'series': series,
        'table': {
            'columns': ['time', 'checkin', 'security', 'cbp', 'lounge', 'boarding', 'immigration', 'baggage'],
            'rows': table_rows,
        },
        'flights': _load_flights(base_dir),
        'insights': _build_insights(series),
    }


def _time_index(labels, hhmm):
    try:
        return labels.index(str(hhmm))
    except ValueError:
        return min(range(len(labels)), key=lambda idx: abs(_mins(labels[idx]) - _mins(hhmm))) if labels else 0


def _mins(hhmm):
    h, m = str(hhmm or '00:00').split(':')[:2]
    return int(h) * 60 + int(m)


def _mins_to_hhmm(total_mins):
    total_mins = max(0, int(total_mins))
    return f"{total_mins // 60:02d}:{total_mins % 60:02d}"


def _shift_range(arr, start_idx, end_idx, shift):
    out = arr[:]
    segment = out[start_idx:end_idx]
    for idx in range(start_idx, end_idx):
        out[idx] = 0.0
    for offset, val in enumerate(segment):
        dest = min(len(out) - 1, start_idx + shift + offset)
        out[dest] += val
    return out


def _shift_component(arr, start_idx, end_idx, shift, share):
    out = arr[:]
    if shift <= 0 or share <= 0:
        return out

    start_idx = max(0, min(len(out), int(start_idx)))
    end_idx = max(start_idx, min(len(out), int(end_idx)))
    component = [max(0.0, out[idx] * share) for idx in range(start_idx, end_idx)]

    for offset, val in enumerate(component):
        src = start_idx + offset
        out[src] = max(0.0, out[src] - val)
        dest = min(len(out) - 1, src + shift)
        out[dest] += val
    return out


def _window_minutes(labels, start_idx, end_idx, shift=0):
    if not labels:
        return None
    start_idx = max(0, min(len(labels) - 1, int(start_idx)))
    end_idx = max(start_idx + 1, min(len(labels), int(end_idx)))
    start_min = _mins(labels[start_idx]) + int(shift)
    if end_idx < len(labels):
        end_min = _mins(labels[end_idx]) + int(shift)
    else:
        end_min = _mins(labels[-1]) + 1 + int(shift)
    return [max(0, start_min), min(1440, max(start_min + 1, end_min))]


def _add_shift_windows(target, touchpoint, labels, start_idx, end_idx, shift):
    original = _window_minutes(labels, start_idx, end_idx, 0)
    shifted = _window_minutes(labels, start_idx, end_idx, shift)
    target.setdefault(touchpoint, [])
    for window in (original, shifted):
        if window and window not in target[touchpoint]:
            target[touchpoint].append(window)


def _cap_series(arr, start_idx, end_idx, reduction_pct):
    out = arr[:]
    window = out[start_idx:end_idx + 1]
    nominal = max(window) if window else max(out or [1])
    limit = nominal * max(0.05, 1 - reduction_pct / 100.0)
    queue = 0.0
    for idx, demand in enumerate(out):
        in_window = start_idx <= idx <= end_idx
        if in_window:
            allowed = min(demand + queue, limit)
            queue = (demand + queue) - allowed
            out[idx] = allowed
        elif queue > 0:
            spare = max(0.0, nominal - demand)
            drain = min(queue, spare)
            out[idx] = demand + drain
            queue -= drain
    return out


def simulate_intraday_pax(base_dir, scenario, params):
    base = build_intraday_pax_pulse(base_dir)
    baseline_series = deepcopy(base['series'])
    series = deepcopy(base['series'])
    labels = series.get('labels', [])
    cascade = []
    affected_terminal = None
    affected_touchpoints = set()
    affected_windows = {}
    selected_flight_meta = {}

    if scenario == 'flight_delay':
        delay = int(float(params.get('delay_min') or 30))
        shift = max(0, delay)
        flight_no = params.get('flight_no') or 'selected flight'

        flights = _load_flights(base_dir)
        selected = next((f for f in flights if f['flight_no'] == flight_no), None)
        if selected:
            selected_flight_meta = {
                'selected_flight_no': selected.get('flight_no'),
                'selected_terminal': _terminal_for_flight(selected),
                'scheduled_departure': selected.get('etd'),
                'scheduled_arrival': selected.get('eta'),
            }

        if selected and selected.get('etd'):
            etd = _mins(str(selected['etd']))
            eta = _mins(str(selected.get('eta') or '0')) if selected.get('eta') else None
            cbp_req = str(selected.get('cbp_required') or '').strip().lower() == 'true'
            affected_terminal = _terminal_for_flight(selected)

            # Upstream departure windows: delay moves the selected flight's
            # check-in/security demand later alongside the departure.
            checkin_start_idx = _time_index(labels, _mins_to_hhmm(max(0, etd - 180)))
            security_start_idx = _time_index(labels, _mins_to_hhmm(max(0, etd - 165)))
            pre_board_end_idx = _time_index(labels, _mins_to_hhmm(max(0, etd - 30)))

            # Boarding window: 90 min lead up to departure
            board_start_idx = _time_index(labels, _mins_to_hhmm(max(0, etd - 90)))
            board_end_idx = _time_index(labels, _mins_to_hhmm(etd))

            # Lounge/CBP window: 180 min lead up to departure
            lc_start_idx = _time_index(labels, _mins_to_hhmm(max(0, etd - 180)))

            # Move only the selected flight's estimated passenger contribution.
            # The same component deducted from the original window is added to
            # the delayed window, preserving passenger volume by touchpoint.
            checkin_share = _flight_touchpoint_share(flights, selected, 'checkin')
            security_share = _flight_touchpoint_share(flights, selected, 'security')
            boarding_share = _flight_touchpoint_share(flights, selected, 'boarding')
            lounge_share = _flight_touchpoint_share(flights, selected, 'lounge')
            series['checkin'] = _shift_component(series['checkin'], checkin_start_idx, pre_board_end_idx, shift, checkin_share)
            series['security'] = _shift_component(series['security'], security_start_idx, board_start_idx, shift, security_share)
            series['boarding'] = _shift_component(series['boarding'], board_start_idx, board_end_idx, shift, boarding_share)
            series['lounge'] = _shift_component(series['lounge'], lc_start_idx, board_end_idx, shift, lounge_share)
            if checkin_share > 0:
                affected_touchpoints.add('checkin')
                _add_shift_windows(affected_windows, 'checkin', labels, checkin_start_idx, pre_board_end_idx, shift)
            if security_share > 0:
                affected_touchpoints.add('security')
                _add_shift_windows(affected_windows, 'security', labels, security_start_idx, board_start_idx, shift)
            if boarding_share > 0:
                affected_touchpoints.add('boarding')
                _add_shift_windows(affected_windows, 'boarding', labels, board_start_idx, board_end_idx, shift)
            if lounge_share > 0:
                affected_touchpoints.add('lounge')
                _add_shift_windows(affected_windows, 'lounge', labels, lc_start_idx, board_end_idx, shift)
            if cbp_req:
                cbp_share = _flight_touchpoint_share(flights, selected, 'cbp')
                series['cbp'] = _shift_component(series['cbp'], lc_start_idx, board_end_idx, shift, cbp_share)
                if cbp_share > 0:
                    affected_touchpoints.add('cbp')
                    _add_shift_windows(affected_windows, 'cbp', labels, lc_start_idx, board_end_idx, shift)
            if eta is not None and _flight_serves_touchpoint(selected, 'immigration'):
                arr_start_idx = _time_index(labels, _mins_to_hhmm(eta))
                arr_end_idx = _time_index(labels, _mins_to_hhmm(min(1439, eta + 90)))
                immigration_share = _flight_touchpoint_share(flights, selected, 'immigration')
                baggage_share = _flight_touchpoint_share(flights, selected, 'baggage')
                series['immigration'] = _shift_component(series['immigration'], arr_start_idx, arr_end_idx, shift, immigration_share)
                series['baggage'] = _shift_component(series['baggage'], arr_start_idx, arr_end_idx, shift, baggage_share)
                if immigration_share > 0:
                    affected_touchpoints.add('immigration')
                    _add_shift_windows(affected_windows, 'immigration', labels, arr_start_idx, arr_end_idx, shift)
                if baggage_share > 0:
                    affected_touchpoints.add('baggage')
                    _add_shift_windows(affected_windows, 'baggage', labels, arr_start_idx, arr_end_idx, shift)
        else:
            for key in ('boarding', 'lounge', 'cbp'):
                series[key] = _shift_range(series[key], 0, len(labels), shift)
                affected_touchpoints.add(key)
                affected_windows[key] = [[0, 1440]]

        cascade = [
            f"{flight_no} delayed by {delay} minutes; boarding and lounge demand shift later.",
            "Gate pressure moves into the next departure wave.",
            "CBP demand is shifted where the selected flight requires preclearance.",
        ]

    elif scenario == 'ground_stop':
        start = params.get('start') or '07:30'
        duration = int(float(params.get('duration_min') or 45))
        start_idx = _time_index(labels, start)
        end_idx = min(len(labels) - 1, start_idx + duration)
        for key in TOUCHPOINTS:
            series[key] = _shift_range(series[key], start_idx, end_idx, duration)
        cascade = [
            f"Ground stop from {start} for {duration} minutes compresses affected demand into the recovery window.",
            "Arrivals and departures bunch after release, raising downstream queue risk.",
            "Lounge occupancy stays elevated while passengers wait airside.",
        ]

    elif scenario == 'security_reduction':
        reduction = float(params.get('reduction_pct') or 40)
        start = params.get('start') or '06:00'
        end = params.get('end') or '08:30'
        start_idx = _time_index(labels, start)
        end_idx = _time_index(labels, end)
        series['security'] = _cap_series(series['security'], start_idx, end_idx, reduction)
        series['boarding'] = _shift_range(series['boarding'], start_idx, end_idx, max(5, round(reduction / 2)))
        cascade = [
            f"Security throughput reduced by {reduction:.0f}% between {start} and {end}.",
            "Security queue pressure is capped during the restriction and drains afterward.",
            "Boarding demand shifts later as passengers clear security later.",
        ]

    elif scenario == 'checkin_reduction':
        reduction = float(params.get('reduction_pct') or 40)
        start = params.get('start') or '06:00'
        end = params.get('end') or '08:00'
        start_idx = _time_index(labels, start)
        end_idx = _time_index(labels, end)
        series['checkin'] = _cap_series(series['checkin'], start_idx, end_idx, reduction)
        series['security'] = _shift_range(series['security'], start_idx, end_idx, max(5, round(reduction / 2)))
        cascade = [
            f"Check-in desk capacity reduced by {reduction:.0f}% between {start} and {end}.",
            "Check-in backlog pushes later demand into security.",
            "Downstream resource pressure increases after the restriction window.",
        ]

    else:
        raise ValueError(f"Unsupported scenario: {scenario}")

    sim_meta = {
        'active': True,
        'scenario': scenario,
        'cascade': cascade,
    }
    if scenario == 'flight_delay':
        sim_meta['delay_min'] = delay
        sim_meta.update({k: v for k, v in selected_flight_meta.items() if v})
    if scenario == 'flight_delay' and affected_terminal:
        sim_meta['affected_terminal'] = affected_terminal
        sim_meta['affected_touchpoints'] = sorted(affected_touchpoints)
        sim_meta['affected_windows'] = affected_windows

    return {
        **base,
        'series': series,
        'labels': series.get('labels', []),
        'baseline': {key: baseline_series.get(key, []) for key in TOUCHPOINTS},
        'simulated': {key: series.get(key, []) for key in TOUCHPOINTS},
        'impact': _build_impact(baseline_series, series),
        'table': {
            'columns': base['table']['columns'],
            'rows': _table_from_series(series),
        },
        'insights': _build_insights(series),
        'simulation': sim_meta,
    }
