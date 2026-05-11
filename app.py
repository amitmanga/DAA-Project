from flask import Flask, jsonify, request, render_template
import csv
import json
import math
import os
import re
from datetime import datetime, timedelta
from collections import defaultdict

# CP-SAT optimiser (optional — falls back to greedy if OR-Tools is absent)
try:
    from intraday_optimizer import (
        optimize_intraday_assignments as _cpsat_optimize,
        ORTOOLS_AVAILABLE as _CPSAT_AVAILABLE,
    )
except ImportError:  # pragma: no cover
    _CPSAT_AVAILABLE = False
    def _cpsat_optimize(*_a, **_kw):  # type: ignore[misc]
        return None

# Long-term MIP optimiser (optional — requires PuLP or scipy)
try:
    from long_term_mip import (
        optimize_weekly_staffing as _lt_mip_optimize,
        MIP_AVAILABLE as _LT_MIP_AVAILABLE,
    )
except ImportError:  # pragma: no cover
    _LT_MIP_AVAILABLE = False
    def _lt_mip_optimize(*_a, **_kw):  # type: ignore[misc]
        raise RuntimeError("long_term_mip module not found")

# Roster optimiser — two-phase shift-pattern generation + staff assignment
try:
    from roster_optimizer import (
        generate_roster       as _roster_generate,
        tasks_to_demand_windows as _roster_tasks_to_dw,
        format_as_on_duty     as _roster_fmt_on_duty,
        DemandWindow          as _RosterDemandWindow,
        SOLVER_AVAILABLE      as _ROSTER_SOLVER_AVAILABLE,
    )
    _ROSTER_AVAILABLE = True
except ImportError:  # pragma: no cover
    _ROSTER_AVAILABLE = False
    _ROSTER_SOLVER_AVAILABLE = False
    def _roster_generate(*_a, **_kw):  # type: ignore[misc]
        raise RuntimeError("roster_optimizer module not found")
    def _roster_tasks_to_dw(*_a, **_kw):  # type: ignore[misc]
        return []
    def _roster_fmt_on_duty(*_a, **_kw):  # type: ignore[misc]
        return []

# Simulation engine — Monte Carlo intraday stress-testing (validation layer only)
try:
    from simulation_engine import run_simulation as _sim_run_simulation
    _SIM_AVAILABLE = True
except ImportError:  # pragma: no cover
    _SIM_AVAILABLE = False
    def _sim_run_simulation(*_a, **_kw):  # type: ignore[misc]
        raise RuntimeError("simulation_engine module not found")

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TASK_SLOT_MINS = 30
PAX_SLOT_MINS = 15
PAX_DEMAND_SLOT_MINS = 60


@app.after_request
def add_api_no_cache_headers(response):
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

def read_csv(filename):
    path = os.path.join(BASE_DIR, 'data', filename)
    with open(path, encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
    return [r for r in rows if any(v.strip() for v in r.values())]


def clean_employee_id(value):
    """Normalise employee ids that may contain hidden control/NUL characters."""
    return re.sub(r'[\x00-\x1f\x7f]', '', str(value or '')).strip()

def parse_date(s, fmt='%d-%m-%y'):
    s = s.strip()
    for f in ('%d-%m-%y', '%d-%m-%Y', '%d-%b-%y', '%d-%b-%Y'):
        try:
            return datetime.strptime(s, f)
        except ValueError:
            pass
    return None

# ---------------------------------------------------------------------------
# Skill Normalization: Maps Staff_schedule names to Config.csv names
# ---------------------------------------------------------------------------
SKILL_MAP = {
    'GNIB':                 'GNIB / Immigration',
    'CBP Pre-clearance':    'CBP Pre-clearance',
    'Bussing':              'Bussing',
    'PBZ':                  'PBZ',
    'Mezz Operation':       'Mezz Operation',
    'Litter Picking':       'Litter Picking',
    'Ramp / Marshalling':   'Ramp / Marshalling',
    'Arr Customer Service': 'Arr Customer Service',
    'Check-in/Trolleys':    'Check-in/Trolleys',
    'Dep/Trolleys':         'Dep / Trolleys',
    'Dep / Trolleys':       'Dep / Trolleys',
    'T1/T2 Trolleys L/UL':  'Dep / Trolleys', # Proxy for T1 zone work
    'Transfer Corridor':   'Transfer Corridor'
}

PAX_WORK_SKILL_MAP = {
    'checkin':    'Checkin',
    'security':   'Security',
    'cbp':        'CBP',
    'lounge':     'Lounge',
    'boarding':   'Boarding',
    'immigration':'Immigration',
    'baggage':    'Baggage',
}

def normalize_skill(sk):
    sk = sk.strip()
    low = sk.lower()
    return PAX_WORK_SKILL_MAP.get(low, SKILL_MAP.get(sk, sk))

def _leave_days_for_month(row, month_start, month_end):
    """Return this row's leave-day contribution within one calendar month."""
    d_from = parse_date(row.get('DATE FROM', ''))
    d_to = parse_date(row.get('DATE TO', ''))
    if not d_from or not d_to:
        return 0
    if d_to < d_from:
        d_from, d_to = d_to, d_from

    overlap_start = max(d_from, month_start)
    overlap_end = min(d_to, month_end)
    if overlap_start > overlap_end:
        return 0

    total_calendar_days = (d_to - d_from).days + 1
    overlap_calendar_days = (overlap_end - overlap_start).days + 1

    try:
        duration_days = float(row.get('DURATION DAYS', '') or 0)
    except (TypeError, ValueError):
        duration_days = 0

    if duration_days > 0 and total_calendar_days > 0:
        return duration_days * (overlap_calendar_days / total_calendar_days)
    return overlap_calendar_days

def get_monthly_absence_impact():
    """
    Calculate leave days per month, grouped by LEAVE TYPE.

    DATE FROM/DATE TO define the calendar span, including Monday-Sunday. The
    CSV's DURATION DAYS is the actual leave quantity, so cross-month records are
    apportioned by calendar-day overlap and category totals are returned as
    whole-day counts for charting.
    """
    absences = read_csv('Staff_absence_schedule.csv')
    monthly_absent = defaultdict(lambda: defaultdict(float))
    leave_types_set = set()
    # Assume 2026 based on data, or dynamic based on now
    current_year = 2026
    
    for month in range(1, 13):
        month_start = datetime(current_year, month, 1)
        if month == 12:
            month_end = datetime(current_year, 12, 31)
        else:
            month_end = datetime(current_year, month + 1, 1) - timedelta(days=1)
        month_key = month_start.strftime('%b %Y')

        for a in absences:
            leave_days = _leave_days_for_month(a, month_start, month_end)
            if leave_days <= 0:
                continue

            leave_type = a.get('LEAVE TYPE', 'Unknown').strip()
            if not leave_type:
                leave_type = 'Unknown'

            monthly_absent[month_key][leave_type] += leave_days
            leave_types_set.add(leave_type)
                
    # Chart consumers expect simple whole-day counts, not decimal FTE-like values.
    int_absent = defaultdict(dict)
    for month, lt_dict in monthly_absent.items():
        for lt, val in lt_dict.items():
            int_absent[month][lt] = int(math.ceil(val))
            
    return dict(int_absent), sorted(list(leave_types_set))


def get_weekly_absence_impact():
    """Calculate leave days per ISO week of 2026, grouped by LEAVE TYPE."""
    absences = read_csv('Staff_absence_schedule.csv')
    weekly_absent = defaultdict(lambda: defaultdict(float))
    leave_types_set = set()

    # ISO week 1 of 2026 starts Mon 29 Dec 2025 (Jan 4 is always in W01)
    jan4 = datetime(2026, 1, 4)
    week1_monday = jan4 - timedelta(days=jan4.weekday())

    week_ranges = []
    for w in range(53):
        wk_start = week1_monday + timedelta(weeks=w)
        wk_end = wk_start + timedelta(days=6)
        if wk_start.year > 2026:
            break
        week_key = wk_start.strftime('%G-W%V')
        label = str(wk_start.day) + ' ' + wk_start.strftime('%b')
        week_ranges.append((week_key, wk_start, wk_end, label))

    for week_key, wk_start, wk_end, _label in week_ranges:
        for a in absences:
            leave_days = _leave_days_for_month(a, wk_start, wk_end)
            if leave_days <= 0:
                continue
            leave_type = a.get('LEAVE TYPE', 'Unknown').strip() or 'Unknown'
            weekly_absent[week_key][leave_type] += leave_days
            leave_types_set.add(leave_type)

    int_absent = {wk: {lt: int(math.ceil(v)) for lt, v in lt_dict.items()}
                  for wk, lt_dict in weekly_absent.items()}
    week_labels = [[wk, label] for wk, _s, _e, label in week_ranges]
    return int_absent, sorted(leave_types_set), week_labels


# ---------------------------------------------------------------------------
# Per-movement staff-minutes derived directly from Config.csv task rules.
# Key: (Flight_Category, Status)
#
# DEPARTURES apply: GNIB / Immigration + Ramp / Marshalling + Bussing (22% remote stands)
# ARRIVALS apply:   Arr Customer Service + Check-in/Trolleys + Bussing (22% remote stands)
# CBP hall is session-based (see FIXED_FTE below); per-movement only covers GNIB + Ramp.
#
# Derivations:
#   Short-Haul DEP : GNIB(2×60=120) + Ramp(1×30=30) + Bussing(0.22×1×30≈7)  = 157
#   Short-Haul ARR : ArrCS(1×55=55) + Trolleys(1×45=45) + Bussing(0.22×1×35≈8) = 108
#   Long-Haul  DEP : GNIB(3×80=240) + Ramp(2×50=100) + Bussing(0.10×1×30≈3)  = 343
#   Long-Haul  ARR : ArrCS(1×90=90) + Bussing(0.10×1×35≈4)                   =  94
#   CBP        DEP : GNIB(3×80=240) + Ramp(2×50=100) + Bussing(0.05×1×30≈2)  = 342
#   CBP        ARR : ArrCS(1×90=90) + Bussing(0.05×1×35≈2)                   =  92
#   Domestic   DEP : Ramp(1×30=30) + Bussing(0.20×1×30≈6)                    =  36
#   Domestic   ARR : ArrCS(1×45=45) + Bussing(0.20×1×35≈7)                   =  52
#   Cargo      DEP : Ramp(2×50=100)                                            = 100
#   Cargo      ARR : Ramp(2×40=80)                                             =  80
# ---------------------------------------------------------------------------
STAFF_MINS_PER_MOVEMENT = {
    ('International Short-Haul', 'Departure'): 157,
    ('International Short-Haul', 'Arrival'):   108,
    ('International Long-Haul',  'Departure'): 343,
    ('International Long-Haul',  'Arrival'):    94,
    ('Transatlantic CBP',        'Departure'): 342,
    ('Transatlantic CBP',        'Arrival'):    92,
    ('Domestic',                 'Departure'):  36,
    ('Domestic',                 'Arrival'):    52,
    ('Cargo',                    'Departure'): 100,
    ('Cargo',                    'Arrival'):    80,
}

# Skill share of each (category, status) task bundle — must sum to 1.0
SKILL_SPLIT = {
    ('International Short-Haul', 'Departure'): {'GNIB': 0.76, 'Ramp / Marshalling': 0.19, 'Bussing': 0.05},
    ('International Short-Haul', 'Arrival'):   {'Arr Customer Service': 0.51, 'Check-in/Trolleys': 0.42, 'Bussing': 0.07},
    ('International Long-Haul',  'Departure'): {'GNIB': 0.70, 'Ramp / Marshalling': 0.29, 'Bussing': 0.01},
    ('International Long-Haul',  'Arrival'):   {'Arr Customer Service': 0.96, 'Bussing': 0.04},
    ('Transatlantic CBP',        'Departure'): {'GNIB': 0.70, 'Ramp / Marshalling': 0.29, 'Bussing': 0.01},
    ('Transatlantic CBP',        'Arrival'):   {'Arr Customer Service': 0.96, 'Bussing': 0.04},
    ('Domestic',                 'Departure'): {'Ramp / Marshalling': 0.83, 'Bussing': 0.17},
    ('Domestic',                 'Arrival'):   {'Arr Customer Service': 0.87, 'Bussing': 0.13},
    ('Cargo',                    'Departure'): {'Ramp / Marshalling': 1.0},
    ('Cargo',                    'Arrival'):   {'Ramp / Marshalling': 1.0},
}

# Fixed posts from Config.csv — run regardless of flight volume.
# FTE = (staff_count × duration_mins × operating_days) / NET_WORKING_MINS_PER_WEEK
# NET_WORKING_MINS_PER_WEEK = 630 min/shift × 5 days = 3150
#
#  Mezz Operation  : 2 staff × 300 min × 5 days / 3150 = 0.95 FTE
#  Litter Picking  : 2 hrs at day shift, 2 hrs at night shift by 2 FTE = 480 mins/day × 7 days / 3150 = 1.07 FTE
#  CBP Hall        : 3 staff × 300 min/session × 5 days / 3150 = 1.43 FTE
#  PBZ (T2 pier)   : 4 roster slots (session-based, consistent) = ~1.27 FTE
FIXED_FTE = {
    'Mezz Operation':     3.00,
    'Litter Picking':     1.07,
    'CBP Pre-clearance':  1.43,   # hall session staffing on top of per-movement GNIB/Ramp
    'PBZ':                1.27,
}
FIXED_FTE_TOTAL = sum(FIXED_FTE.values())  # = 6.58 FTE

# Calibration constants
# Historical 2025 average weekly flight-driven staff-minutes (W24 + S25 + W25 data).
# Calculated from actual 2025 movements × STAFF_MINS_PER_MOVEMENT above.
# Represents the demand that the current 50-person workforce is calibrated against.
HISTORICAL_AVG_STAFF_MINS = 763_301.0
BASELINE_STAFF = 60
NET_WORKING_MINS_PER_WEEK = 630 * 5  # 3150 mins per FTE per week
PAX_SLOTS_PER_FTE_WEEK = NET_WORKING_MINS_PER_WEEK / PAX_SLOT_MINS

AIRCRAFT_PAX_CAPACITY = {
    'E190': 100,
    'A319': 140,
    'A320': 180,
    'A321': 220,
    'B737': 189,
    'B738': 189,
    'B38M': 197,
    'B752': 220,
    'B767': 260,
    'A330': 300,
    'A332': 260,
    'A333': 300,
    'A350': 325,
    'B787': 290,
    'B789': 290,
    'B777': 365,
}

FLIGHT_CATEGORY_PAX_CAPACITY = {
    'Domestic': 160,
    'International Short-Haul': 180,
    'International Long-Haul': 310,
    'Transatlantic CBP': 310,
    'Cargo': 0,
}

PAX_CONFIG_CACHE = None
PAX_PROFILE_CACHE = None


def _pax_work_from_col(col_name):
    parts = str(col_name or '').strip().split('_', 1)
    if len(parts) != 2:
        return ''
    return parts[1].strip().lower()


def load_pax_config():
    """Return {pax_column: passengers handled by one FTE in one 15-min slot}."""
    global PAX_CONFIG_CACHE
    if PAX_CONFIG_CACHE is not None:
        return PAX_CONFIG_CACHE

    path = os.path.join(BASE_DIR, 'data', 'PAX Config.xlsx')
    if not os.path.exists(path):
        PAX_CONFIG_CACHE = {}
        return PAX_CONFIG_CACHE

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        values = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), [])
        cfg = {}
        for h, v in zip(headers, values):
            if not h:
                continue
            try:
                rate = float(v or 0)
            except (TypeError, ValueError):
                rate = 0.0
            if rate > 0:
                cfg[h] = rate
        PAX_CONFIG_CACHE = cfg
    except Exception:
        PAX_CONFIG_CACHE = {}
    return PAX_CONFIG_CACHE


def load_pax_profile():
    """Return passenger profile rows keyed by their source workbook date."""
    global PAX_PROFILE_CACHE
    if PAX_PROFILE_CACHE is not None:
        return PAX_PROFILE_CACHE

    path = os.path.join(BASE_DIR, 'data', 'short term PAX.xlsx')
    if not os.path.exists(path):
        PAX_PROFILE_CACHE = {'dates': [], 'rows_by_date': {}}
        return PAX_PROFILE_CACHE

    rows_by_date = defaultdict(list)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        for vals in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, vals))
            ts = row.get('Timestamp')
            if not isinstance(ts, datetime):
                continue
            rows_by_date[ts.date()].append(row)
    except Exception:
        rows_by_date = defaultdict(list)

    dates = sorted(rows_by_date.keys())
    PAX_PROFILE_CACHE = {'dates': dates, 'rows_by_date': dict(rows_by_date)}
    return PAX_PROFILE_CACHE


def _parse_any_date(date_str):
    for fmt in ('%Y-%m-%d', '%d-%b-%y', '%d-%b-%Y', '%d-%m-%Y', '%d-%m-%y'):
        try:
            return datetime.strptime(str(date_str).strip(), fmt)
        except ValueError:
            pass
    return None


def _pax_source_date_for(requested_dt):
    """Map current D..D+3 dates to the four dates provided in the PAX workbook."""
    profile = load_pax_profile()
    dates = profile.get('dates', [])
    if not dates:
        return None

    req_date = requested_dt.date()
    if req_date in dates:
        return req_date

    today = datetime.now().date()
    offset = (req_date - today).days
    if 0 <= offset < len(dates):
        return dates[offset]
    return None


def build_pax_demand_tasks(date_str, current_time_mins=None):
    """Build hourly passenger-handling demand windows from the PAX workbooks."""
    requested_dt = _parse_any_date(date_str)
    if requested_dt is None:
        return []

    source_date = _pax_source_date_for(requested_dt)
    if source_date is None:
        return []

    cfg = load_pax_config()
    profile = load_pax_profile()
    rows = profile.get('rows_by_date', {}).get(source_date, [])
    if not cfg or not rows:
        return []

    hourly = defaultdict(lambda: {'passengers': 0.0, 'rate': 0.0, 'col': '', 'source_count': 0})
    for row in sorted(rows, key=lambda r: r.get('Timestamp')):
        ts = row.get('Timestamp')
        if not isinstance(ts, datetime):
            continue
        start_mins = ts.hour * 60
        end_mins = min(1440, start_mins + PAX_DEMAND_SLOT_MINS)

        for col, rate in cfg.items():
            try:
                pax = float(row.get(col) or 0)
            except (TypeError, ValueError):
                pax = 0.0
            if pax <= 0 or rate <= 0:
                continue

            bucket = hourly[(start_mins, col)]
            bucket['passengers'] += pax
            bucket['rate'] = float(rate)
            bucket['col'] = col
            bucket['source_count'] += 1

    tasks = []
    rate_multiplier = PAX_DEMAND_SLOT_MINS / PAX_SLOT_MINS
    for (start_mins, col), bucket in sorted(hourly.items()):
        pax = bucket['passengers']
        rate = bucket['rate']
        if pax <= 0 or rate <= 0:
            continue

        end_mins = min(1440, start_mins + PAX_DEMAND_SLOT_MINS)
        hourly_rate = rate * rate_multiplier
        work = _pax_work_from_col(col)
        skill = PAX_WORK_SKILL_MAP.get(work, work)
        terminal = col.split('_', 1)[0] if '_' in col else 'ALL'
        needed = max(1, int(math.ceil(pax / hourly_rate)))
        cap = max(needed, int(math.ceil(needed * 1.5)))
        task = {
            'id':              f"PAX_{col}_{start_mins}",
            'flight_no':       'PAX',
            'task':            f"{terminal} {work.title()} PAX",
            'role':            skill,
            'skill':           skill,
            'priority':        'Critical' if work in ('security', 'immigration', 'cbp') else 'High',
            'start_mins':      start_mins,
            'end_mins':        end_mins,
            'start':           mins_to_time(start_mins),
            'end':             mins_to_time(end_mins),
            'staff_needed':    needed,
            'staff_capacity':  cap,
            'assigned':        [],
            'alert':           None,
            'time_mins':       start_mins,
            'flights_covered': [],
            'terminal':        terminal,
            'pier':            'ALL',
            'sharing_mode':    'pax_hourly',
            'passengers':      int(round(pax)),
            'pax_rate':        hourly_rate,
            'pax_rate_15m':    rate,
            'slot_mins':       PAX_DEMAND_SLOT_MINS,
            'source_slot_mins': PAX_SLOT_MINS,
            'source_slots':    bucket['source_count'],
            'time_window':     f"{mins_to_time(start_mins)}-{mins_to_time(end_mins)}",
            'source_date':     source_date.isoformat(),
        }
        if current_time_mins is not None and end_mins <= current_time_mins:
            task['is_past'] = True
        tasks.append(task)

    return tasks


def _load_factor(row):
    raw = str(row.get('Avg_Load_Factor_Pct', '') or '').strip().replace('%', '')
    try:
        val = float(raw)
    except (TypeError, ValueError):
        val = 85.0
    if val > 1:
        val /= 100.0
    return max(0.0, min(1.0, val))


def _aircraft_capacity(row):
    """Resolve a passenger capacity for a flight row.

    This is intentionally tolerant of varying CSV column names that may
    contain aircraft family/type (e.g. 'Aircraft_Family', 'aircraft_type',
    'Aircraft_Type', 'Aircraft'). If a specific type is not recognised, the
    function falls back to the broad `FLIGHT_CATEGORY_PAX_CAPACITY` mapping.
    """
    fam = str(
        (row.get('Aircraft_Family') or row.get('aircraft_type') or
         row.get('Aircraft_Type') or row.get('Aircraft Type') or
         row.get('Aircraft') or row.get('AircraftFamily') or '')
    ).strip().upper()

    if fam:
        if fam in AIRCRAFT_PAX_CAPACITY:
            return AIRCRAFT_PAX_CAPACITY[fam]
        for key, cap in AIRCRAFT_PAX_CAPACITY.items():
            if fam.startswith(key):
                return cap

    # Fallback to category-level capacity when aircraft family unknown
    return FLIGHT_CATEGORY_PAX_CAPACITY.get(str(row.get('Flight_Category', '')).strip(), 180)


def _long_term_pax_work_mix(category, status, terminal):
    terminal = terminal if terminal in ('T1', 'T2') else 'T1'
    if status == 'Departure':
        work = ['Checkin', 'Security', 'Lounge', 'Boarding']
        if category == 'Transatlantic CBP':
            work.append('CBP')
    else:
        work = ['Immigration', 'Baggage']
    return [f'{terminal}_{w}' for w in work]


def compute_week_start(d):
    """Return Monday of the week containing date d."""
    return d - timedelta(days=d.weekday())


def load_weekly_demand():
    rows = read_csv('Weekly_flight_demand.csv')
    return [r for r in rows if r.get('Season_Code') and r.get('Flight_Category', '').strip().lower() != 'cargo']


def load_staff():
    rows = read_csv('Staff_schedule.csv')
    # deduplicate: one record per employee (use first date seen)
    seen = {}
    for r in rows:
        emp = clean_employee_id(r.get('EMPLOYEE NUMBER', ''))
        if emp and emp not in seen:
            r = dict(r)
            r['EMPLOYEE NUMBER'] = emp
            seen[emp] = r
    return list(seen.values())


def load_absences():
    rows = read_csv('Staff_absence_schedule.csv')
    cleaned = []
    for r in rows:
        emp = clean_employee_id(r.get('EMPLOYEE NUMBER', ''))
        if not emp:
            continue
        r = dict(r)
        r['EMPLOYEE NUMBER'] = emp
        cleaned.append(r)
    return cleaned


def weekly_demand_2026():
    """Returns weekly movement groups with enough metadata to estimate passengers.

    Uses S26/W26 Forecast for weeks W15-W53.
    Uses W25 Historical for weeks W01-W14 (Jan-Mar 2026, winter season).
    """
    rows = load_weekly_demand()
    by_week = defaultdict(lambda: defaultdict(float))

    for r in rows:
        sc    = r.get('Season_Code', '').strip()
        dtype = r.get('Data_type', '').strip()
        d     = parse_date(r.get('Week_Start', '').strip())
        if not d or d.year != 2026:
            continue

        include = sc in ('S26', 'W26') or \
                  (dtype == 'Historical' and sc == 'W25')
        if not include:
            continue

        week_key = d.strftime('%Y-W%V')

        # Be permissive with input column names for aircraft and load-factor.
        flight_category = r.get('Flight_Category', '').strip()
        status = r.get('Status', '').strip()
        terminal = r.get('Terminal', '').strip()
        aircraft = (r.get('Aircraft_Family') or r.get('aircraft_type') or r.get('Aircraft_Type') or r.get('Aircraft') or r.get('AircraftFamily') or '').strip()
        load_factor = (r.get('Avg_Load_Factor_Pct') or r.get('Avg_Load_Factor') or r.get('Load_Factor_Pct') or r.get('Load_Factor') or '').strip()

        key = (flight_category, status, terminal, aircraft, load_factor)
        try:
            mvmt = float(r.get('Weekly_Movements', 0) or 0)
        except:
            mvmt = 0
        by_week[week_key][key] += mvmt

    return dict(by_week)


def weekly_staff_required(demand_by_week):
    """Returns {week_key: fte_required} and {week_key: {skill: fte}}.

    Methodology:
    1. Passenger estimate per row = Weekly_Movements × aircraft_capacity × load_factor
    2. All passengers flow through every applicable stage (no even-split across stages):
         DEP: Checkin, Security, Lounge, Boarding (+ CBP for Transatlantic)
         ARR: Immigration, Baggage
    3. FTE per skill = (weekly_pax / PAX_Config_rate) / PAX_SLOTS_PER_FTE_WEEK
       where PAX_Config_rate is passengers handled per FTE per 15-min slot.
    """
    total = {}
    by_skill = {}
    pax_cfg = load_pax_config()
    fallback_rate = 100.0

    for wk, grouped in demand_by_week.items():
        skill_slots = defaultdict(float)
        total_slots = 0.0

        for key, mvmt in grouped.items():
            if len(key) == 5:
                cat, status, terminal, aircraft, load_factor = key
            else:
                cat, status = key[:2]
                terminal, aircraft, load_factor = 'T1', '', '85%'

            row = {
                'Flight_Category': cat,
                'Aircraft_Family': aircraft,
                'Avg_Load_Factor_Pct': load_factor,
            }
            pax = float(mvmt or 0) * _aircraft_capacity(row) * _load_factor(row)
            work_cols = _long_term_pax_work_mix(cat, status, terminal)
            if pax <= 0 or not work_cols:
                continue

            for col in work_cols:
                rate = pax_cfg.get(col, fallback_rate)
                if rate <= 0:
                    continue
                slots = pax / rate
                skill = PAX_WORK_SKILL_MAP.get(_pax_work_from_col(col), _pax_work_from_col(col))
                skill_slots[skill] += slots
                total_slots += slots

        skill_fte = defaultdict(float)
        for sk, slots in skill_slots.items():
            skill_fte[sk] = slots / PAX_SLOTS_PER_FTE_WEEK
        total_fte = total_slots / PAX_SLOTS_PER_FTE_WEEK

        total[wk] = round(total_fte, 1)
        by_skill[wk] = {k: round(v, 1) for k, v in skill_fte.items()}

    return total, by_skill


def weekly_staff_available():
    """Returns {week_key: net_staff} and {week_key: {skill: count}}."""
    staff = load_staff()
    absences = load_absences()
    total_staff = len(staff)

    # Build employee -> skills map (normalized) and a skill pool.
    # Employees who have only `Mezz Operation` are treated as mezz-only
    # and counted exclusively against the Mezz Operation pool.
    emp_skills = {}
    skill_pool = defaultdict(float)
    for s in staff:
        emp = s.get('EMPLOYEE NUMBER', '').strip()
        skills = set()
        for sk_col in ['Skill1', 'Skill2', 'Skill3', 'Skill4']:
            sk_name = s.get(sk_col, '').strip()
            if sk_name:
                skills.add(normalize_skill(sk_name))
        emp_skills[emp] = skills
        # If the employee only holds Mezz Operation, count only towards Mezz
        if skills == {'Mezz Operation'}:
            skill_pool['Mezz Operation'] += 1.0
        elif len(skills) > 0:
            weight = 1.0 / len(skills)
            for sk in skills:
                skill_pool[sk] += weight

    # Build absence windows: {employee: [(from_date, to_date)]}
    absence_map = defaultdict(list)
    for a in absences:
        emp = a['EMPLOYEE NUMBER'].strip()
        d_from = parse_date(a.get('DATE FROM', ''))
        d_to = parse_date(a.get('DATE TO', ''))
        if d_from and d_to:
            absence_map[emp].append((d_from, d_to))

    # For each ISO week in 2026, count absent staff
    result = {}
    skill_result = {}

    # Generate all Monday dates for 2026
    start = datetime(2026, 1, 5)  # first Monday of 2026
    end = datetime(2026, 12, 28)
    d = start
    while d <= end:
        wk_key = d.strftime('%Y-W%V')
        week_end = d + timedelta(days=4)  # Friday
        absent_emps = set()
        for emp, windows in absence_map.items():
            for (f, t) in windows:
                if f <= week_end and t >= d:
                    absent_emps.add(emp)
                    break
        net = total_staff - len(absent_emps)
        result[wk_key] = net

        # Skill-level: subtract absent staff from their skill pool using
        # our precomputed employee->skills map so mezz-only staff remain
        # exclusive to Mezz Operation.
        sk = dict(skill_pool)
        for emp_id in absent_emps:
            skills = emp_skills.get(emp_id, set())
            if skills == {'Mezz Operation'}:
                sk['Mezz Operation'] = max(0.0, sk.get('Mezz Operation', 0.0) - 1.0)
            elif len(skills) > 0:
                weight = 1.0 / len(skills)
                for sname in skills:
                    sk[sname] = max(0.0, sk.get(sname, 0.0) - weight)
        skill_result[wk_key] = sk
        d += timedelta(weeks=1)

    return result, skill_result


# Pre-compute (cached at module load)
_demand = None
_staff_req = None
_skill_req = None
_staff_avail = None
_skill_avail = None

# Long-term MIP state
_lt_use_optimisation = False   # Toggle: False = baseline, True = MIP
_lt_mip_cache        = None    # Cached MIP result (invalidated on POST reset)


def get_data():
    global _demand, _staff_req, _skill_req, _staff_avail, _skill_avail
    if _demand is None:
        _demand = weekly_demand_2026()
        _staff_req, _skill_req = weekly_staff_required(_demand)
        _staff_avail, _skill_avail = weekly_staff_available()
    return _demand, _staff_req, _skill_req, _staff_avail, _skill_avail


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/long-term/summary')
def lt_summary():
    demand, staff_req, skill_req, staff_avail, skill_avail = get_data()

    all_rows = load_weekly_demand()
    # Include both Historical (W25/S26 actuals) and Forecast rows for 2026
    rows_2026 = [r for r in all_rows if r.get('Data_type', '').strip() in ('Forecast', 'Historical')]

    # Annual flights 2026
    annual_flights = 0
    monthly_flights = defaultdict(float)
    # Also compute passenger estimates for the year (movements × aircraft capacity × load factor)
    annual_passengers = 0.0
    monthly_passengers = defaultdict(float)
    for r in rows_2026:
        d = parse_date(r.get('Week_Start', ''))
        if d and d.year == 2026:
            try:
                mv = float(r.get('Weekly_Movements', 0) or 0)
            except:
                mv = 0
            annual_flights += mv
            monthly_flights[d.month] += mv

            # Estimate passengers for these movements using aircraft capacity + load factor
            try:
                cap = _aircraft_capacity(r)
                lf = _load_factor(r)
                pax_week = mv * cap * lf
            except Exception:
                pax_week = 0.0
            annual_passengers += pax_week
            monthly_passengers[d.month] += pax_week

    avg_weekly = annual_flights / 52 if annual_flights else 0
    avg_weekly_passengers = annual_passengers / 52 if annual_passengers else 0

    # Choose peak month by passenger footfall when available, otherwise by movements
    if monthly_passengers:
        peak_month_num = max(monthly_passengers, key=monthly_passengers.get)
    else:
        peak_month_num = max(monthly_flights, key=monthly_flights.get) if monthly_flights else 7
    peak_month = datetime(2026, peak_month_num, 1).strftime('%B')

    # Peak week
    # Peak week: prefer passenger-weighted peak when possible
    weekly_flights = defaultdict(float)
    weekly_passengers = defaultdict(float)
    for r in rows_2026:
        d = parse_date(r.get('Week_Start', ''))
        if d and d.year == 2026:
            try:
                mv = float(r.get('Weekly_Movements', 0) or 0)
            except:
                mv = 0
            wk = d.strftime('%Y-W%V')
            weekly_flights[wk] += mv
            try:
                cap = _aircraft_capacity(r)
                lf = _load_factor(r)
                pax_week = mv * cap * lf
            except Exception:
                pax_week = 0.0
            weekly_passengers[wk] += pax_week
    if weekly_passengers:
        peak_wk = max(weekly_passengers, key=weekly_passengers.get)
    else:
        peak_wk = max(weekly_flights, key=weekly_flights.get) if weekly_flights else '2026-W28'

    # Staff utilisation: avg (staff_req / staff_avail) across 2026 weeks
    utils = []
    for wk in staff_req:
        avail = staff_avail.get(wk, 50)
        if avail > 0:
            utils.append(min(staff_req[wk] / avail * 100, 100))
    avg_util = round(sum(utils) / len(utils), 1) if utils else 0

    # Avg passengers handled by 1 FTE per day
    # = avg_weekly_passengers / 7 days / total_staff
    total_staff = len(load_staff())
    avg_pax_per_fte_per_day = round(avg_weekly_passengers / 7 / total_staff) if total_staff else 0

    return jsonify({
        'annual_flights': int(annual_flights),
        'avg_weekly_flights': round(avg_weekly, 0),
        'annual_passengers': int(round(annual_passengers)),
        'avg_weekly_passengers': int(round(avg_weekly_passengers)),
        'peak_month': peak_month,
        'peak_week': peak_wk,
        'staff_utilisation_pct': avg_util,
        'total_staff': total_staff,
        'avg_pax_per_fte_per_day': int(avg_pax_per_fte_per_day),
    })


def _build_heatmap_row(wk_key, staff_req, skill_req, staff_avail, demand):
    """Build one heatmap row dict for a given week key."""
    try:
        year_str, w_str = wk_key.split('-W')
        d = datetime.strptime(f'{year_str}-W{int(w_str):02d}-1', '%G-W%V-%u')
    except:
        return None
    req   = staff_req.get(wk_key, 0)
    avail = staff_avail.get(wk_key, 50)
    gap   = round(avail - req, 1)
    util  = round(min(req / avail * 100, 150) if avail > 0 else 100, 1)

    # Per-skill FTE for the week
    skills = skill_req.get(wk_key, {})

    # Per-category flight volumes for the week (raw movements)
    cat_mvmt = {}
    for key, mvmt in demand.get(wk_key, {}).items():
        cat = key[0] if key else 'Unknown'
        cat_mvmt[cat] = round(cat_mvmt.get(cat, 0) + mvmt, 0)

    # Also estimate passengers for the week (movements × aircraft capacity × load factor)
    weekly_passengers = 0.0
    for key, mvmt in demand.get(wk_key, {}).items():
        if len(key) == 5:
            cat, status, terminal, aircraft, load_factor = key
        else:
            cat, status = key[:2]
            terminal, aircraft, load_factor = 'T1', '', '85%'
        row = {
            'Flight_Category': cat,
            'Aircraft_Family': aircraft,
            'Avg_Load_Factor_Pct': load_factor,
        }
        try:
            cap = _aircraft_capacity(row)
            lf = _load_factor(row)
            weekly_passengers += float(mvmt or 0) * cap * lf
        except Exception:
            pass

    return {
        'week':              wk_key,
        'week_start':        d.strftime('%d %b'),
        'week_end':          (d + timedelta(days=6)).strftime('%d %b %Y'),
        'month':             d.strftime('%b'),
        'month_num':         d.month,
        'week_in_month':     (d.day - 1) // 7,
        'required':          req,
        'available':         avail,
        'gap':               gap,
        'utilisation':       util,
        'skills':            {k: round(v, 1) for k, v in skills.items()},
        'categories':        cat_mvmt,
        'weekly_passengers': int(round(weekly_passengers)),
    }


@app.route('/api/long-term/demand-heatmap')
def lt_demand_heatmap():
    demand, staff_req, skill_req, staff_avail, skill_avail = get_data()
    weeks_data = []
    for wk_key in sorted(staff_req.keys()):
        row = _build_heatmap_row(wk_key, staff_req, skill_req, staff_avail, demand)
        if row:
            weeks_data.append(row)
    return jsonify(weeks_data)


@app.route('/api/long-term/week/<week_key>')
def lt_week_detail(week_key):
    """Detailed data for a single week — drives the week-filter mode in the UI."""
    demand, staff_req, skill_req, staff_avail, skill_avail = get_data()
    row = _build_heatmap_row(week_key, staff_req, skill_req, staff_avail, demand)
    if not row:
        return jsonify({'error': 'Week not found'}), 404

    # Weekly flight movements by (category, status) for a mini breakdown chart
    cat_status = {}
    for key, mvmt in demand.get(week_key, {}).items():
        cat = key[0] if len(key) > 0 else 'Unknown'
        status = key[1] if len(key) > 1 else 'Unknown'
        cat_status[f'{cat} ({status})'] = round(cat_status.get(f'{cat} ({status})', 0) + mvmt, 0)

    # Absent staff this week
    absences = load_absences()
    staff    = load_staff()
    absence_map = defaultdict(list)
    for a in absences:
        emp   = a['EMPLOYEE NUMBER'].strip()
        d_from = parse_date(a.get('DATE FROM', ''))
        d_to   = parse_date(a.get('DATE TO', ''))
        if d_from and d_to:
            absence_map[emp].append((d_from, d_to))
    try:
        year_str, w_str = week_key.split('-W')
        wk_start = datetime.strptime(f'{year_str}-W{int(w_str):02d}-1', '%G-W%V-%u')
    except:
        wk_start = None
    wk_end = wk_start + timedelta(days=4) if wk_start else None

    absent_emps = []
    if wk_start and wk_end:
        for emp, windows in absence_map.items():
            for (f, t) in windows:
                if f <= wk_end and t >= wk_start:
                    emp_data = next((s for s in staff if s['EMPLOYEE NUMBER'] == emp), None)
                    if emp_data:
                        absent_emps.append({
                            'id':    emp,
                            'skill': emp_data.get('Skill1', ''),
                            'leave': next((a['LEAVE TYPE'] for a in absences
                                           if a['EMPLOYEE NUMBER'].strip() == emp), ''),
                        })
                    break

    row['cat_status_breakdown'] = cat_status
    row['absent_staff']         = absent_emps
    row['absent_count']         = len(absent_emps)
    row['weekly_passengers']    = int(round(weekly_passengers))
    return jsonify(row)


@app.route('/api/long-term/staff-allocation')
def lt_staff_allocation():
    demand, staff_req, skill_req, staff_avail, skill_avail = get_data()

    # Monthly summary by skill
    monthly_skill = defaultdict(lambda: defaultdict(list))
    monthly_total_req = defaultdict(list)
    monthly_total_avail = defaultdict(list)

    req_skill_set = set()
    for wk_key, sk_req in skill_req.items():
        try:
            year_str, w_str = wk_key.split('-W')
            d = datetime.strptime(f'{year_str}-W{int(w_str):02d}-1', '%G-W%V-%u')
        except:
            continue
        month_key = d.strftime('%b %Y')
        for skill, fte in sk_req.items():
            monthly_skill[month_key][skill].append(fte)
            req_skill_set.add(skill)
        monthly_total_req[month_key].append(staff_req.get(wk_key, 0))
        monthly_total_avail[month_key].append(staff_avail.get(wk_key, 50))

    # Also include staff-availability counts for roles only in staff data (not in demand model)
    for wk_key, sk_av in skill_avail.items():
        try:
            year_str, w_str = wk_key.split('-W')
            d = datetime.strptime(f'{year_str}-W{int(w_str):02d}-1', '%G-W%V-%u')
        except:
            continue
        month_key = d.strftime('%b %Y')
        for skill, count in sk_av.items():
            if skill not in req_skill_set:
                monthly_skill[month_key][skill].append(count)

    # Collect all skills from both demand and staff availability data
    all_skill_set = set()
    for wk_sk in skill_req.values():
        all_skill_set.update(wk_sk.keys())
    for wk_sk in skill_avail.values():
        all_skill_set.update(wk_sk.keys())
    PREFERRED_ORDER = [
        'GNIB', 'CBP Pre-clearance', 'Arr Customer Service', 'Check-in/Trolleys',
        'Dep / Trolleys', 'T1/T2 Trolleys L/UL', 'Gate 335', 'Departures',
        'Transfer Corridor', 'Ramp / Marshalling', 'Bussing',
        'PBZ', 'Mezz Operation', 'Litter Picking',
    ]
    skills = [s for s in PREFERRED_ORDER if s in all_skill_set]
    skills += sorted(s for s in all_skill_set if s not in skills)
    months_ordered = []
    d = datetime(2026, 1, 5)
    seen_months = set()
    while d.year == 2026:
        mk = d.strftime('%b %Y')
        if mk not in seen_months:
            seen_months.add(mk)
            months_ordered.append({'key': mk, 'month_num': d.month})
        d += timedelta(weeks=1)

    def avg(lst): return sum(lst) / len(lst) if lst else 0

    # Gate headcount distribution
    # Contact piers handle GNIB, Ramp, Arr CS, Check-in, CBP, fixed posts
    # Remote Apron handles all Bussing FTE
    # Cargo Apron gets a small share of Ramp FTE (cargo movements)
    PIER_STANDS = {
        'Pier 1 (T1)':      41,
        'Pier 2 (T1)':      28,
        'Pier 3 (T1)':      28,
        'Pier 4 (T2)':      45,
    }
    CONTACT_TOTAL = 142
    GATE_ROWS = list(PIER_STANDS.keys()) + ['Remote Apron', 'Cargo Apron']

    result = []
    gate_monthly = {g: [] for g in GATE_ROWS}

    for mo in months_ordered:
        mk = mo['key']
        row = {'month': mk, 'month_num': mo['month_num']}
        for sk in skills:
            vals = monthly_skill[mk].get(sk, [])
            row[sk] = round(avg(vals), 1)
        row['total_required'] = round(avg(monthly_total_req.get(mk, [])), 1)
        row['total_available'] = round(avg(monthly_total_avail.get(mk, [])), 1)
        row['gap'] = round(row['total_available'] - row['total_required'], 1)
        result.append(row)

        # Gate headcount: split non-bussing FTE across contact piers by stand ratio
        bussing_fte  = row.get('Bussing', 0)
        cargo_fte    = row.get('Ramp / Marshalling', 0) * 0.07   # ~7% cargo ramp
        contact_fte  = row['total_required'] - bussing_fte - cargo_fte
        for pier, stands in PIER_STANDS.items():
            gate_monthly[pier].append(round(contact_fte * (stands / CONTACT_TOTAL), 1))
        gate_monthly['Remote Apron'].append(round(bussing_fte, 1))
        gate_monthly['Cargo Apron'].append(round(cargo_fte, 1))

    # Build by_gate: list of {gate, months: [fte per month]}
    by_gate = []
    for g in GATE_ROWS:
        by_gate.append({
            'gate':   g,
            'values': gate_monthly[g],
        })

    return jsonify({'months': result, 'skills': skills, 'by_gate': by_gate, 'gate_rows': GATE_ROWS})


@app.route('/api/long-term/daily-heatmap')
def lt_daily_heatmap():
    """Daily heatmap data for the entire 2026 planning year."""
    demand, staff_req, skill_req, staff_avail, skill_avail = get_data()
    
    skills = [
        'GNIB', 'CBP Pre-clearance', 'Arr Customer Service', 'Check-in/Trolleys',
        'Dep / Trolleys', 'T1/T2 Trolleys L/UL', 'Transfer Corridor',
        'Ramp / Marshalling', 'Bussing', 'PBZ', 'Mezz Operation', 'Litter Picking',
    ]

    # Generate all days for 2026
    start_date = datetime(2026, 1, 1)
    end_date = datetime(2026, 12, 31)
    
    today_str = datetime.now().strftime('%Y-%m-%d')
    days = []
    curr = start_date
    while curr <= end_date:
        # ISO week key
        wk_key = curr.strftime('%Y-W%V')
        iso_str = curr.strftime('%Y-%m-%d')
        
        day_data = {
            'date': iso_str,
            'label': curr.strftime('%d %b'),
            'is_today': iso_str == today_str,
            'values': {}
        }
        
        wk_sk_req = skill_req.get(wk_key, {})
        wk_sk_avail = skill_avail.get(wk_key, {})
        
        day_total_req = 0
        day_total_avail = 0
        
        for sk in skills:
            req = wk_sk_req.get(sk, 0)
            avail = wk_sk_avail.get(sk, 0)
            
            day_total_req += req
            day_total_avail += avail
            
            gap = avail - req
            
            # Status Logic
            if gap < -2.0:
                status = 'gap'
            elif gap < 0:
                status = 'warning'
            elif gap > 1.0:
                status = 'surplus'
            else:
                status = 'adequate'
                
            day_data['values'][sk] = {
                'req': round(req, 1),
                'avail': round(avail, 1),
                'status': status
            }
            
        day_data['totals'] = {
            'req': round(day_total_req, 1),
            'avail': round(day_total_avail, 1),
            'gap': round(day_total_avail - day_total_req, 1)
        }
        
        days.append(day_data)
        curr += timedelta(days=1)
        
    return jsonify({
        'skills': skills,
        'days': days
    })


@app.route('/api/long-term/imbalance')
def lt_imbalance():
    demand, staff_req, skill_req, staff_avail, skill_avail = get_data()

    result = []
    for wk_key in sorted(staff_req.keys()):
        try:
            year_str, w_str = wk_key.split('-W')
            d = datetime.strptime(f'{year_str}-W{int(w_str):02d}-1', '%G-W%V-%u')
        except:
            continue
        req = staff_req[wk_key]
        avail = staff_avail.get(wk_key, 50)
        gap = round(avail - req, 1)
        result.append({
            'week': wk_key,
            'date': d.strftime('%d %b %Y'),
            'month': d.strftime('%b'),
            'required': req,
            'available': avail,
            'gap': gap,
            'status': 'ok' if gap > 0 else ('warning' if gap == 0 else 'critical'),
        })

    return jsonify(result)


@app.route('/api/long-term/skill-breakdown')
def lt_skill_breakdown():
    staff = load_staff()
    absences = load_absences()
    absence_map = defaultdict(list)
    for a in absences:
        emp = a['EMPLOYEE NUMBER'].strip()
        d_from = parse_date(a.get('DATE FROM', ''))
        d_to = parse_date(a.get('DATE TO', ''))
        if d_from and d_to:
            absence_map[emp].append((d_from, d_to))

    # Total staff by all qualified skills
    total_by_skill = defaultdict(int)
    for s in staff:
        for sk_col in ['Skill1', 'Skill2', 'Skill3', 'Skill4']:
            sk_name = s.get(sk_col, '').strip()
            if sk_name:
                total_by_skill[sk_name] += 1

    skills = sorted(total_by_skill.keys())
    monthly_absent, leave_types = get_monthly_absence_impact()
    
    return jsonify({
        'total_by_skill': dict(total_by_skill),
        'monthly_absent': monthly_absent,
        'leave_types': leave_types,
        'skills': skills,
    })


@app.route('/api/long-term/flight-trend')
def lt_flight_trend():
    rows = load_weekly_demand()

    # Monthly passenger footfall: movements × aircraft_capacity × load_factor
    monthly = defaultdict(lambda: {'historical': 0.0, 'forecast': 0.0})
    for r in rows:
        if not r.get('Season_Code'):
            continue
        d = parse_date(r.get('Week_Start', ''))
        if not d:
            continue
        try:
            mv = float(r.get('Weekly_Movements', 0) or 0)
        except:
            mv = 0
        pax = mv * _aircraft_capacity(r) * _load_factor(r)
        dtype = r.get('Data_type', '').strip()
        mk = d.strftime('%b')
        if d.year == 2025 and dtype == 'Historical':
            monthly[mk]['historical'] += pax
        elif d.year == 2026 and dtype in ('Forecast', 'Historical'):
            monthly[mk]['forecast'] += pax

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    return jsonify([{
        'month': m,
        'historical': int(round(monthly[m]['historical'])),
        'forecast':   int(round(monthly[m]['forecast'])),
    } for m in months])


@app.route('/api/long-term/gap-skill-data')
def lt_merged_gap_skill():
    demand, staff_req, skill_req, staff_avail, skill_avail = get_data()

    # Get all active skills from both req and avail
    all_skills = set()
    for wk in skill_req:
        all_skills.update(skill_req[wk].keys())
    for wk in skill_avail:
        all_skills.update(skill_avail[wk].keys())
    all_skills = sorted(list(all_skills))

    weekly_data = []
    for wk_key in sorted(staff_req.keys()):
        try:
            year_str, w_str = wk_key.split('-W')
            d = datetime.strptime(f'{year_str}-W{int(w_str):02d}-1', '%G-W%V-%u')
        except:
            continue
        
        req_total = staff_req[wk_key]
        avail_total = staff_avail.get(wk_key, 50)
        gap_total = round(avail_total - req_total, 1)

        sk_gaps = {}
        sk_reqs = {}
        sk_avails = {}
        for sk in all_skills:
            s_req = skill_req.get(wk_key, {}).get(sk, 0)
            s_avail = skill_avail.get(wk_key, {}).get(sk, 0)
            sk_gaps[sk] = round(s_avail - s_req, 1)
            sk_reqs[sk] = round(s_req, 1)
            sk_avails[sk] = round(s_avail, 1)

        weekly_data.append({
            'week': wk_key,
            'date': d.strftime('%d %b %Y'),
            'month': d.strftime('%b'),
            'required': req_total,
            'available': avail_total,
            'gap': gap_total,
            'skill_gaps': sk_gaps,
            'skill_reqs': sk_reqs,
            'skill_avails': sk_avails,
            'status': 'ok' if gap_total > 0 else ('warning' if gap_total == 0 else 'critical')
        })

    # Summary by skill (average across all weeks)
    skill_summary = []
    for sk in all_skills:
        # Extract data for this skill across all weeks
        weeks_data = []
        for w in weekly_data:
            weeks_data.append({
                'gap': w['skill_gaps'].get(sk, 0),
                'req': w['skill_reqs'].get(sk, 0),
                'avail': w['skill_avails'].get(sk, 0)
            })
        
        if not weeks_data:
            continue
            
        gaps = [d['gap'] for d in weeks_data]
        reqs = [d['req'] for d in weeks_data]
        avails = [d['avail'] for d in weeks_data]
        
        avg_req = round(sum(reqs) / len(reqs), 1)
        avg_avail = round(sum(avails) / len(avails), 1)
        avg_gap = round(avg_avail - avg_req, 1)
        
        # Define "Peak" as the week with the maximum demand (highest requirement)
        # This ensures that Peak Avail - Peak Req = Peak Gap
        peak_week = max(weeks_data, key=lambda x: x['req'])
        peak_req = round(peak_week['req'], 1)
        peak_avail = round(peak_week['avail'], 1)
        peak_gap = round(peak_avail - peak_req, 1)
        
        min_gap = round(min(gaps), 1)

        skill_summary.append({
            'skill': sk,
            'avg_gap': avg_gap,
            'peak_gap': peak_gap,
            'min_gap': min_gap,
            'avg_req': avg_req,
            'avg_avail': avg_avail,
            'peak_req': peak_req,
            'peak_avail': peak_avail,
            'status': 'ok' if avg_gap > 0 else ('warning' if avg_gap == 0 else 'critical')
        })

    # Historical monthly absences (from original skills API)
    staff = load_staff()
    
    total_by_skill = defaultdict(int)
    for s in staff:
        total_by_skill[s.get('Skill1', '').strip()] += 1

    weekly_absent, leave_types, week_labels = get_weekly_absence_impact()

    total_by_skill = defaultdict(int)
    for s in staff:
        total_by_skill[s.get('Skill1', '').strip()] += 1

    return jsonify({
        'weekly': weekly_data,
        'skill_summary': skill_summary,
        'total_by_skill': dict(total_by_skill),
        'weekly_absent': weekly_absent,
        'weekly_absence_labels': week_labels,
        'leave_types': leave_types,
        'skills': all_skills
    })


# ---------------------------------------------------------------------------
# Shift optimisation constants & MIP engine
# ---------------------------------------------------------------------------

_ROSTER_SHIFTS = [
    {'id': 'Early',   'label': 'Early (00:00–12:00)',   'color': '#f97316',
     'hours': list(range(0, 12))},
    {'id': 'Mid',     'label': 'Mid (06:00–18:00)',     'color': '#3b82f6',
     'hours': list(range(6, 18))},
    {'id': 'Late',    'label': 'Late (12:00–00:00)',    'color': '#8b5cf6',
     'hours': list(range(12, 24))},
    {'id': 'Evening', 'label': 'Evening (16:00–04:00)', 'color': '#10b981',
     'hours': list(range(16, 24)) + list(range(0, 4))},
    {'id': 'Night',   'label': 'Night (22:00–10:00)',   'color': '#ec4899',
     'hours': list(range(22, 24)) + list(range(0, 10))},
]

# ── Staggered Break Templates ──────────────────────────────────────────────────
# Key: (shift_start_mins, shift_end_mins) → {group: [break_dict, ...]}
# Each group has one 30-min short break and one 60-min meal break, staggered
# by 1 hour relative to the other group to maintain continuous coverage.
STAGGER_BREAK_TEMPLATES = {
    (0, 720): {        # Early  00:00–12:00
        'name': 'Early',
        'A': [
            {'start': '03:00', 'end': '03:30', 'start_mins': 180,  'end_mins': 210,  'type': 'Short Break'},
            {'start': '07:00', 'end': '08:00', 'start_mins': 420,  'end_mins': 480,  'type': 'Meal Break'},
        ],
        'B': [
            {'start': '04:00', 'end': '04:30', 'start_mins': 240,  'end_mins': 270,  'type': 'Short Break'},
            {'start': '08:00', 'end': '09:00', 'start_mins': 480,  'end_mins': 540,  'type': 'Meal Break'},
        ],
    },
    (360, 1080): {     # Mid    06:00–18:00
        'name': 'Mid',
        'A': [
            {'start': '09:00', 'end': '09:30', 'start_mins': 540,  'end_mins': 570,  'type': 'Short Break'},
            {'start': '13:00', 'end': '14:00', 'start_mins': 780,  'end_mins': 840,  'type': 'Meal Break'},
        ],
        'B': [
            {'start': '10:00', 'end': '10:30', 'start_mins': 600,  'end_mins': 630,  'type': 'Short Break'},
            {'start': '14:00', 'end': '15:00', 'start_mins': 840,  'end_mins': 900,  'type': 'Meal Break'},
        ],
    },
    (720, 1440): {     # Late   12:00–00:00
        'name': 'Late',
        'A': [
            {'start': '15:00', 'end': '15:30', 'start_mins': 900,  'end_mins': 930,  'type': 'Short Break'},
            {'start': '19:00', 'end': '20:00', 'start_mins': 1140, 'end_mins': 1200, 'type': 'Meal Break'},
        ],
        'B': [
            {'start': '16:00', 'end': '16:30', 'start_mins': 960,  'end_mins': 990,  'type': 'Short Break'},
            {'start': '20:00', 'end': '21:00', 'start_mins': 1200, 'end_mins': 1260, 'type': 'Meal Break'},
        ],
    },
    (960, 1680): {     # Evening 16:00–04:00
        'name': 'Evening',
        'A': [
            {'start': '19:00', 'end': '19:30', 'start_mins': 1140, 'end_mins': 1170, 'type': 'Short Break'},
            {'start': '23:00', 'end': '00:00', 'start_mins': 1380, 'end_mins': 1440, 'type': 'Meal Break'},
        ],
        'B': [
            {'start': '20:00', 'end': '20:30', 'start_mins': 1200, 'end_mins': 1230, 'type': 'Short Break'},
            {'start': '00:00', 'end': '01:00', 'start_mins': 1440, 'end_mins': 1500, 'type': 'Meal Break'},
        ],
    },
    (1320, 2040): {    # Night  22:00–10:00
        'name': 'Night',
        'A': [
            {'start': '01:00', 'end': '01:30', 'start_mins': 1500, 'end_mins': 1530, 'type': 'Short Break'},
            {'start': '05:00', 'end': '06:00', 'start_mins': 1740, 'end_mins': 1800, 'type': 'Meal Break'},
        ],
        'B': [
            {'start': '02:00', 'end': '02:30', 'start_mins': 1560, 'end_mins': 1590, 'type': 'Short Break'},
            {'start': '06:00', 'end': '07:00', 'start_mins': 1800, 'end_mins': 1860, 'type': 'Meal Break'},
        ],
    },
}

# 24-hour aviation demand profile (relative units, index 0=00:00 … 23=23:00)
_HOURLY_PROFILE = [
    0.5, 0.3, 0.2, 0.2, 0.3, 0.8,   # 00–05
    2.0, 3.5, 4.5, 4.0, 3.0, 2.5,   # 06–11
    2.5, 2.8, 3.5, 4.0, 4.5, 4.2,   # 12–17
    3.8, 3.0, 2.2, 1.5, 1.0, 0.7,   # 18–23
]

# Coverage matrix: _SHIFT_COV[shift_id][h] = 1 if shift covers hour h
_SHIFT_COV = {sh['id']: [1 if h in sh['hours'] else 0 for h in range(24)]
              for sh in _ROSTER_SHIFTS}

# 3-hour time block definitions
_TIME_BLOCKS = [
    {'id': 'b00_03', 'label': '00–03', 'hours': [0, 1, 2]},
    {'id': 'b03_06', 'label': '03–06', 'hours': [3, 4, 5]},
    {'id': 'b06_09', 'label': '06–09', 'hours': [6, 7, 8]},
    {'id': 'b09_12', 'label': '09–12', 'hours': [9, 10, 11]},
    {'id': 'b12_15', 'label': '12–15', 'hours': [12, 13, 14]},
    {'id': 'b15_18', 'label': '15–18', 'hours': [15, 16, 17]},
    {'id': 'b18_21', 'label': '18–21', 'hours': [18, 19, 20]},
    {'id': 'b21_24', 'label': '21–24', 'hours': [21, 22, 23]},
]
_TOTAL_PROFILE = sum(_HOURLY_PROFILE)


def _solve_shift_mip(daily_total_fte: float) -> dict:
    """Allocate daily FTE across shift templates proportional to demand covered.

    Each shift receives a share of daily_total_fte proportional to the sum of
    the demand profile over its covered hours.  The five allocations always sum
    to daily_total_fte.

    Returns {shift_id: fte_float}.
    """
    if daily_total_fte <= 0:
        return {sh['id']: 0.0 for sh in _ROSTER_SHIFTS}

    covered = {sh['id']: sum(_HOURLY_PROFILE[h] for h in sh['hours'])
               for sh in _ROSTER_SHIFTS}
    total_covered = sum(covered.values()) or 1.0

    return {sh['id']: round(daily_total_fte * covered[sh['id']] / total_covered, 2)
            for sh in _ROSTER_SHIFTS}


# ---------------------------------------------------------------------------
# Four-Week Roster endpoint
# ---------------------------------------------------------------------------

@app.route('/api/long-term/four-week-roster')
def lt_four_week_roster():
    """Return daily FTE breakdown and shift plan for the next 4 ISO weeks.

    Methodology:
    1. Compute day-of-week weights from the ratio of Peak_Day_Movements /
       Weekly_Movements across all historical rows, then normalise a canonical
       aviation DOW distribution to those weights.
    2. Retrieve weekly skill FTE for the next 4 ISO weeks via
       weekly_staff_required().
    3. Expand each week to 7 days using the DOW weights.
    4. Derive Early / Late / Night shift headcount from daily totals.
    """
    # ── 1. Day-of-week weights from historical peak-day ratios ──────────────
    # Aviation baseline weights (Mon–Sun index 0–6).
    # These reflect typical airport traffic patterns (Fri/Sat peaks, Tue troughs).
    _BASE_DOW = [0.150, 0.120, 0.135, 0.145, 0.170, 0.165, 0.115]

    rows = load_weekly_demand()
    # Collect peak_day / weekly ratios per row to compute average "peakiness"
    ratios = []
    for r in rows:
        dtype = r.get('Data_type', '').strip()
        if dtype != 'Historical':
            continue
        try:
            wm = float(r.get('Weekly_Movements', 0) or 0)
            pd_val = float(r.get('Peak_Day_Movements', 0) or 0)
            if wm > 0 and pd_val > 0:
                ratios.append(pd_val / wm)
        except (ValueError, TypeError):
            pass

    # Average peak-day share across historical rows
    avg_peak_ratio = sum(ratios) / len(ratios) if ratios else 1.0 / 7

    # The peak DOW (index 4 = Fri) should equal avg_peak_ratio.
    # Scale the base weights so that peak matches, then re-normalise.
    peak_idx = _BASE_DOW.index(max(_BASE_DOW))  # Fri = index 4
    scale = avg_peak_ratio / _BASE_DOW[peak_idx] if _BASE_DOW[peak_idx] > 0 else 1.0
    scaled = [w * scale for w in _BASE_DOW]
    total_w = sum(scaled)
    dow_weights = [round(w / total_w, 4) for w in scaled]  # Mon–Sun, sum=1.0

    DOW_LABELS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    # ── 2. Next 4 ISO weeks' weekly skill FTE ──────────────────────────────
    today = datetime.now()
    # Compute start of next full ISO week (Monday)
    days_until_monday = (7 - today.weekday()) % 7
    if days_until_monday == 0:
        days_until_monday = 7  # always start next week
    next_monday = today + timedelta(days=days_until_monday)

    target_weeks = []
    for i in range(4):
        monday = next_monday + timedelta(weeks=i)
        wk_key = monday.strftime('%Y-W%V')
        target_weeks.append({'monday': monday, 'wk_key': wk_key})

    # Fetch weekly demand data for those weeks
    demand_all = weekly_demand_2026()
    staff_req_all, skill_req_all = weekly_staff_required(demand_all)
    _staff_avail_all, skill_avail_all = weekly_staff_available()

    # Collect all skill names
    all_skills_set = set()
    for wk_data in skill_req_all.values():
        all_skills_set.update(wk_data.keys())
    all_skills = sorted(all_skills_set)

    # ── 3. Helper: build day objects for a given week ─────────────────────
    def _build_day(date, dow_idx, weekly_skill_fte, weekly_total_fte,
                   weekly_skill_avail=None, weekly_total_avail=0,
                   coverage='roster_only'):
        weight = dow_weights[dow_idx]
        skill_fte_day = {}
        for sk in all_skills:
            skill_fte_day[sk] = round(weekly_skill_fte.get(sk, 0) * weight * 7, 2)

        daily_total_fte = round(weekly_total_fte * weight * 7, 2)

        # Availability (same DOW decomposition)
        skill_avail_day = {}
        wsa = weekly_skill_avail or {}
        for sk in all_skills:
            skill_avail_day[sk] = round(wsa.get(sk, 0) * weight * 7, 2)
        daily_total_avail = round(weekly_total_avail * weight * 7, 2)

        # Gap = Available − Required  (positive = surplus, negative = shortfall)
        skill_gap_day = {sk: round(skill_avail_day[sk] - skill_fte_day[sk], 2)
                         for sk in all_skills}
        daily_total_gap = round(daily_total_avail - daily_total_fte, 2)

        # MIP-based shift allocation
        shift_solution = _solve_shift_mip(daily_total_fte)
        skill_fracs = ({sk: skill_fte_day[sk] / daily_total_fte for sk in all_skills}
                       if daily_total_fte > 0 else {sk: 0.0 for sk in all_skills})

        shifts = []
        for sh in _ROSTER_SHIFTS:
            sh_fte = shift_solution[sh['id']]
            shift_skills = {sk: round(sh_fte * skill_fracs.get(sk, 0), 2)
                            for sk in all_skills}
            shifts.append({
                'id':        sh['id'],
                'name':      sh['label'],
                'color':     sh['color'],
                'total_fte': round(sh_fte, 2),
                'skills':    shift_skills,
            })

        # 3-hour time block allocation (demand-weighted staffing level per block)
        time_blocks = []
        for tb in _TIME_BLOCKS:
            tb_profile = sum(_HOURLY_PROFILE[h] for h in tb['hours'])
            # Same logic as shift plan: block gets share proportional to demand it covers.
            # Blocks are non-overlapping so total_covered = _TOTAL_PROFILE and sum of
            # all tb_scale = 1.0, meaning sum of all block FTEs = daily_total_fte.
            tb_scale = tb_profile / _TOTAL_PROFILE
            time_blocks.append({
                'id':        tb['id'],
                'label':     tb['label'],
                'total_fte': round(daily_total_fte * tb_scale, 2),
                'skills':    {sk: round(skill_fte_day[sk] * tb_scale, 2) for sk in all_skills},
            })

        return {
            'date':           date.strftime('%Y-%m-%d'),
            'label':          date.strftime('%a %d %b'),
            'dow':            dow_idx,
            'dow_label':      DOW_LABELS[dow_idx],
            'weight':         weight,
            'total_fte':      daily_total_fte,
            'skill_fte':      skill_fte_day,
            'total_avail':    daily_total_avail,
            'skill_avail':    skill_avail_day,
            'total_gap':      daily_total_gap,
            'skill_gap':      skill_gap_day,
            'shifts':         shifts,
            'time_blocks':    time_blocks,
            'coverage':       coverage,
        }

    # ── 4. Current-week partial block ──────────────────────────────────────
    #   Intraday  = today
    #   Short-term = today + 0..3  (4-day CSV: Flights_schedule_4days.csv)
    #   Roster-only = today+4 .. Sunday of current ISO week
    today_date = today.date() if hasattr(today, 'date') else today
    today_d    = today if not hasattr(today, 'date') else today.replace(hour=0, minute=0, second=0, microsecond=0)
    cur_week_monday = today_d - timedelta(days=today_d.weekday())
    cur_week_sunday = cur_week_monday + timedelta(days=6)
    cur_wk_key      = cur_week_monday.strftime('%Y-W%V')

    cur_weekly_skill_fte   = skill_req_all.get(cur_wk_key, {})
    cur_weekly_total_fte   = staff_req_all.get(cur_wk_key, 0)
    cur_weekly_skill_avail = skill_avail_all.get(cur_wk_key, {})
    cur_week_avail         = _staff_avail_all.get(cur_wk_key, 0)

    # Short term covers 4 days: today (day 0) through today+3
    ST_WINDOW = 4  # days covered by Flights_schedule_4days.csv

    current_week_days = []
    curr_d = today_d  # start from today itself
    while curr_d <= cur_week_sunday:
        delta = (curr_d - today_d).days
        if delta == 0:
            coverage = 'intraday'
        elif delta < ST_WINDOW:
            coverage = 'short_term'
        else:
            coverage = 'roster_only'

        dow_idx = curr_d.weekday()  # 0=Mon … 6=Sun
        current_week_days.append(
            _build_day(curr_d, dow_idx,
                       cur_weekly_skill_fte, cur_weekly_total_fte,
                       cur_weekly_skill_avail, cur_week_avail,
                       coverage)
        )
        curr_d += timedelta(days=1)

    current_week_block = {
        'week':         cur_wk_key,
        'start_date':   cur_week_monday.strftime('%d %b %Y'),
        'end_date':     cur_week_sunday.strftime('%d %b %Y'),
        'weekly_fte':   round(cur_weekly_total_fte, 1),
        'weekly_avail': cur_week_avail,
        'days':         current_week_days,
        'is_current_week': True,
    }

    # ── 5. Build full 4-week output ─────────────────────────────────────────
    weeks_output = [current_week_block]  # prepend current week

    for wk_info in target_weeks:
        monday = wk_info['monday']
        wk_key = wk_info['wk_key']

        weekly_skill_fte   = skill_req_all.get(wk_key, {})
        weekly_total_fte   = staff_req_all.get(wk_key, 0)
        weekly_skill_avail = skill_avail_all.get(wk_key, {})
        week_avail         = _staff_avail_all.get(wk_key, 0)

        days = []
        for dow_idx in range(7):
            date = monday + timedelta(days=dow_idx)
            days.append(_build_day(date, dow_idx,
                                   weekly_skill_fte, weekly_total_fte,
                                   weekly_skill_avail, week_avail,
                                   'roster_only'))

        weeks_output.append({
            'week':            wk_key,
            'start_date':      monday.strftime('%d %b %Y'),
            'end_date':        (monday + timedelta(days=6)).strftime('%d %b %Y'),
            'weekly_fte':      round(weekly_total_fte, 1),
            'weekly_avail':    week_avail,
            'days':            days,
            'is_current_week': False,
        })

    return jsonify({
        'weeks':            weeks_output,
        'dow_weights':      dict(zip(DOW_LABELS, dow_weights)),
        'skills':           all_skills,
        'shift_templates':  [{'id': sh['id'], 'name': sh['label'], 'color': sh['color']}
                             for sh in _ROSTER_SHIFTS],
        'time_block_defs':  [{'id': tb['id'], 'label': tb['label']}
                             for tb in _TIME_BLOCKS],
        'st_window_days':   ST_WINDOW,
    })


# ---------------------------------------------------------------------------
# Long-term MIP optimised staffing endpoint
# ---------------------------------------------------------------------------

@app.route('/api/long-term/optimised-staffing', methods=['GET', 'POST'])
def lt_optimised_staffing():
    """MIP-based weekly workforce optimisation.

    GET  — return current result (or prompt to enable).
    POST — toggle optimisation on/off; optionally pass constraints overrides.

    POST body (all optional):
        {
          "use_optimisation": true,        // enable / disable MIP
          "reset": true,                   // clear cached result and re-solve
          "constraints": {
            "surge_demand_factor":        1.0,
            "max_ot_hrs_per_person_per_week": 8,
            "regular_hrs_per_week":       40
          }
        }

    Response shape (when enabled and solved):
        {
          "use_optimisation": true,
          "status": "Optimal",
          "solver": "CBC via PuLP (MIP)",
          "objective_value": 12345.6,
          "summary": { ... aggregate stats ... },
          "weeks": {
            "2026-W01": {
              "total_demand_fte": 42.3,
              "total_available": 48,
              "total_assigned": 41,
              "total_shortage_fte": 1.3,
              "overtime_hrs": 0.0,
              "utilisation_pct": 85.4,
              "gap": 5.7,
              "status_flag": "minor",
              "skills": {
                "GNIB": {
                  "demand_fte": 12.5, "available": 18,
                  "assigned": 12, "shortage_fte": 0.5,
                  "excess_fte": 0.0, "coverage_pct": 96.0
                }, ...
              }
            }, ...
          }
        }
    """
    global _lt_use_optimisation, _lt_mip_cache

    if request.method == 'POST':
        body = request.get_json(force=True) or {}

        # Toggle
        if 'use_optimisation' in body:
            _lt_use_optimisation = bool(body['use_optimisation'])

        # Invalidate cache on reset or when constraints change
        if body.get('reset') or 'constraints' in body or 'use_optimisation' in body:
            _lt_mip_cache = None

        # Store per-request constraint overrides in the body for use below
        _lt_extra_constraints = body.get('constraints', {})
    else:
        _lt_extra_constraints = {}

    # --- Status: disabled ---
    if not _lt_use_optimisation:
        return jsonify({
            'use_optimisation': False,
            'mip_available':    _LT_MIP_AVAILABLE,
            'message': (
                'MIP optimisation is disabled. '
                'POST {"use_optimisation": true} to enable.'
            ),
            # Include the baseline demand/availability for comparison
            'baseline': _lt_baseline_summary(),
        })

    # --- Status: solver unavailable ---
    if not _LT_MIP_AVAILABLE:
        return jsonify({
            'use_optimisation': True,
            'mip_available':    False,
            'error': 'No MIP solver available.',
            'message': 'Install PuLP:  pip install pulp',
        }), 503

    # --- Solve (or return cached result) ---
    if _lt_mip_cache is None:
        demand, staff_req, skill_req, staff_avail, skill_avail = get_data()
        try:
            _lt_mip_cache = _lt_mip_optimize(
                skill_req, staff_req,
                skill_avail, staff_avail,
                _lt_extra_constraints or None,
            )
            print(
                f"[MIP] {_lt_mip_cache['status']}  "
                f"obj={_lt_mip_cache['objective_value']}  "
                f"shortage={_lt_mip_cache['summary'].get('total_shortage_fte', '?')} FTE"
            )
        except Exception as exc:
            return jsonify({'error': str(exc), 'use_optimisation': True}), 500

    return jsonify(_lt_mip_cache)


@app.route('/api/long-term/optimisation-status', methods=['GET'])
def lt_optimisation_status():
    """Quick status check: is MIP enabled, available, and what does it report."""
    return jsonify({
        'use_optimisation': _lt_use_optimisation,
        'mip_available':    _LT_MIP_AVAILABLE,
        'result_cached':    _lt_mip_cache is not None,
        'solver_status':    _lt_mip_cache.get('status')  if _lt_mip_cache else None,
        'objective_value':  _lt_mip_cache.get('objective_value') if _lt_mip_cache else None,
        'summary':          _lt_mip_cache.get('summary') if _lt_mip_cache else None,
    })


def _lt_baseline_summary():
    """Return a lightweight baseline summary for comparison in the disabled-state response."""
    try:
        _, staff_req, skill_req, staff_avail, _ = get_data()
        weeks = sorted(staff_req.keys())
        if not weeks:
            return {}
        total_short = sum(
            max(staff_req[w] - staff_avail.get(w, 0), 0) for w in weeks
        )
        weeks_short = sum(
            1 for w in weeks if staff_req[w] > staff_avail.get(w, 0)
        )
        return {
            'method':               'demand_estimation_baseline',
            'total_shortage_fte':   round(total_short, 1),
            'weeks_with_shortage':  weeks_short,
            'total_weeks':          len(weeks),
        }
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Route Map API
# ---------------------------------------------------------------------------

def _get_map_data_from_flights(flights_list):
    """Aggregates arrivals and departures by airport code."""
    res = {
        'arrivals': defaultdict(int),
        'departures': defaultdict(int)
    }
    for f in flights_list:
        code = f.get('origin_code', '').strip()
        status = f.get('Status', '').strip()
        if not code or code == 'DUB':
            continue
        if status == 'Arrival':
            res['arrivals'][code] += 1
        elif status == 'Departure':
            res['departures'][code] += 1
    
    # Convert to standard dict for jsonify
    return {
        'arrivals': dict(res['arrivals']),
        'departures': dict(res['departures'])
    }

@app.route('/api/map-data/long-term')
def map_data_long_term():
    """Returns aggregated route counts for the entire 4-day schedule as a representative network."""
    flights = read_csv_flights()
    return jsonify(_get_map_data_from_flights(flights))

@app.route('/api/map-data/short-term/<date_str>')
def map_data_short_term(date_str):
    """Returns route counts for a specific date."""
    # Convert date_str (YYYY-MM-DD or DD-MMM-YY) to DD-MMM-YY for CSV match
    d = None
    for fmt in ('%Y-%m-%d', '%d-%b-%y', '%d-%b-%Y', '%d-%m-%y'):
        try:
            d = datetime.strptime(date_str.strip(), fmt)
            break
        except ValueError:
            pass
    if not d:
        return jsonify({'error': 'Invalid date'}), 400
    
    target_date = d.strftime('%d-%b-%y')
    flights = [f for f in read_csv_flights() if f.get('date') == target_date]
    return jsonify(_get_map_data_from_flights(flights))

@app.route('/api/map-data/intraday')
def map_data_intraday():
    """Returns route counts for today (hardcoded to 13-Apr-26 in this project)."""
    target_date = '13-Apr-26'
    flights = [f for f in read_csv_flights() if f.get('date') == target_date]
    return jsonify(_get_map_data_from_flights(flights))


# ===========================================================================
# SHORT-TERM & INTRADAY OPTIMISATION
# ===========================================================================

TASK_SKILL = {
    'checkin':               'Checkin',
    'security':              'Security',
    'cbp':                   'CBP',
    'lounge':                'Lounge',
    'boarding':              'Boarding',
    'immigration':           'Immigration',
    'baggage':               'Baggage',
    # New Config.csv task names
    'GNIB':                  'GNIB',
    'Gate 335':              'Gate 335',
    'Departures':            'Departures',
    'Dep/Trolleys':          'Dep/Trolleys',
    # Legacy / alternate spellings kept for backward compat
    'GNIB / Immigration':    'GNIB',
    'CBP Pre-clearance':     'CBP Pre-clearance',
    'Ramp / Marshalling':    'GNIB',
    'Bussing':               'Bussing',
    'Transfer Corridor':     'Transfer Corridor',
    'Check-in/Trolleys':     'Check-in/Trolleys',
    'Dep / Trolleys':        'Dep/Trolleys',
    'Arr Customer Service':  'Arr Customer Service',
    'Mezz Operation':        'Mezz Operation',
    'Litter Picking':        'Litter Picking',
    'PBZ':                   'PBZ',
    'T1/T2 Trolleys L/UL':   'T1/T2 Trolleys L/UL',
}

# ---------------------------------------------------------------------------
# Task sharing classification
#
# SHARED        — One pool of staff covers ALL flights that share the same
#                 terminal, pier, and 30-minute time-bucket.  These are
#                 area-based duties: the officer handles all passengers in the
#                 zone, regardless of which individual aircraft they arrived on.
#
# PARTIALLY_SHARED — Staff can be shared across concurrent flights but headcount
#                 scales with the combined passenger volume in the window.
#                 Thresholds are defined in PARTIAL_SHARE_PAX_THRESHOLDS.
#
# DEDICATED     — Every flight requires its own dedicated staff member(s).
#                 Ramp marshalling is aircraft-specific; bussing is bus-specific.
# ---------------------------------------------------------------------------

TASK_SHARED = frozenset({
    'GNIB / Immigration',
    'CBP Pre-clearance',    # session-level handled separately; included for completeness
    'Transfer Corridor',
    'Arr Customer Service',
    'Mezz Operation',
    'Litter Picking',
    'T1/T2 Trolleys L/UL',
})

TASK_PARTIALLY_SHARED = frozenset({
    'Check-in/Trolleys',
    'Dep / Trolleys',
    'PBZ',
})

TASK_DEDICATED = frozenset({
    'Ramp / Marshalling',
    'Bussing',
})

# Approximate seated-pax capacity by ICAO wake-turbulence category.
# Used to estimate passenger volume when the flights CSV has no pax column.
PAX_BY_ICAO = {
    'A': 50,    # light piston / turboprop
    'B': 150,   # narrow-body small  (A319, B737-700)
    'C': 180,   # narrow-body large  (A321, B757)
    'D': 280,   # wide-body medium   (B767, A330-200)
    'E': 350,   # wide-body large    (B777, A340)
    'F': 450,   # super-heavy        (B747, A380)
}
PAX_DEFAULT = 150   # fallback for unknown ICAO category

# Partially-shared staffing thresholds: (min_pax_inclusive, max_pax_exclusive, staff_count).
# staff_count is the number needed for the *whole group* in that time-window.
PARTIAL_SHARE_PAX_THRESHOLDS = [
    (0,   200, 1),
    (200, 400, 2),
    (400, float('inf'), 3),
]

# Width of the time-bucket (minutes) used to decide whether flights are
# "concurrent" for sharing purposes.  Configurable here; used throughout.
SHARING_WINDOW_MINS: int = 30

# Module-level state
_config_rules = None
_stands_map = None
_intraday_overrides = {}   # {flight_no: {delay_mins: int, cancelled: bool}}
_manual_assigns = {}       # {date_key: {task_id: [extra_staff_ids]}}
_manual_unassigns = {}     # {date_key: {task_id: [blocked_staff_ids]}}
_manual_block_assigns = {} # {date_key: {block_key: {skill, terminal, block_start, block_end, staff_ids}}}


def _apply_manual_unassigns_to_result(result, date_key):
    """Remove manually blocked staff from specific task ids in an optimiser result."""
    blocks = _manual_unassigns.get(date_key, {})
    if not blocks:
        return

    task_lookup = {t.get('id'): t for t in result.get('tasks', [])}
    removed = set()
    for task_id, staff_ids in blocks.items():
        task = task_lookup.get(task_id)
        if not task:
            continue
        blocked = {str(sid) for sid in (staff_ids or [])}
        assigned = [str(sid) for sid in (task.get('assigned') or [])]
        next_assigned = [sid for sid in assigned if sid not in blocked]
        if len(next_assigned) == len(assigned):
            continue
        task['assigned'] = next_assigned
        needed = int(task.get('staff_needed') or 0)
        if len(next_assigned) < needed and not task.get('is_past'):
            gap = needed - len(next_assigned)
            task['alert'] = f'Under-staffed: need {needed}, assigned {len(next_assigned)} (gap {gap})'
        else:
            task['alert'] = ''
        for sid in blocked:
            removed.add((task_id, sid))

    if not removed:
        return

    for staff in result.get('staff', []):
        sid = str(staff.get('id', ''))
        staff['assignments'] = [
            a for a in (staff.get('assignments') or [])
            if (a.get('task_id'), sid) not in removed
        ]

    for flight in result.get('flights', []):
        for task in flight.get('tasks', []):
            src = task_lookup.get(task.get('id'))
            if src is not None:
                task['assigned'] = list(src.get('assigned') or [])
                task['alert'] = src.get('alert', '')


def _count_staff_in_block(result, staff_id, skill, terminal, blk_start, blk_end):
    count = 0
    for task in result.get('tasks', []):
        if task.get('skill') != skill:
            continue
        if terminal != 'ALL' and task.get('terminal', 'ALL') != terminal:
            continue
        if not (blk_start <= (task.get('start_mins') or 0) < blk_end):
            continue
        if staff_id in [str(x) for x in (task.get('assigned') or [])]:
            count += 1
    return count


def _matching_block_tasks(result, skill, terminal, blk_start, blk_end):
    return [
        task for task in result.get('tasks', [])
        if task.get('skill') == skill
        and (terminal == 'ALL' or task.get('terminal', 'ALL') == terminal)
        and blk_start <= (task.get('start_mins') or 0) < blk_end
    ]


def _block_required_fte(tasks):
    slot_req = defaultdict(int)
    for task in tasks:
        slot_req[task.get('start_mins') or 0] += int(task.get('staff_needed') or 0)
    return max(slot_req.values()) if slot_req else 0


def _block_assigned_staff(tasks):
    assigned = set()
    for task in tasks:
        assigned.update(str(sid) for sid in (task.get('assigned') or []) if sid)
    return assigned


def _staff_has_task_skill(staff, task_skill):
    selected = TASK_SKILL.get(str(task_skill or '').strip(), str(task_skill or '').strip())
    staff_skills = {
        TASK_SKILL.get(str(staff.get(k, '')).strip(), str(staff.get(k, '')).strip())
        for k in ('skill1', 'skill2', 'skill3', 'skill4')
        if staff.get(k)
    }
    return selected in staff_skills


def _staff_shift_covers_block(staff, blk_start, blk_end):
    sh_start = staff.get('shift_start_mins', staff.get('shift_start', 0))
    sh_end = staff.get('shift_end_mins', staff.get('shift_end', 0))
    return sh_start <= blk_start and sh_end >= blk_end


def _staff_busy_outside_tasks(result, staff_id, excluded_task_ids, blk_start, blk_end):
    excluded = set(excluded_task_ids)
    busy = []
    for task in result.get('tasks', []):
        if task.get('id') in excluded:
            continue
        if not ((task.get('start_mins') or 0) < blk_end and (task.get('end_mins') or 0) > blk_start):
            continue
        if staff_id in [str(x) for x in (task.get('assigned') or [])]:
            busy.append(task)
    return busy


def _manual_block_key(skill, terminal, blk_start, blk_end):
    return f'{skill}|{terminal}|{int(blk_start)}|{int(blk_end)}'


def _sync_flight_tasks_from_task_lookup(result, task_lookup):
    for flight in result.get('flights', []):
        for task in flight.get('tasks', []):
            src = task_lookup.get(task.get('id'))
            if src is not None:
                task['assigned'] = list(src.get('assigned') or [])
                task['alert'] = src.get('alert', '')
                task['staff_needed'] = src.get('staff_needed', task.get('staff_needed', 1))


def _apply_manual_block_assigns_to_result(result, date_key):
    """Apply 3-hour reallocation-panel assignments as block-level crews."""
    blocks = _manual_block_assigns.get(date_key, {})
    if not blocks:
        return

    staff_by_id = {str(s.get('id')): s for s in result.get('staff', []) if s.get('id')}
    task_lookup = {t.get('id'): t for t in result.get('tasks', [])}

    for block in blocks.values():
        skill = block.get('skill', '')
        terminal = block.get('terminal', 'ALL')
        blk_start = int(block.get('block_start', 0))
        blk_end = int(block.get('block_end', 0))
        tasks = _matching_block_tasks(result, skill, terminal, blk_start, blk_end)
        if not tasks:
            continue

        required = _block_required_fte(tasks)
        if required <= 0:
            continue

        existing_counts = defaultdict(int)
        for task in tasks:
            for sid in task.get('assigned') or []:
                existing_counts[str(sid)] += 1

        preferred = []
        for sid in block.get('staff_ids') or []:
            sid = str(sid)
            staff = staff_by_id.get(sid)
            if not staff:
                continue
            if not _staff_has_task_skill(staff, skill):
                continue
            if not _staff_shift_covers_block(staff, blk_start, blk_end):
                continue
            if sid not in preferred:
                preferred.append(sid)

        carry_forward = [
            sid for sid, _count in sorted(existing_counts.items(), key=lambda item: (-item[1], item[0]))
            if sid not in preferred and sid in staff_by_id
        ]
        crew = (preferred + carry_forward)[:required]

        task_ids = {t.get('id') for t in tasks}
        for staff in staff_by_id.values():
            staff['assignments'] = [
                a for a in (staff.get('assignments') or [])
                if a.get('task_id') not in task_ids
            ]

        for task in tasks:
            task['assigned'] = list(crew)
            needed = int(task.get('staff_needed') or 0)
            if len(task['assigned']) >= needed:
                task['alert'] = None
            elif not task.get('is_past'):
                gap = needed - len(task['assigned'])
                task['alert'] = f'Under-staffed: need {needed}, assigned {len(task["assigned"])} (gap {gap})'
            for sid in crew:
                staff = staff_by_id.get(sid)
                if not staff:
                    continue
                staff.setdefault('assignments', []).append({
                    'task_id':    task.get('id'),
                    'task':       task.get('task'),
                    'skill':      task.get('skill', skill),
                    'terminal':   task.get('terminal', 'ALL'),
                    'start':      task.get('start'),
                    'end':        task.get('end'),
                    'start_mins': task.get('start_mins'),
                    'end_mins':   task.get('end_mins'),
                    'skill_mismatch': False,
                })

    _sync_flight_tasks_from_task_lookup(result, task_lookup)


def _staff_has_break_in_window(staff, start, end):
    return any(
        start < (br.get('end_mins') or 0) and end > (br.get('start_mins') or 0)
        for br in (staff.get('breaks') or [])
    )


def _set_task_status(task):
    needed = int(task.get('staff_needed') or 0)
    assigned_count = len(task.get('assigned') or [])
    if assigned_count >= needed:
        task['alert'] = None
    elif not task.get('is_past'):
        gap = needed - assigned_count
        task['alert'] = f'Under-staffed: need {needed}, assigned {assigned_count} (gap {gap})'


def _append_staff_assignment(staff, task):
    staff.setdefault('assignments', []).append({
        'task_id':    task.get('id'),
        'task':       task.get('task'),
        'skill':      task.get('skill', ''),
        'terminal':   task.get('terminal', 'ALL'),
        'start':      task.get('start'),
        'end':        task.get('end'),
        'start_mins': task.get('start_mins'),
        'end_mins':   task.get('end_mins'),
        'skill_mismatch': False,
    })


def _enforce_single_terminal_per_staff_hour(result):
    """Prevent one staff member from covering multiple terminals in one hour."""
    staff = result.get('staff', []) if isinstance(result, dict) else []
    tasks = result.get('tasks', []) if isinstance(result, dict) else []
    if not staff or not tasks:
        return

    task_lookup = {t.get('id'): t for t in tasks}
    period_terminal_by_staff = {}

    for task in sorted(tasks, key=lambda t: (
        t.get('start_mins', 0),
        0 if t.get('terminal', 'ALL') != 'ALL' else 1,
        t.get('priority', ''),
        t.get('skill', ''),
        t.get('terminal', 'ALL'),
    )):
        terminal = task.get('terminal', 'ALL')
        if terminal == 'ALL':
            continue

        start = int(task.get('start_mins') or 0)
        end = int(task.get('end_mins') or start)
        period_key = (start // 60) * 60
        kept = []
        for sid in task.get('assigned') or []:
            sid = str(sid)
            key = (sid, period_key)
            chosen_terminal = period_terminal_by_staff.get(key)
            if chosen_terminal is None:
                period_terminal_by_staff[key] = terminal
                kept.append(sid)
            elif chosen_terminal == terminal:
                kept.append(sid)

        max_staff = int(task.get('staff_needed') or len(kept) or 0)
        task['assigned'] = kept[:max_staff]
        _set_task_status(task)

    valid_task_ids_by_staff = defaultdict(set)
    for task in tasks:
        for sid in task.get('assigned') or []:
            valid_task_ids_by_staff[str(sid)].add(task.get('id'))

    for s in staff:
        sid = str(s.get('id', ''))
        s['assignments'] = [
            a for a in (s.get('assignments') or [])
            if a.get('task_id') in valid_task_ids_by_staff.get(sid, set())
        ]

    _sync_flight_tasks_from_task_lookup(result, task_lookup)


def _normalize_reallocation_blocks(result, date_key):
    """Make each terminal x skill x hourly block use one capped crew."""
    staff = result.get('staff', []) if isinstance(result, dict) else []
    tasks = result.get('tasks', []) if isinstance(result, dict) else []
    if not staff or not tasks:
        return

    staff_by_id = {str(s.get('id')): s for s in staff if s.get('id')}
    task_lookup = {t.get('id'): t for t in tasks}
    blocks = defaultdict(list)
    for task in tasks:
        skill = task.get('skill')
        if not skill:
            continue
        block_start = ((task.get('start_mins') or 0) // 60) * 60
        terminal = task.get('terminal', 'ALL')
        blocks[(block_start, block_start + 60, skill, terminal)].append(task)

    task_ids_in_blocks = {t.get('id') for group in blocks.values() for t in group}
    for s in staff:
        s['assignments'] = [
            a for a in (s.get('assignments') or [])
            if a.get('task_id') not in task_ids_in_blocks
        ]

    manual_blocks = _manual_block_assigns.get(date_key, {})
    blocked_by_task = _manual_unassigns.get(date_key, {})

    used_by_period = defaultdict(set)
    ordered_groups = []
    for (block_start, block_end, skill, terminal), group_tasks in blocks.items():
        required = _block_required_fte(group_tasks)
        current = len(_block_assigned_staff(group_tasks))
        gap = required - current
        ordered_groups.append((block_start, block_end, skill, terminal, group_tasks, required, gap))
    ordered_groups.sort(key=lambda item: (item[0], -item[6], item[2], item[3]))

    for block_start, block_end, skill, terminal, group_tasks, required, _gap in ordered_groups:
        if required <= 0:
            for task in group_tasks:
                task['assigned'] = []
                _set_task_status(task)
            continue

        period_key = (block_start, block_end)
        task_ids = {t.get('id') for t in group_tasks}
        blocked = {
            str(sid)
            for task in group_tasks
            for sid in (blocked_by_task.get(task.get('id'), []) or [])
        }

        existing_counts = defaultdict(int)
        for task in group_tasks:
            for sid in task.get('assigned') or []:
                existing_counts[str(sid)] += 1

        manual_preferred = []
        for block in manual_blocks.values():
            if block.get('skill') != skill:
                continue
            if block.get('terminal', 'ALL') != terminal:
                continue
            if int(block.get('block_start', -1)) != block_start or int(block.get('block_end', -1)) != block_end:
                continue
            for sid in block.get('staff_ids') or []:
                sid = str(sid)
                if sid not in manual_preferred:
                    manual_preferred.append(sid)

        candidates = []
        for sid in manual_preferred + [sid for sid, _ in sorted(existing_counts.items(), key=lambda x: (-x[1], x[0]))]:
            if sid in candidates:
                continue
            candidates.append(sid)
        for s in staff:
            sid = str(s.get('id', ''))
            if sid not in candidates:
                candidates.append(sid)

        crew = []
        for sid in candidates:
            if len(crew) >= required:
                break
            if sid in blocked or sid in used_by_period[period_key]:
                continue
            s = staff_by_id.get(sid)
            if not s:
                continue
            if not _staff_has_task_skill(s, skill):
                continue
            if not _staff_shift_covers_block(s, block_start, block_end):
                continue
            crew.append(sid)

        used_by_period[period_key].update(crew)

        for task in group_tasks:
            task['assigned'] = list(crew)
            _set_task_status(task)
            for sid in crew:
                s = staff_by_id.get(sid)
                if s:
                    _append_staff_assignment(s, task)

    for s in staff:
        seen = set()
        deduped = []
        for a in s.get('assignments') or []:
            key = (a.get('task_id'), a.get('start_mins'), a.get('end_mins'))
            if key in seen:
                continue
            seen.add(key)
            deduped.append(a)
        s['assignments'] = sorted(deduped, key=lambda a: (a.get('start_mins', 0), a.get('end_mins', 0)))
        shift_len = s.get('shift_end', 0) - s.get('shift_start', 0)
        total_busy = sum((a.get('end_mins') or 0) - (a.get('start_mins') or 0) for a in s['assignments'])
        s['utilisation_pct'] = round(min(total_busy / shift_len * 100, 100), 1) if shift_len > 0 else 0

    _sync_flight_tasks_from_task_lookup(result, task_lookup)
    _enforce_single_terminal_per_staff_hour(result)
_intraday_custom_constraints = {
    'permitted_shifts': [
        (0, 720, 'Day'),
        (720, 1440, 'Night')
    ],
    'use_cpsat': False,   # Set True to engage the CP-SAT optimiser
}

_st_custom_constraints = {
    'permitted_shifts': [
        (0, 720, 'Day'),
        (720, 1440, 'Night')
    ]
}

_st_custom_constraints_by_date = {}


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def read_csv_flights():
    """Read Flights_schedule_4days.csv with cp1252 encoding, strip \xa0 from all values."""
    path = os.path.join(BASE_DIR, 'data', 'Flights_schedule_4days.csv')
    stands_map = get_stands_map()
    with open(path, encoding='cp1252') as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            clean = {k: v.replace('\xa0', '').strip() for k, v in row.items()}
            if any(clean.values()):
                gate = clean.get('gate', '').strip()
                stand_info = stands_map.get(gate, {})
                # If stands map assigns this to the Cargo terminal, exclude it
                if stand_info.get('terminal', '').lower() == 'cargo':
                    continue
                rows.append(clean)
    return rows


def parse_time(s):
    """Convert 'H:MM' or 'HH:MM' to minutes since midnight. Return None if invalid."""
    if not s:
        return None
    s = s.strip().replace('\xa0', '')
    try:
        parts = s.split(':')
        if len(parts) != 2:
            return None
        h, m = int(parts[0]), int(parts[1])
        return h * 60 + m
    except (ValueError, AttributeError):
        return None


def mins_to_time(m):
    """Convert minutes to 'HH:MM' string, wrapping around midnight."""
    m = int(m) % 1440
    return f'{m // 60:02d}:{m % 60:02d}'


def icao_to_haul(icao_cat, cbp_flag):
    """Return haul type from ICAO category and CBP flag."""
    if str(cbp_flag).strip().upper() == 'TRUE':
        return 'US/Canada'
    cat = str(icao_cat).strip().upper()
    if cat in ('B', 'C'):
        return 'Short'
    if cat in ('D', 'E'):
        return 'Long'
    return 'Short'  # default


def load_config_rules():
    """Parse Config.csv (new multi-column format) and cache in _config_rules.

    Columns: Task, Terminal 1, Terminal 2, Pier 1-4, Priority,
             Short Haul, Long Haul, Arrival, Departure,
             Max Staff Count, Dependent on Flights/Terminal
    """
    global _config_rules
    if _config_rules is not None:
        return _config_rules
    path = os.path.join(BASE_DIR, 'data', 'Config.csv')
    with open(path, encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
    rules = []
    for r in rows:
        task = r.get('Task', '').strip()
        if not task:
            continue

        def _yes(col):
            return r.get(col, '').strip().lower() == 'yes'

        applicable_terminals = []
        if _yes('Terminal 1'): applicable_terminals.append('T1')
        if _yes('Terminal 2'): applicable_terminals.append('T2')

        applicable_piers = []
        if _yes('Pier 1'): applicable_piers.append('P1')
        if _yes('Pier 2'): applicable_piers.append('P2')
        if _yes('Pier 3'): applicable_piers.append('P3')
        if _yes('Pier 4'): applicable_piers.append('P4')

        count_raw = r.get('Max Staff Count', '1').strip()
        try:
            import math as _math
            max_staff_count = max(1, _math.ceil(float(count_raw)))
        except (ValueError, TypeError):
            max_staff_count = 1

        rules.append({
            'task':                   task,
            'applicable_terminals':   applicable_terminals,
            'applicable_piers':       applicable_piers,
            'priority':               r.get('Priority', 'Medium').strip(),
            'applies_to_short_haul':  _yes('Short Haul'),
            'applies_to_long_haul':   _yes('Long Haul'),
            'applies_to_arrivals':    _yes('Arrival'),
            'applies_to_departures':  _yes('Departure'),
            'max_staff_count':        max_staff_count,
            'scope':                  r.get('Dependent on Flights/Terminal', 'All Flights').strip(),
        })
    _config_rules = rules
    return _config_rules


def _infer_terminal(gate: str) -> str:
    """Best-effort terminal inference from a gate/stand identifier.

    Checks for explicit 'T2' token; everything else defaults to 'T1'.
    """
    g = gate.upper()
    if 'T2' in g:
        return 'T2'
    return 'T1'


def _infer_pier(gate: str) -> str:
    """Best-effort pier inference from a gate/stand identifier.

    Handles common naming conventions:
      • Explicit token  : 'P1-101', 'PIER2', 'T1-P3-22'
        → captures the single digit immediately after P/PIER
      • Three-digit gate: leading digit → pier  (e.g. '101' → P1, '204' → P2)
      • Single-letter   : 'A…' → P1, 'B…' → P2, 'C…' → P3, 'D…' → P4
    Defaults to 'P1' when no pattern matches.
    """
    g = gate.upper().replace(' ', '').replace('-', '')
    # Match 'P' or 'PIER' followed by a SINGLE digit (pier numbers are 1-9).
    # No lookahead needed — we only want the first digit after the P token.
    m = re.search(r'P(?:IER)?(\d)', g)
    if m:
        return f'P{m.group(1)}'
    # Three-or-more digit gate number: first digit is the pier
    m = re.match(r'[A-Z]*(\d)(\d{2,})', g)
    if m:
        return f'P{m.group(1)}'
    # Single letter prefix A-D maps to P1-P4
    m = re.match(r'([A-D])\d+', g)
    if m:
        return f'P{ord(m.group(1)) - ord("A") + 1}'
    return 'P1'


_PIER_NORM = {'1': 'P1', '2': 'P2', '3': 'P3', '4': 'P4'}

def _norm_pier_value(p: str) -> str:
    """Normalise Stands.csv pier values ('1'-'4') to 'P1'-'P4'.
    Non-numeric labels (North, Central …) are returned unchanged.
    """
    return _PIER_NORM.get(p, p)


def get_stands_map():
    """Return {stand_id: {'type': str, 'terminal': str, 'pier': str}}, cached.

    Reads Stands.csv.  Pier values are normalised: '1'→'P1', '2'→'P2', etc.
    so they match the 'P1'-'P4' format used in Config.csv rules.
    """
    global _stands_map
    if _stands_map is not None:
        return _stands_map
    path = os.path.join(BASE_DIR, 'data', 'Stands.csv')
    with open(path, encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
    _stands_map = {}
    for r in rows:
        sid   = r.get('stand_id', '').strip()
        stype = r.get('stand_type', '').strip() or 'Contact'
        term  = (r.get('terminal', '') or r.get('Terminal', '')).strip()
        pier  = (r.get('pier',     '') or r.get('Pier',     '')).strip()
        if sid:
            raw_pier = pier or _infer_pier(sid)
            _stands_map[sid] = {
                'type':     stype,
                'terminal': term or _infer_terminal(sid),
                'pier':     _norm_pier_value(raw_pier),
            }
    return _stands_map


def get_staff_for_date(date_str, custom_constraints=None, use_roster_optimiser=False,
                       demand_windows=None, prev_shift_ends=None):
    """Return (on_duty_list, absent_list) for the given date string.

    When use_roster_optimiser=True the two-phase roster engine is used to
    assign staff to optimal shift patterns instead of the density heuristic.
    The caller may supply pre-built demand_windows (list[DemandWindow]) and
    prev_shift_ends ({staff_id: shift_end_mins}) for the rest-rule check.
    Both default to empty / auto-derived when omitted.
    """
    if custom_constraints is None:
        custom_constraints = {}
    
    leave_types_excluded = custom_constraints.get('leave_types_excluded', ['Annual Leave', 'Paternity Leave', 'Jury Duty', 'Sick Leave', 'Training'])
    shift_duration_hrs = int(custom_constraints.get('shift_duration_hrs', 12))
    shift_duration_mins = shift_duration_hrs * 60
    # Parse date_str
    d = None
    for fmt in ('%Y-%m-%d', '%d-%b-%y', '%d-%m-%Y', '%d-%m-%y', '%d-%m-%Y'):
        try:
            d = datetime.strptime(date_str.strip(), fmt)
            break
        except ValueError:
            pass
    if d is None:
        return [], []

    # Staff CSV can use DD-MM-YYYY (4-digit) or DD-MM-YY (2-digit)
    staff_date_key_full = d.strftime('%d-%m-%Y') 
    staff_date_key_short = d.strftime('%d-%m-%y')

    # Load staff schedule
    path_staff = os.path.join(BASE_DIR, 'data', 'Staff_schedule.csv')
    with open(path_staff, encoding='utf-8-sig') as f:
        staff_rows = list(csv.DictReader(f))
    staff_rows = [{k: v.replace('\x00', '').strip() if isinstance(v, str) else v
                   for k, v in row.items()} for row in staff_rows]

    # Load absences
    path_abs = os.path.join(BASE_DIR, 'data', 'Staff_absence_schedule.csv')
    with open(path_abs, encoding='utf-8-sig') as f:
        abs_rows = list(csv.DictReader(f))
    abs_rows = [{k: v.replace('\x00', '').strip() if isinstance(v, str) else v
                 for k, v in row.items()} for row in abs_rows]

    # Build absent_set
    absent_set = {}
    for a in abs_rows:
        emp_id = a.get('EMPLOYEE NUMBER', '').strip()
        if not emp_id:
            continue
        d_from_raw = a.get('DATE FROM', '').strip()
        d_to_raw = a.get('DATE TO', '').strip()
        d_from = parse_date(d_from_raw)
        d_to = parse_date(d_to_raw)
        if d_from and d_to and d_from <= d <= d_to:
            leave_type = a.get('LEAVE TYPE', '').strip()
            if leave_type in leave_types_excluded:
                absent_set[emp_id] = leave_type

    # Filter staff for this date
    day_staff = [r for r in staff_rows
                 if (r.get('DATE', '').strip() == staff_date_key_full or r.get('DATE', '').strip() == staff_date_key_short)
                 and r.get('EMPLOYEE NUMBER', '').strip()]

    # Carry-forward fallback: if no staff data for requested date, use the
    # most recent available date in the CSV so D+3 still renders meaningfully.
    _carried_forward_from = None
    if not day_staff:
        # Collect all distinct dates present in the CSV
        known_dates = []
        for fmt in ('%d-%m-%Y', '%d-%m-%y', '%d-%m'):
            for r in staff_rows:
                raw = r.get('DATE', '').strip()
                try:
                    pd = datetime.strptime(raw, fmt)
                    if fmt == '%d-%m':
                        pd = pd.replace(year=d.year)
                    if pd < d:
                        known_dates.append(pd)
                except ValueError:
                    pass
        if known_dates:
            fallback_d = max(known_dates)
            fb_full  = fallback_d.strftime('%d-%m-%Y')
            fb_short = fallback_d.strftime('%d-%m-%y')
            day_staff = [r for r in staff_rows
                         if (r.get('DATE', '').strip() == fb_full or r.get('DATE', '').strip() == fb_short)
                         and r.get('EMPLOYEE NUMBER', '').strip()]
            _carried_forward_from = fallback_d.strftime('%d %b %Y')

    on_duty = []
    absent_staff = []

    # Shift duration in minutes
    shift_duration_hrs = int(custom_constraints.get('shift_duration_hrs', 12))
    shift_duration_mins = shift_duration_hrs * 60

    # Permitted shifts from constraints
    sh_options = custom_constraints.get('permitted_shifts')

    # ── Density-based shift allocation ──────────────────────────────────────
    # Only allocate shifts WHEN THERE ARE FLIGHTS.
    # Empty 30-minute windows are excluded from candidates entirely.
    # Busier windows attract proportionally more staff.
    density_shifts = None
    if not sh_options:
        try:
            flt_path     = os.path.join(BASE_DIR, 'data', 'Flights_schedule_4days.csv')
            flt_date_key = d.strftime('%d-%b-%y')
            flight_times = []
            with open(flt_path, encoding='cp1252') as ff:
                for row in csv.DictReader(ff):
                    clean = {k: v.replace('\xa0', '').strip() for k, v in row.items()}
                    if clean.get('date', '') == flt_date_key:
                        t = parse_time(clean.get('sta', ''))
                        if t is not None:
                            flight_times.append(t)

            # 30-minute blocks: 0..47
            # block_density now stores WEIGHTED demand (proxy for FTE)
            block_density = [0.0] * 48
            with open(flt_path, encoding='cp1252') as ff:
                for row in csv.DictReader(ff):
                    clean = {k: v.replace('\xa0', '').strip() for k, v in row.items()}
                    if clean.get('date', '') == flt_date_key:
                        t = parse_time(clean.get('sta', ''))
                        if t is not None:
                            # Proxy for task demand: higher weight for larger aircraft categories
                            # Cat C (~1.0), Cat E (~2.0), Cat F (~3.0)
                            cat = clean.get('icao_cat', 'C').upper()
                            weight = 1.0
                            if cat == 'D':   weight = 1.5
                            elif cat == 'E': weight = 2.0
                            elif cat == 'F': weight = 3.0
                            
                            block_density[min(47, int(t) // TASK_SLOT_MINS)] += weight

            # Find the span of active blocks
            active_blocks = [bi for bi, cnt in enumerate(block_density) if cnt > 0]

            if active_blocks:
                # Build candidates ONLY for requested start timings
                # Mandatory shift start timings: 00:00, 3:00, 7:00, 12:00
                requested_starts = [0, 180, 420, 720]
                candidates = []
                for s_start in requested_starts:
                    s_end = min(1440, s_start + shift_duration_mins)
                    
                    # Coverage = weighted flights in the duration of this shift
                    bi_start = s_start // TASK_SLOT_MINS
                    bi_end = (s_end - 1) // TASK_SLOT_MINS
                    coverage = sum(block_density[j] for j in range(bi_start, min(48, bi_end + 1)))
                    
                    if coverage == 0:
                        continue
                        
                    candidates.append({
                        'start':   s_start,
                        'end':     s_end,
                        'label':   f"{mins_to_time(s_start)}\u2013{mins_to_time(s_end)}",
                        'density': coverage,
                    })

                if candidates:
                    total_density = sum(c['density'] for c in candidates)
                    n_staff_count = len(day_staff)
                    density_shifts = []
                    # Proportionally allocate staff based on weighted demand
                    for cand in sorted(candidates, key=lambda x: -x['density']):
                        n_alloc = max(0, round(n_staff_count * cand['density'] / total_density))
                        for _ in range(n_alloc):
                            density_shifts.append((cand['start'], cand['end'], cand['label']))

                    # Pad/Trim to exactly n_staff_count (use busiest shift)
                    best = max(candidates, key=lambda x: x['density'])
                    while len(density_shifts) < n_staff_count:
                        density_shifts.append((best['start'], best['end'], best['label']))
                    density_shifts = density_shifts[:n_staff_count]
        except Exception:
            density_shifts = None


    # Current Day
    mezz_shifts_map = {}
    mezz_staff_list = [
        r for r in day_staff
        if 'Mezz Operation' in [
            r.get('Skill1', '').strip(),
            r.get('Skill2', '').strip(),
            r.get('Skill3', '').strip(),
            r.get('Skill4', '').strip()
        ]
        and r.get('EMPLOYEE NUMBER', '').strip() not in absent_set
    ]
    if mezz_staff_list:
        # Mandatory 24-hour coverage anchors (hard constraint):
        #   Worker 0 → 00:00-12:00 (covers midnight-to-noon)
        #   Worker 1 → 12:00-24:00 (covers noon-to-midnight)
        # Additional Mezz workers may overlap freely with any start time.
        mezz_coverage_anchors = [
            (0,   720,  '00:00'),  # anchor 1: must-have for 0:00-12:00 coverage
            (720, 1440, '12:00'),  # anchor 2: must-have for 12:00-24:00 coverage
        ]
        mezz_extra_starts = [
            (0,   720,  '00:00'),
            (180, 900,  '03:00'),
            (420, 1140, '07:00'),
            (720, 1440, '12:00'),
            (360, 1080, '06:00'),
        ]
        for idx, r_mezz in enumerate(mezz_staff_list):
            if idx < len(mezz_coverage_anchors):
                # First two workers must fill the 24h coverage anchors
                st_m, en_m, lb_m = mezz_coverage_anchors[idx]
            else:
                # Additional workers overlap freely across all start options
                extra_idx = (idx - len(mezz_coverage_anchors)) % len(mezz_extra_starts)
                st_m, en_m, lb_m = mezz_extra_starts[extra_idx]
            mezz_shifts_map[r_mezz.get('EMPLOYEE NUMBER', '').strip()] = (st_m, en_m, lb_m)

    for i, r in enumerate(day_staff):
        emp_id = r.get('EMPLOYEE NUMBER', '').strip()
        skill1 = r.get('Skill1', '').strip()
        skill2 = r.get('Skill2', '').strip()
        skill3 = r.get('Skill3', '').strip()
        skill4 = r.get('Skill4', '').strip()
        employment = r.get('EMPLOYMENT TYPE', '').strip()

        if emp_id in absent_set:
            absent_staff.append({'id': emp_id, 'skill1': skill1, 'leave_type': absent_set[emp_id], 'absent': True})
            continue

        if emp_id in mezz_shifts_map:
            st, en, lb = mezz_shifts_map[emp_id]
        elif sh_options and len(sh_options) > 0:
            sh_data = sh_options[i % len(sh_options)]
            st, en, lb = sh_data[0], sh_data[1], sh_data[2]
        elif density_shifts:
            # Use the density-optimised shift for this staff member's position
            idx = min(i, len(density_shifts) - 1)
            st, en, lb = density_shifts[idx]
        else:
            # Hard fallback: round-robin across the requested shift starts
            sys_defaults = [
                (0,   720,  '00:00'),
                (180, 900,  '03:00'),
                (420, 1140, '07:00'),
                (720, 1440, '12:00'),
            ]
            st, en, lb = sys_defaults[i % len(sys_defaults)]


        _shift_name = {0:'Early',360:'Mid',720:'Late',960:'Evening',1320:'Night'}.get(st, lb.upper().replace(' SHIFT',''))
        on_duty.append({
            'id': emp_id,
            'skill1': skill1, 'skill2': skill2, 'skill3': skill3, 'skill4': skill4,
            'employment': employment, 'shift': _shift_name,
            'shift_start': st, 'shift_end': en,
            'shift_label': f"{_shift_name} {mins_to_time(st)}–{mins_to_time(en)}",
            'assignments': [], 'breaks': [], 'utilisation_pct': 0
        })

    # ── Roster optimiser path ────────────────────────────────────────────────
    # Replace the density-heuristic shift labels with optimiser-assigned patterns.
    if use_roster_optimiser and _ROSTER_AVAILABLE and on_duty:
        try:
            dw_list  = demand_windows or []
            pe_map   = prev_shift_ends or {}
            use_mip  = custom_constraints.get('use_mip', True)

            permitted_shifts = custom_constraints.get('permitted_shifts')
            if permitted_shifts:
                permitted_starts = [int(p[0]) for p in permitted_shifts]
            else:
                permitted_starts = [0, 180, 420, 720]

            roster_result = _roster_generate(
                demand_windows  = dw_list,
                staff_list      = [
                    {
                        'id':         s['id'],
                        'skill1':     s.get('skill1', ''),
                        'skill2':     s.get('skill2', ''),
                        'skill3':     s.get('skill3', ''),
                        'skill4':     s.get('skill4', ''),
                        'employment': s.get('employment', ''),
                    }
                    for s in on_duty
                ],
                constraints     = {
                    'b1_duration_mins': custom_constraints.get('b1_duration_mins', 30),
                    'b2_duration_mins': custom_constraints.get('b2_duration_mins', 60),
                    'max_shift_mins':   custom_constraints.get('shift_duration_hrs', 12) * 60,
                    'min_rest_mins':    custom_constraints.get('min_rest_mins', 660),
                    'permitted_starts': permitted_starts,
                },
                prev_shift_ends = pe_map,
                use_mip         = use_mip,
            )

            # Merge optimiser assignments back onto the on_duty list.
            # Mezz Operation staff keep their 24h-coverage-anchored shifts
            # (set above in mezz_shifts_map) and are NOT overridden by the
            # roster optimizer, which assigns generic demand-driven patterns.
            # This preserves the hard 0:00-24:00 coverage guarantee for Mezz.
            optimised_lookup = {e['id']: e for e in roster_result.get('roster', [])}
            merged = []
            for s in on_duty:
                is_mezz = 'Mezz Operation' in [s.get('skill1', ''), s.get('skill2', ''), s.get('skill3', ''), s.get('skill4', '')]
                if is_mezz:
                    # Mezz workers always keep their 24h-anchored shifts;
                    # overlapping workers are also kept as-is (no dedup).
                    merged.append(s)
                    continue

                oe = optimised_lookup.get(s['id'])
                if oe and oe.get('pattern_id') != 'unassigned':
                    s = dict(s)
                    _oe_st = oe.get('shift_start', 0)
                    _oe_name = {0:'Early',360:'Mid',720:'Late',960:'Evening',1320:'Night'}.get(_oe_st, oe.get('pattern_id', oe['shift_label']))
                    s['shift']          = _oe_name
                    s['shift_start']    = _oe_st
                    s['shift_end']      = oe['shift_end']
                    s['shift_label']    = f"{_oe_name} {mins_to_time(_oe_st)}–{mins_to_time(oe['shift_end'])}"
                    s['breaks']         = oe.get('breaks', [])
                    s['utilisation_pct'] = oe.get('utilisation_pct', 0)
                    s['pattern_id']     = oe.get('pattern_id', '')
                    s['skill_match']    = oe.get('skill_match', '')
                merged.append(s)
            on_duty = merged

            # Attach top-level roster metadata for callers that inspect it
            on_duty._roster_meta = roster_result  # type: ignore[attr-defined]
        except Exception as _exc:
            import logging
            logging.getLogger(__name__).warning(
                "Roster optimiser failed — falling back to density heuristic: %s", _exc
            )

    return on_duty, absent_staff, _carried_forward_from


def schedule_breaks(staff, assigned_windows, custom_constraints=None):
    """Return mandatory shift-clock breaks for one staff member."""
    shift_start = staff['shift_start']
    shift_end   = staff['shift_end']

    if custom_constraints is None:
        custom_constraints = {}

    b1_dur     = int(custom_constraints.get('b1_duration_mins', 30))
    b2_dur     = int(custom_constraints.get('b2_duration_mins', 60))
    delay      = int(staff.get('_break_delay_mins', 0))
    work_limit = int(custom_constraints.get('break_after_work_mins', 180))
    break_gap  = int(custom_constraints.get('break_gap_after_b1_mins', 180))

    b1_start = shift_start + work_limit + delay
    b1_end   = b1_start + b1_dur
    b2_start = b1_end + break_gap
    b2_end   = b2_start + b2_dur

    # Use absolute end (shift_start + duration) so night shifts crossing midnight
    # are handled correctly — raw shift_end can be < shift_start for night workers.
    shift_dur     = (shift_end - shift_start) % 1440 or 1440
    shift_abs_end = shift_start + shift_dur
    if b2_end > shift_abs_end:
        return []

    return [
        {
            'start_mins': b1_start,
            'end_mins':   b1_end,
            'start':      mins_to_time(b1_start),
            'end':        mins_to_time(b1_end),
            'type':       'Short Break',
        },
        {
            'start_mins': b2_start,
            'end_mins':   b2_end,
            'start':      mins_to_time(b2_start),
            'end':        mins_to_time(b2_end),
            'type':       'Meal Break',
        },
    ]

    b1_dur          = int(custom_constraints.get('b1_duration_mins', 30))
    b2_dur          = int(custom_constraints.get('b2_duration_mins', 60))
    break_durations = [b1_dur, b2_dur]
    break_types     = ['Short Break', 'Meal Break']
    MAX_BREAKS      = 2
    WORK_LIMIT      = 180   # continuous work limit before break is mandatory
    BREAK_GAP       = 180   # minimum gap between end of one break and start of next

    # ─ Helper ───────────────────────────────────────────────────
    def find_free_slot(duration, search_start, search_end, busy_sorted):
        """Return start of first contiguous free window of `duration` mins, or None."""
        t       = max(search_start, shift_start)
        end_cap = min(search_end, shift_end - duration)
        while t + duration <= end_cap:
            conflict = False
            for (ws, we) in busy_sorted:
                if t < we and t + duration > ws:
                    conflict = True
                    t = max(t + 1, we)
                    break
            if not conflict:
                return t
        return None

    busy = sorted(assigned_windows)

    # ─ Step 1: Detect trigger points from continuous-work analysis ──────────
    # Walk the sorted busy windows; accumulate work time and record a trigger
    # each time we cross WORK_LIMIT minutes of uninterrupted work.
    trigger_points = []
    cum_work   = 0
    last_event = shift_start

    for (ws, we) in busy:
        ws = max(ws, shift_start)
        we = min(we, shift_end)
        if we <= ws:
            continue
        if ws > last_event:          # idle gap → reset continuous counter
            cum_work = 0
        cum_work  += we - ws
        if cum_work >= WORK_LIMIT:
            trigger_points.append(we)
            cum_work = 0
        last_event = we

    # Fallback: no triggers but shift is long → place one midshift trigger
    shift_dur = (shift_end - shift_start) % 1440 or 1440
    if not trigger_points and shift_dur >= WORK_LIMIT:
        trigger_points.append(shift_start + WORK_LIMIT)

    trigger_points = trigger_points[:MAX_BREAKS]

    # ─ Step 2: Place breaks, enforcing the 3-hour inter-break gap ──────────
    breaks       = []
    earliest_b2  = None   # will be set to break1_end + BREAK_GAP after B1

    for i, trigger in enumerate(trigger_points):
        if len(breaks) >= MAX_BREAKS:
            break

        dur   = break_durations[i] if i < len(break_durations) else b1_dur
        btype = break_types[i]     if i < len(break_types)     else 'Break'

        # Include previously placed breaks in busy list so we avoid overlap
        cur_busy = sorted(busy + [(b['start_mins'], b['end_mins']) for b in breaks])

        # For break-2 onwards, the search must not start before earliest_b2
        search_from = trigger
        if i > 0 and earliest_b2 is not None:
            search_from = max(trigger, earliest_b2)

        # Earliest the FIRST break may start: at least 3 hours into the shift.
        # All subsequent breaks already respect earliest_b2 (break_end + 180).
        earliest_b1 = shift_start + WORK_LIMIT
        search_from = max(search_from, earliest_b1) if i == 0 else search_from

        # Try right after trigger (within 90-min window), then anywhere in shift
        b_start = (
            find_free_slot(dur, search_from, search_from + 90, cur_busy)
            or find_free_slot(dur, search_from, shift_end - dur, cur_busy)
            or (find_free_slot(dur, earliest_b1, shift_end - dur, cur_busy) if i == 0 else None)
        )

        if b_start is not None:
            b_end = b_start + dur
            breaks.append({
                'start_mins': b_start,
                'end_mins':   b_end,
                'start':      mins_to_time(b_start),
                'end':        mins_to_time(b_end),
                'type':       btype,
            })
            # Next break may not start before break_end + BREAK_GAP
            earliest_b2 = b_end + BREAK_GAP

    return sorted(breaks, key=lambda b: b['start_mins'])


def assign_stagger_breaks(staff_list, custom_constraints=None):
    """Assign fixed staggered break templates (Group A / Group B) to all staff.

    Staff within each shift are split 50/50: first half → Group A, second half → Group B.
    Shifts not matching a fixed template fall back to dynamic schedule_breaks with a
    60-minute stagger for Group B members.

    Side-effect: sets s['break_group'] = 'A' or 'B' on every staff dict.
    Returns {staff_id: [break_dict, ...]} matching the existing breaks API shape.
    """
    from copy import deepcopy
    if custom_constraints is None:
        custom_constraints = {}

    planned = {}
    groups = defaultdict(list)
    for s in staff_list:
        key = (s.get('shift_start', 0), s.get('shift_end', 720))
        groups[key].append(s)

    for (sh_start, sh_end), members in groups.items():
        ordered = sorted(members, key=lambda s: str(s.get('id', s.get('name', ''))))
        template = STAGGER_BREAK_TEMPLATES.get((sh_start, sh_end))
        half = (len(ordered) + 1) // 2  # Group A gets the larger half when odd

        if template:
            for i, s in enumerate(ordered):
                grp = 'A' if i < half else 'B'
                s['break_group'] = grp
                sid = str(s.get('id', s.get('name', '')))
                planned[sid] = deepcopy(template[grp])
        else:
            # Fallback: dynamic scheduling with 60-min stagger for second half
            delayed_ids = {
                str(s.get('id', s.get('name', '')))
                for s in ordered[half:]
            }
            for i, s in enumerate(ordered):
                grp = 'A' if i < half else 'B'
                s['break_group'] = grp
                sid = str(s.get('id', s.get('name', '')))
                s['_break_delay_mins'] = 60 if sid in delayed_ids else 0
                planned[sid] = schedule_breaks(s, [], custom_constraints)

    return planned


# Keep old name as alias so any external callers continue to work
def schedule_staff_breaks(staff_list, custom_constraints=None):
    return assign_stagger_breaks(staff_list, custom_constraints)


def enforce_break_conflicts(result):
    """Remove staff task assignments that overlap their break windows."""
    staff = result.get('staff', []) if isinstance(result, dict) else []
    if not staff:
        return result

    removed = set()
    for s in staff:
        sid = s.get('id')
        if not sid:
            continue
        breaks = s.get('breaks') or []
        kept = []
        for a in s.get('assignments') or []:
            if a.get('task') == 'Mezz Operation' or a.get('skill') == 'Mezz Operation':
                kept.append(a)
                continue
            overlaps_break = any(
                a.get('start_mins', 0) < b.get('end_mins', b.get('end', 0))
                and a.get('end_mins', 0) > b.get('start_mins', b.get('start', 0))
                for b in breaks
                if isinstance(b.get('start_mins'), int) and isinstance(b.get('end_mins'), int)
            )
            if overlaps_break:
                removed.add((a.get('task_id'), sid))
            else:
                kept.append(a)
        s['assignments'] = kept

        shift_len = s.get('shift_end', 0) - s.get('shift_start', 0)
        total_busy = sum(a.get('end_mins', 0) - a.get('start_mins', 0) for a in kept)
        s['utilisation_pct'] = round(min(total_busy / shift_len * 100, 100), 1) if shift_len > 0 else 0

    if not removed:
        return result

    for task in result.get('tasks', []) or []:
        assigned = task.get('assigned') or []
        new_assigned = [
            sid for sid in assigned
            if (task.get('id'), sid) not in removed
        ]
        if len(new_assigned) != len(assigned):
            task['assigned'] = new_assigned
            needed = task.get('staff_needed', 1)
            if len(new_assigned) < needed and not task.get('is_past'):
                task['alert'] = (
                    f'Under-staffed: need {needed}, assigned {len(new_assigned)} '
                    f'(gap {needed - len(new_assigned)})'
                )

    for flight in result.get('flights', []) or []:
        for task in flight.get('tasks', []) or []:
            assigned = task.get('assigned') or []
            new_assigned = [
                sid for sid in assigned
                if (task.get('id'), sid) not in removed
            ]
            if len(new_assigned) != len(assigned):
                task['assigned'] = new_assigned
                needed = task.get('staff_needed', 1)
                if len(new_assigned) < needed and not task.get('is_past'):
                    task['alert'] = (
                        f'Under-staffed: need {needed}, assigned {len(new_assigned)} '
                        f'(gap {needed - len(new_assigned)})'
                    )

    result['break_conflict_removals'] = len(removed)
    return result


def enforce_fte_shift_containment(result):
    """Ensure FTE staff do not have any assignments that fall outside their shift."""
    staff = result.get('staff', []) if isinstance(result, dict) else []
    tasks = result.get('tasks', []) if isinstance(result, dict) else []
    if not staff or not tasks:
        return result

    for s in staff:
        is_fte = s.get('employment', '').strip().lower() in ('full-time', 'full time', 'fte', 'permanent')
        if not is_fte:
            continue
        
        sh_start = s.get('shift_start', 0)
        sh_end = s.get('shift_end', 0)
        D_shift = (sh_end - sh_start) % 1440
        if D_shift == 0 and sh_end == sh_start:
            D_shift = 1440

        valid_assignments = []
        for a in s.get('assignments') or []:
            t_start = a.get('start_mins', 0)
            t_end = a.get('end_mins', 0)
            task_dur = (t_end - t_start) % 1440
            if task_dur == 0 and t_end != t_start:
                task_dur = 1440
            ts_rel = (t_start - sh_start) % 1440

            if ts_rel + task_dur > D_shift:
                # Task falls outside the shift of the FTE staff member
                for task in tasks:
                    if task.get('id') == a.get('task_id'):
                        if s['id'] in task.get('assigned', []):
                            task['assigned'].remove(s['id'])
            else:
                valid_assignments.append(a)

        s['assignments'] = valid_assignments

    return result


def refill_unassigned_tasks(result):
    """Fill task gaps after optimized shifts/breaks have been merged."""
    staff = result.get('staff', []) if isinstance(result, dict) else []
    tasks = result.get('tasks', []) if isinstance(result, dict) else []
    if not staff or not tasks:
        return result

    staff_by_id = {s.get('id'): s for s in staff if s.get('id')}
    priority_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}

    for s in staff:
        s['assignments'] = sorted(s.get('assignments') or [], key=lambda a: (a.get('start_mins', 0), a.get('end_mins', 0)))

    def has_skill(s, skill):
        needed = TASK_SKILL.get(str(skill).strip(), str(skill).strip())
        return any(
            TASK_SKILL.get(str(s.get(f'skill{k}', '')).strip(), str(s.get(f'skill{k}', '')).strip()) == needed
            for k in range(1, 5)
            if s.get(f'skill{k}', '')
        )

    def can_take(s, task):
        start = task.get('start_mins', 0)
        end = task.get('end_mins', 0)
        shift_start = s.get('shift_start', 0)
        shift_end = s.get('shift_end', 0)
        D_shift = (shift_end - shift_start) % 1440
        if D_shift == 0 and shift_end == shift_start:
            D_shift = 1440
        task_dur = (end - start) % 1440
        if task_dur == 0 and end != start:
            task_dur = 1440
        ts_rel = (start - shift_start) % 1440
        if ts_rel + task_dur > D_shift:
            return False
        if not task.get('post_coverage', False):
            for b in s.get('breaks') or []:
                if start < b.get('end_mins', 0) and end > b.get('start_mins', 0):
                    return False
        for a in s.get('assignments') or []:
            if start < a.get('end_mins', 0) and end > a.get('start_mins', 0):
                return False
        return True

    refilled = 0
    for task in sorted(tasks, key=lambda t: (
        0 if t.get('sharing_mode') == 'fixed' else 1,
        priority_order.get(t.get('priority', 'Medium'), 2),
        t.get('start_mins', 0),
    )):
        if task.get('is_past'):
            continue
        # Re-evaluate existing assignments: drop any staff who are no longer
        # able to cover the task due to a shift change (or missing staff).
        orig_assigned = list(dict.fromkeys(task.get('assigned') or []))
        assigned = []
        for sid in orig_assigned:
            s = staff_by_id.get(sid)
            if not s:
                # Staff not present for this date
                continue
            # If the staff's shift no longer covers the task window, remove the
            # assignment from both the task and the staff member.
            t_start = task.get('start_mins', 0)
            t_end = task.get('end_mins', 0)
            sh_start = s.get('shift_start', 0)
            sh_end = s.get('shift_end', 0)
            D_shift = (sh_end - sh_start) % 1440
            if D_shift == 0 and sh_end == sh_start:
                D_shift = 1440
            task_dur = (t_end - t_start) % 1440
            if task_dur == 0 and t_end != t_start:
                task_dur = 1440
            ts_rel = (t_start - sh_start) % 1440
            if ts_rel + task_dur > D_shift:
                # Remove matching assignment entries from the staff record
                s['assignments'] = [a for a in (s.get('assignments') or []) if a.get('task_id') != task.get('id')]
                continue
            assigned.append(sid)
        max_staff = task.get('staff_needed', 1)  # cap at requirement, not excess capacity
        if len(assigned) > max_staff:
            assigned = assigned[:max_staff]
        task['assigned'] = assigned
        needed = task.get('staff_needed', 1) - len(assigned)
        if needed <= 0:
            task['alert'] = None
            continue

        skill = task.get('skill', '')
        # Strict skill enforcement: only consider staff who actually have the
        # required skill. Do not fall back to assigning unskilled staff here.
        skill_candidates = [
            s for s in staff
            if s.get('id') not in assigned and has_skill(s, skill) and can_take(s, task)
        ]

        for s in skill_candidates:
            if needed <= 0:
                break
            sid = s.get('id')
            assignment = {
                'task_id':    task.get('id'),
                'task':       task.get('task'),
                'skill':      skill,
                'terminal':   task.get('terminal', 'ALL'),
                'start':      task.get('start'),
                'end':        task.get('end'),
                'start_mins': task.get('start_mins'),
                'end_mins':   task.get('end_mins'),
                'skill_mismatch': False,
            }
            assigned.append(sid)
            s.setdefault('assignments', []).append(assignment)
            needed -= 1
            refilled += 1

        if needed <= 0:
            task['alert'] = None
        else:
            total = task.get('staff_needed', 1)
            task['alert'] = f'Under-staffed: need {total}, assigned {len(task["assigned"])} (gap {needed})'

    for s in staff:
        seen_assignments = set()
        deduped = []
        for a in s.get('assignments') or []:
            key = (a.get('task_id'), a.get('start_mins'), a.get('end_mins'))
            if key in seen_assignments:
                continue
            seen_assignments.add(key)
            deduped.append(a)
        s['assignments'] = sorted(deduped, key=lambda a: (a.get('start_mins', 0), a.get('end_mins', 0)))
        shift_len = s.get('shift_end', 0) - s.get('shift_start', 0)
        total_busy = sum(a.get('end_mins', 0) - a.get('start_mins', 0) for a in s['assignments'])
        s['utilisation_pct'] = round(min(total_busy / shift_len * 100, 100), 1) if shift_len > 0 else 0

    task_by_id = {t.get('id'): t for t in tasks}
    for flight in result.get('flights', []) or []:
        for ft in flight.get('tasks', []) or []:
            src = task_by_id.get(ft.get('id'))
            if not src:
                continue
            ft['assigned'] = list(src.get('assigned') or [])
            ft['alert'] = src.get('alert')
            ft['staff_needed'] = src.get('staff_needed', ft.get('staff_needed', 1))

    result['refilled_assignments'] = result.get('refilled_assignments', 0) + refilled
    return result



# ---------------------------------------------------------------------------
# Task-generation helper — applies sharing logic before greedy assignment
# ---------------------------------------------------------------------------

def _stand_info(gate: str, stands_map: dict) -> dict:
    """Return the stand info dict for a gate, with safe defaults."""
    info = stands_map.get(gate)
    if isinstance(info, dict):
        return info
    # Fallback: old-format string value or missing entry
    stype = info if isinstance(info, str) else 'Contact'
    return {
        'type':     stype,
        'terminal': _infer_terminal(gate),
        'pier':     _infer_pier(gate),
    }


def _pax_for_icao(icao_cat: str) -> int:
    """Estimate pax count from ICAO wake-turbulence category."""
    return PAX_BY_ICAO.get(icao_cat.strip().upper(), PAX_DEFAULT)


def _partial_staff_count(total_pax: int) -> int:
    """Return staff count for a partially-shared task given combined pax volume."""
    for pax_lo, pax_hi, count in PARTIAL_SHARE_PAX_THRESHOLDS:
        if pax_lo <= total_pax < pax_hi:
            return count
    return PARTIAL_SHARE_PAX_THRESHOLDS[-1][2]


def _generate_day_tasks(processed_flights: list, rules: list, stands_map: dict,
                        window_mins: int = SHARING_WINDOW_MINS) -> list:
    """Generate all per-day tasks driven by Config.csv rules.

    Each rule carries: task, applicable_terminals, applicable_piers, scope,
    applies_to_short_haul, applies_to_long_haul, applies_to_arrivals,
    applies_to_departures, max_staff_count, priority.

    Scope:
      'Terminal'    -> one pooled 30-minute block per (terminal, block_idx, direction)
      'All Flights' -> one pooled 30-minute block per (terminal, pier, block_idx, direction)
      'US Flights'  -> like 'All Flights' but haul must be 'US/Canada'
      'Fixed'       -> handled in optimize_day() -- skipped here
    """
    import math as _math

    # Pax-based staffing ratios: {task_name: (pax_per_staff, min_staff)}
    PAX_RATIOS = {
        "GNIB":                 (300, 2),
        "T1/T2 Trolleys L/UL":  (500, 1),
        "Dep/Trolleys":         (500, 1),
        "Dep / Trolleys":       (500, 1),
        "Check-in/Trolleys":    (400, 1),
        "Arr Customer Service": (400, 1),
        "Transfer Corridor":    (500, 1),
        "PBZ":                  (400, 1),
        "CBP Pre-clearance":    (300, 1),
    }

    # terminal_blocks key: (terminal, block_idx, direction, task_name)
    terminal_blocks = defaultdict(lambda: {"pax": 0, "flights": [], "rule": None})
    # flight_blocks key:  (terminal, pier, block_idx, direction, task_name)
    flight_blocks   = defaultdict(lambda: {"pax": 0, "flights": [], "rule": None})

    for flight in processed_flights:
        fn        = flight["flight_no"]
        t_mins    = flight["time_mins"]
        status    = flight["status"]
        haul      = flight["haul"]
        gate      = flight["gate"]
        icao_cat  = flight["icao_cat"]

        si        = _stand_info(gate, stands_map)
        terminal  = si["terminal"]
        pier      = si["pier"]

        block_idx = t_mins // TASK_SLOT_MINS
        direction = "ARR" if status == "Arrival" else "DEP"
        pax       = _pax_for_icao(icao_cat)

        is_short  = haul == "Short"
        is_long   = haul in ("Long", "US/Canada")
        is_us     = haul == "US/Canada"

        for rule in rules:
            scope = rule["scope"]
            if scope == "Fixed":
                continue

            if terminal not in rule["applicable_terminals"]:
                continue
            if pier not in rule["applicable_piers"]:
                continue

            if status == "Arrival"   and not rule["applies_to_arrivals"]:
                continue
            if status == "Departure" and not rule["applies_to_departures"]:
                continue

            haul_flagged = rule["applies_to_short_haul"] or rule["applies_to_long_haul"]
            if haul_flagged:
                if not ((rule["applies_to_short_haul"] and is_short) or
                        (rule["applies_to_long_haul"]  and is_long)):
                    continue

            if scope == "US Flights" and not is_us:
                continue

            task_name = rule["task"]

            if scope == "Terminal":
                key = (terminal, block_idx, direction, task_name)
                d = terminal_blocks[key]
                d["pax"] += pax
                if fn not in d["flights"]:
                    d["flights"].append(fn)
                d["rule"] = rule
            else:
                key = (terminal, pier, block_idx, direction, task_name)
                d = flight_blocks[key]
                d["pax"] += pax
                if fn not in d["flights"]:
                    d["flights"].append(fn)
                d["rule"] = rule

    all_tasks = []

    # Terminal-scope block tasks
    for (terminal, block_idx, direction, task_name), data in terminal_blocks.items():
        if not data["flights"]:
            continue
        rule      = data["rule"]
        total_pax = data["pax"]
        fns       = data["flights"]
        blk_start = block_idx * TASK_SLOT_MINS
        blk_end   = min(1440, blk_start + TASK_SLOT_MINS)
        skill     = TASK_SKILL.get(task_name, "GNIB")

        if task_name in PAX_RATIOS and total_pax > 0:
            ratio, min_s = PAX_RATIOS[task_name]
            needed = max(min_s, _math.ceil(total_pax / ratio))
        else:
            needed = max(1, _math.ceil(len(fns) / 1))
        needed = min(needed, rule["max_staff_count"])

        safe_name  = task_name[:8].replace(" ", "").replace("/", "")
        task_id    = f"TBLK_{terminal}_{block_idx}_{direction}_{safe_name}"
        task_label = f"{task_name} -- {terminal} {direction}"
        all_tasks.append({
            "id":              task_id,
            "flight_no":       fns[0],
            "task":            task_label,
            "role":            task_name,
            "skill":           skill,
            "priority":        rule["priority"],
            "start_mins":      blk_start,
            "end_mins":        blk_end,
            "start":           mins_to_time(blk_start),
            "end":             mins_to_time(blk_end),
            "staff_needed":    needed,
            "staff_capacity":  rule["max_staff_count"],
            "assigned":        [],
            "alert":           None,
            "time_mins":       blk_start,
            "flights_covered": fns,
            "terminal":        terminal,
            "pier":            "ALL",
            "sharing_mode":    "block_shared",
            "total_pax":       total_pax,
            "time_window":     f"{mins_to_time(blk_start)}-{mins_to_time(blk_end)}",
        })

    # All Flights / US Flights scope (pier-level block)
    for (terminal, pier, block_idx, direction, task_name), data in flight_blocks.items():
        if not data["flights"]:
            continue
        rule      = data["rule"]
        total_pax = data["pax"]
        fns       = data["flights"]
        blk_start = block_idx * TASK_SLOT_MINS
        blk_end   = min(1440, blk_start + TASK_SLOT_MINS)
        skill     = TASK_SKILL.get(task_name, "GNIB")

        if task_name in PAX_RATIOS and total_pax > 0:
            ratio, min_s = PAX_RATIOS[task_name]
            needed = max(min_s, _math.ceil(total_pax / ratio))
        else:
            needed = max(1, len(fns))
        needed = min(needed, rule["max_staff_count"])

        safe_name  = task_name[:8].replace(" ", "").replace("/", "")
        task_id    = f"FBLK_{terminal}_{pier}_{block_idx}_{direction}_{safe_name}"
        task_label = f"{task_name} -- {pier} {direction}"
        all_tasks.append({
            "id":              task_id,
            "flight_no":       fns[0],
            "task":            task_label,
            "role":            task_name,
            "skill":           skill,
            "priority":        rule["priority"],
            "start_mins":      blk_start,
            "end_mins":        blk_end,
            "start":           mins_to_time(blk_start),
            "end":             mins_to_time(blk_end),
            "staff_needed":    needed,
            "staff_capacity":  rule["max_staff_count"],
            "assigned":        [],
            "alert":           None,
            "time_mins":       blk_start,
            "flights_covered": fns,
            "terminal":        terminal,
            "pier":            pier,
            "sharing_mode":    "pier_block",
            "total_pax":       total_pax,
            "time_window":     f"{mins_to_time(blk_start)}-{mins_to_time(blk_end)}",
        })

    return all_tasks



# ---------------------------------------------------------------------------
# Main optimiser
# ---------------------------------------------------------------------------

def optimize_day(date_str, overrides=None, manual_assigns=None, current_time_mins=None, prefer_early=False, custom_constraints=None):
    """Run the greedy staff-assignment optimiser for a single day.

    Returns a rich dict with flights, tasks, staff, alerts and KPIs.
    """
    if overrides is None:
        overrides = {}
    if manual_assigns is None:
        manual_assigns = {}
    if custom_constraints is None:
        custom_constraints = {}
        
    tt_t1_t2 = int(custom_constraints.get('tt_t1_t2', 15))
    tt_skill_switch = int(custom_constraints.get('tt_skill_switch', 10))
    # Support both singular/plural naming from different JS files
    allow_overlap = custom_constraints.get('allow_overlap') or custom_constraints.get('allow_overlaps') or False
    use_primary_first = custom_constraints.get('use_primary_first', True)

    if current_time_mins is not None:
        current_time_mins = int(current_time_mins)

    # Parse date
    d = None
    for fmt in ('%Y-%m-%d', '%d-%b-%y', '%d-%m-%Y', '%d-%m-%y'):
        try:
            d = datetime.strptime(date_str.strip(), fmt)
            break
        except ValueError:
            pass
    if d is None:
        return {'error': f'Cannot parse date: {date_str}'}

    iso_date_key    = d.strftime('%Y-%m-%d')    # e.g. '2026-04-11'
    date_label      = d.strftime('%A %d %b %Y')

    # Short-term and intraday optimisation is passenger-driven only.
    # Flight schedules, gates, stands, and Config.csv task generation are not
    # used as tactical demand inputs.
    flights_raw = []
    delay_map = {}
    rules = []
    stands_map = {}
    on_duty, absent_staff, _cf = get_staff_for_date(date_str, custom_constraints)
    print(f"[DEBUG] On-duty: {len(on_duty)}, Absent: {len(absent_staff)}")

    # Build skill lookup dicts using raw CSV skill names.
    # Staff CSV uses task-name labels ('Gate 335', 'Dep/Trolleys', etc.).
    # Pools are keyed by these raw names; task-to-staff matching uses SKILL_ALIASES below.
    staff_by_prim = defaultdict(list)   # raw_skill1 → [staff_dict]
    staff_by_any  = defaultdict(list)   # raw skill (any slot) → [staff_dict]
    for s in on_duty:
        raw1 = s.get('skill1', '').strip()
        if raw1:
            staff_by_prim[raw1].append(s)
            staff_by_any[raw1].append(s)

        seen_raw = {raw1}
        for raw_sk in [s.get('skill2',''), s.get('skill3',''), s.get('skill4','')]:
            rsk = raw_sk.strip()
            if rsk and rsk not in seen_raw:
                staff_by_any[rsk].append(s)
                seen_raw.add(rsk)

    # Reverse map: task_skill → set of raw CSV skill names that qualify.
    # e.g. 'GNIB' → {'GNIB', 'Gate 335', 'Departures', 'Check-in/Trolleys', ...}
    #      'Bussing' → {'Bussing', 'Dep/Trolleys', 'T1/T2 Trolleys L/UL', ...}
    _skill_aliases: dict = defaultdict(set)
    for raw_name, ts in TASK_SKILL.items():
        _skill_aliases[ts].add(raw_name)
    # Always include the task_skill itself in its own alias set
    for ts in set(TASK_SKILL.values()):
        _skill_aliases[ts].add(ts)

    def _skill_norm(raw_sk: str) -> str:
        """Map a raw staff skill name to its task-execution skill."""
        return TASK_SKILL.get(raw_sk.strip(), raw_sk.strip()) if raw_sk else ''

    def _candidate_pools(task_skill: str):
        """Return (prim_pool, any_pool) covering all raw skills for task_skill."""
        keys = _skill_aliases.get(task_skill, {task_skill})
        seen_ids: set = set()
        prim: list = []
        any_: list = []
        for k in keys:
            for s in staff_by_prim.get(k, []):
                if id(s) not in seen_ids:
                    seen_ids.add(id(s)); prim.append(s)
        seen_ids.clear()
        for k in keys:
            for s in staff_by_any.get(k, []):
                if id(s) not in seen_ids:
                    seen_ids.add(id(s)); any_.append(s)
        return prim, any_

    # busy_map: emp_id → [(start, end, terminal, skill)]
    busy_map = defaultdict(list)
    planned_breaks_by_id = assign_stagger_breaks(on_duty, custom_constraints)
    for s in on_duty:
        sid = str(s.get('id', s.get('name', '')))
        s['breaks'] = planned_breaks_by_id.get(sid, [])
        for br in s['breaks']:
            busy_map[sid].append((br['start_mins'], br['end_mins'], 'BREAK', 'BREAK'))

    def available(s, task_start, task_end, task_terminal, task_skill, skip_break_check=False):
        """Check if staff member s is available for window [task_start, task_end).

        skip_break_check=True is used for post-coverage tasks (e.g. Mezz Operation)
        where the post is permanently manned and workers rotate breaks while a
        colleague covers — so a worker's own break does not block them from the post.
        """

        # Shift window
        S, E = s['shift_start'], s['shift_end']
        D_shift = (E - S) % 1440
        if D_shift == 0 and E == S: D_shift = 1440
        
        # Task duration and relative start
        task_dur = (task_end - task_start) % 1440
        if task_dur == 0 and task_end != task_start: task_dur = 1440
        ts_rel = (task_start - S) % 1440
        
        # Is task within shift window?
        if ts_rel + task_dur > D_shift:
            return False
            
        if allow_overlap:
            return True
            
        # Check busy map for overlaps with buffer.
        # For post-coverage tasks: skip BREAK entries — breaks are covered by
        # rotating staff at the post, so a worker's break doesn't vacate the post.
        for (ws, we, term, sk) in busy_map[str(s['id'])]:
            if skip_break_check and term == 'BREAK':
                continue
            buffer_mins = 0
            if term != 'BREAK' and term != task_terminal and term != 'ALL' and task_terminal != 'ALL':
                buffer_mins = max(buffer_mins, tt_t1_t2)
            if sk != 'BREAK' and sk != task_skill:
                buffer_mins = max(buffer_mins, tt_skill_switch)
            
            # Normalize busy window to shift-relative coordinates
            ws_rel = (ws - S) % 1440
            w_dur = (we - ws) % 1440
            if w_dur == 0 and we != ws: w_dur = 1440
            
            # Overlap check in relative space
            if ts_rel < (ws_rel + w_dur + buffer_mins) and (ts_rel + task_dur) > (ws_rel - buffer_mins):
                return False
        return True

    def _within_shift(s, task_start, task_end):
        """Shift-window-only check (no busy_map). Used for Mezz pre-assignment."""
        S, E = s['shift_start'], s['shift_end']
        D_shift = (E - S) % 1440
        if D_shift == 0 and E == S:
            D_shift = 1440
        task_dur = (task_end - task_start) % 1440
        if task_dur == 0 and task_end != task_start:
            task_dur = 1440
        ts_rel = (task_start - S) % 1440
        return (ts_rel + task_dur) <= D_shift

    # ── Build processed-flight list (apply overrides once, reuse everywhere) ──
    # Each entry carries enough info for _generate_day_tasks and flights_map.
    processed_flights = []
    cbp_dep_times = []

    for flight in flights_raw:
        fn = flight.get('flight_no', '').strip()
        if fn in cancelled_set:
            continue
        sta_raw = flight.get('sta', '').strip()
        t_mins = parse_time(sta_raw)
        if t_mins is None:
            continue
        t_mins += delay_map.get(fn, 0)

        status   = flight.get('Status', '').strip()
        icao_cat = flight.get('icao_cat', '').strip()
        cbp_flag = flight.get('cbp_flag', '').strip()
        gate     = flight.get('gate', '').strip()
        haul     = icao_to_haul(icao_cat, cbp_flag)

        # Collect US/Canada DEP times for CBP hall session task
        if haul == 'US/Canada' and status == 'Departure':
            cbp_dep_times.append(t_mins)

        processed_flights.append({
            'flight_no': fn,
            'time_mins': t_mins,
            'status':    status,
            'haul':      haul,
            'gate':      gate,
            'icao_cat':  icao_cat,
            'cbp_flag':  cbp_flag,
            # Keep raw flight dict for flights_map construction below
            '_raw':      flight,
        })

    # ── Generate flight tasks from Config.csv rules ───────────────────────────
    all_tasks = []

    # ── CBP hall task — session-level, driven by US/Canada departures ────────
    # Find CBP rule from config for max-staff cap
    _cbp_rule = next((r for r in rules if r['task'] == 'CBP Pre-clearance'), None)
    if cbp_dep_times:
        cbp_start = max(0, min(cbp_dep_times) - 90)
        cbp_end   = max(cbp_dep_times) + 30
        if cbp_end <= cbp_start:
            cbp_end = cbp_start + 120
        cbp_flights = [pf['flight_no'] for pf in processed_flights
                       if pf['haul'] == 'US/Canada' and pf['status'] == 'Departure']
        _cbp_need = min(3, _cbp_rule['max_staff_count']) if _cbp_rule else 3
        _cbp_cap  = _cbp_rule['max_staff_count'] if _cbp_rule else 3
        _cbp_pri  = _cbp_rule['priority'] if _cbp_rule else 'Critical'
        all_tasks.append({
            'id':              f'CBP_HALL_{cbp_start}',
            'flight_no':       'CBP-HALL',
            'task':            'CBP Pre-clearance',
            'skill':           'CBP Pre-clearance',
            'priority':        _cbp_pri,
            'start_mins':      cbp_start,
            'end_mins':        cbp_end,
            'start':           mins_to_time(cbp_start),
            'end':             mins_to_time(cbp_end),
            'staff_needed':    _cbp_need,
            'staff_capacity':  _cbp_cap,
            'assigned':        [],
            'alert':           None,
            'time_mins':       cbp_start,
            'flights_covered': cbp_flights,
            'terminal':        'T2',
            'pier':            'P4',
            'sharing_mode':    'shared',
            'time_window':     f"{mins_to_time(cbp_start)}-{mins_to_time(cbp_end)}",
        })

    # ── Fixed duties — generated from Config.csv rows with scope='Fixed' ──────
    # Each Fixed rule runs two standard shifts (AM 04:00-12:00 / PM 12:00-20:00).
    # Mezz Operation is included as a 24-hour fixed post.
    _fixed_rules = [r for r in rules if r['scope'] == 'Fixed']
    fixed_duties = []
    for _fr in _fixed_rules:
        _task   = _fr['task']
        _skill  = TASK_SKILL.get(_task, _task)
        _pri    = 'Critical' if _task == 'Mezz Operation' else (_fr['priority'] or 'Medium')
        # Use the configured max_staff_count as the required headcount for
        # fixed duties so Mezz Operation slots can be manned by multiple
        # workers concurrently (e.g., max 3 as defined in Config.csv).
        _needed = _fr['max_staff_count']
        _cap    = _fr['max_staff_count']
        
        if _task == 'Litter Picking':
            # 2 hrs daytime, 2 hrs nighttime at 2 FTE
            shifts_to_run = [(600, 720), (1320, 1440)]
        elif _task == 'Mezz Operation':
            # Full 24h coverage: 00:00-12:00 and 12:00-24:00.
            # Multiple workers may be assigned to the same slot (overlapping
            # is permitted for Mezz); staff_capacity = max_staff_count = 3.
            # post_coverage=True: breaks do not block Mezz workers from the post
            # (rotating coverage means someone else covers during a break).
            shifts_to_run = [(0, 720), (720, 1440)]
        else:
            shifts_to_run = [(240, 720), (720, 1200)]
            
        for _shift_idx, (_s, _e) in enumerate(shifts_to_run):
            slot_idx = 0
            for _slot_start in range(_s, _e, TASK_SLOT_MINS):
                _slot_end = min(_e, _slot_start + TASK_SLOT_MINS)
                fixed_duties.append({
                    'id':            f"FIXED_{_task[:6].replace(' ', '')}_{_shift_idx}_{slot_idx}",
                    'task':          _task,
                    'role':          _task,
                    'skill':         _skill,
                    'priority':      _pri,
                    'start_mins':    _slot_start,
                    'end_mins':      _slot_end,
                    'staff_needed':  _needed,
                    'staff_capacity': _cap,
                    # post_coverage tasks (e.g. Mezz) are continuously manned:
                    # workers rotate breaks so a break does not vacate the post.
                    'post_coverage': (_task == 'Mezz Operation'),
                })
                slot_idx += 1
    for fd in fixed_duties:
        fd.update({
            'flight_no':       'FIXED',
            'start':           mins_to_time(fd['start_mins']),
            'end':             mins_to_time(fd['end_mins']),
            'assigned':        [],
            'alert':           None,
            'time_mins':       fd['start_mins'],
            'flights_covered': [],
            'terminal':        'ALL',
            'pier':            'ALL',
            'sharing_mode':    'fixed',
            'time_window':     f"{mins_to_time(fd['start_mins'])}-{mins_to_time(fd['end_mins'])}",
        })
        all_tasks.append(fd)

    pax_tasks = build_pax_demand_tasks(iso_date_key, current_time_mins)
    if pax_tasks:
        all_tasks = pax_tasks

    # Mark completed tasks in live intraday mode instead of deleting them.
    # This preserves them for the timeline and staff utilisation history.
    if current_time_mins is not None:
        for t in all_tasks:
            if t['end_mins'] <= current_time_mins:
                t['is_past'] = True

    # Apply manual assigns FIRST
    for task in all_tasks:
        tid = task['id']
        if tid in manual_assigns:
            for emp_id in manual_assigns[tid]:
                # Find staff member
                s = next((x for x in on_duty if x['id'] == emp_id), None)
                if s and emp_id not in task['assigned']:
                    # Respect staff skills: only apply manual assigns when the
                    # employee holds the required task skill (skill1..skill4).
                    task_skill = task.get('skill', 'GNIB')
                    staff_skills = {
                        _skill_norm(s.get('skill1', '')).strip(),
                        _skill_norm(s.get('skill2', '')).strip(),
                        _skill_norm(s.get('skill3', '')).strip(),
                        _skill_norm(s.get('skill4', '')).strip(),
                    }
                    staff_skills.discard('')
                    if task_skill not in staff_skills:
                        # Skip invalid manual assign (staff lacks skill)
                        print(f"[WARN] Skipping manual assign: {emp_id} lacks skill '{task_skill}' for task {tid}")
                        continue

                    task['assigned'].append(emp_id)
                    s['assignments'].append({
                        'task_id':    tid,
                        'task':       task['task'],
                        'skill':      task.get('skill', 'GNIB'),
                        'terminal':   task.get('terminal', 'ALL'),
                        'start':      task['start'],
                        'end':        task['end'],
                        'start_mins': task['start_mins'],
                        'end_mins':   task['end_mins'],
                    })
                    busy_map[emp_id].append((task['start_mins'], task['end_mins'], task.get('terminal', 'ALL'), task.get('skill', 'GNIB')))

    # ── Block-level PAX pre-assignment ───────────────────────────────────────
    # Group PAX hourly tasks into 3-hour blocks (same skill + terminal).
    # Assign staff for the FULL block window so nobody switches skill/terminal
    # inside the block. The full block is recorded in busy_map, preventing the
    # greedy loop from scattering staff across different positions intra-block.
    # Continuity is tracked across consecutive blocks to minimise skill/terminal
    # changes over the full day.
    _PAX_TIME_BLOCKS = [
        (0, 180), (180, 360), (360, 540), (540, 720),
        (720, 900), (900, 1080), (1080, 1260), (1260, 1440),
    ]

    _pax_blk_groups: dict = defaultdict(list)
    for _t in all_tasks:
        if _t.get('sharing_mode') not in ('pax_15min', 'pax_hourly') or _t.get('is_past'):
            continue
        _sm = _t['start_mins']
        for _bi, (_bs, _be) in enumerate(_PAX_TIME_BLOCKS):
            if _bs <= _sm < _be:
                _pax_blk_groups[(_bi, _t.get('skill', ''), _t.get('terminal', 'ALL'))].append(_t)
                break

    # (skill, terminal) → list of emp_ids assigned in the previous block (continuity)
    _blk_continuity: dict = defaultdict(list)

    for _bi in range(len(_PAX_TIME_BLOCKS)):
        _blk_start, _blk_end = _PAX_TIME_BLOCKS[_bi]
        _next_continuity: dict = defaultdict(list)

        for (_bi2, _skill, _term), _group in sorted(_pax_blk_groups.items()):
            if _bi2 != _bi:
                continue
            _peak_need = max((_t['staff_needed'] for _t in _group), default=0)
            if _peak_need == 0:
                continue

            _already_ids: set = set()
            for _t in _group:
                _already_ids.update(_t['assigned'])
            _needed = max(0, _peak_need - len(_already_ids))

            if _needed > 0:
                # Prefer: (1) continuity staff, (2) primary-skill pool, (3) any-skill pool
                _prim_pool, _any_pool = _candidate_pools(_skill)
                _prev_ids = _blk_continuity.get((_skill, _term), [])
                _seen_cands: set = set()
                _ordered_cands: list = []

                for _sid in _prev_ids:
                    if _sid in _already_ids or _sid in _seen_cands:
                        continue
                    _s = next((x for x in on_duty if x['id'] == _sid), None)
                    if _s and available(_s, _blk_start, _blk_end, _term, _skill):
                        _ordered_cands.append(_s)
                        _seen_cands.add(_sid)

                for _s in _prim_pool:
                    if _s['id'] in _seen_cands or _s['id'] in _already_ids:
                        continue
                    if available(_s, _blk_start, _blk_end, _term, _skill):
                        _ordered_cands.append(_s)
                        _seen_cands.add(_s['id'])

                for _s in _any_pool:
                    if _s['id'] in _seen_cands or _s['id'] in _already_ids:
                        continue
                    if available(_s, _blk_start, _blk_end, _term, _skill):
                        _ordered_cands.append(_s)
                        _seen_cands.add(_s['id'])

                for _s in _ordered_cands:
                    if _needed <= 0:
                        break
                    _sid = str(_s['id'])
                    # Assign this staff to every hourly task in the block where still needed.
                    for _t in _group:
                        if _sid not in _t['assigned'] and len(_t['assigned']) < _t['staff_needed']:
                            _t['assigned'].append(_sid)
                    # Block the FULL 3-hr window so greedy won't re-assign elsewhere
                    busy_map[_sid].append((_blk_start, _blk_end, _term, _skill))
                    # One clean block entry on the staff record (for roster display)
                    _s.setdefault('assignments', []).append({
                        'task_id':    f"BLK_{_bi}_{_skill}_{_term}",
                        'task':       f"{_term} {_skill}",
                        'skill':      _skill,
                        'terminal':   _term,
                        'start':      mins_to_time(_blk_start),
                        'end':        mins_to_time(_blk_end),
                        'start_mins': _blk_start,
                        'end_mins':   _blk_end,
                    })
                    _already_ids.add(_sid)
                    _needed -= 1

            _next_continuity[(_skill, _term)] = list(_already_ids)

        _blk_continuity = _next_continuity

    # Sort tasks for assignment.
    # Key objectives in order:
    #   1. Priority (Critical → High → Medium → Low) from Config.csv
    #   2. Coverage breadth: tasks covering more flights processed first so shared
    #      resources are allocated where they deliver the most coverage.
    #   3. Terminal/pier spread: tasks serving multiple terminals ahead of single-terminal
    #   4. Earliest start time (so deadlines are met in time order)
    #   5. Flight number as final tiebreaker for determinism
    # In live intraday mode (prefer_early=True) start_mins leads so past-due tasks
    # are never skipped in favour of future high-priority ones.
    priority_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}
    def _fixed_rank(t):
        return 0 if t.get('sharing_mode') == 'fixed' else 1

    if prefer_early:
        all_tasks.sort(key=lambda t: (
            _fixed_rank(t),
            t['start_mins'],
            priority_order.get(t['priority'], 2),
            -len(t.get('flights_covered', [])),
            t.get('flight_no', ''),
        ))
    else:
        all_tasks.sort(key=lambda t: (
            _fixed_rank(t),
            priority_order.get(t['priority'], 2),
            -len(t.get('flights_covered', [])),
            t['start_mins'],
            t.get('flight_no', ''),
        ))

    # ── CP-SAT optimisation (enabled via use_cpsat flag) ──────────────────────
    # Replaces the greedy loop when OR-Tools is installed and the flag is set.
    # On solver failure or if the flag is off, falls through to greedy below.
    _cpsat_applied = False
    if custom_constraints.get('use_cpsat', False) and _CPSAT_AVAILABLE:
        try:
            # Build the constraint dict expected by the optimizer module
            _cpsat_cons = {
                'tt_t1_t2':               tt_t1_t2,
                'tt_skill_switch':        tt_skill_switch,
                'b1_duration_mins':       int(custom_constraints.get('b1_duration_mins', 30)),
                'b2_duration_mins':       int(custom_constraints.get('b2_duration_mins', 60)),
                'max_overtime_per_day_hrs': 2,
                'planned_breaks':          planned_breaks_by_id,
            }
            # Only pass tasks that still need staff (manual assigns may have
            # already covered some or all slots for a task)
            _open_tasks = [
                t for t in all_tasks
                if len(t.get('assigned', [])) < t.get('staff_needed', 1)
                and (not t.get('is_past') or t.get('post_coverage', False))
            ]
            _cpsat_result = _cpsat_optimize(_open_tasks, on_duty, _cpsat_cons)

            if _cpsat_result and _cpsat_result.get('solver_status') in ('OPTIMAL', 'FEASIBLE'):
                _task_lookup = {t['id']: t for t in all_tasks}

                # Apply solver assignments to task and staff objects in-place
                for _asgn in _cpsat_result['assignments']:
                    _task = _task_lookup.get(_asgn['task_id'])
                    if not _task:
                        continue
                    for _emp_id in _asgn['staff_ids']:
                        if _emp_id in _task['assigned']:
                            continue
                        _s = next((w for w in on_duty if w['id'] == _emp_id), None)
                        if not _s:
                            continue
                        _task['assigned'].append(_emp_id)
                        _s['assignments'].append({
                            'task_id':    _task['id'],
                            'task':       _task['task'],
                            'skill':      _task.get('skill', 'GNIB'),
                            'terminal':   _task.get('terminal', 'ALL'),
                            'start':      _task['start'],
                            'end':        _task['end'],
                            'start_mins': _task['start_mins'],
                            'end_mins':   _task['end_mins'],
                        })
                        # Update busy_map so break-scheduling later is accurate
                        busy_map[_emp_id].append((
                            _task['start_mins'], _task['end_mins'],
                            _task.get('terminal', 'ALL'),
                            _task.get('skill', 'GNIB'),
                        ))

                # Set alerts on tasks that are still under-staffed after CP-SAT
                for _task in all_tasks:
                    _needed  = _task['staff_needed']
                    _covered = len(_task['assigned'])
                    if _covered < _needed and not _task.get('is_past'):
                        _task['alert'] = (
                            f'Under-staffed: need {_needed}, assigned {_covered}'
                            f' (gap {_needed - _covered})'
                        )

                _cpsat_applied = True
                print(
                    f"[CP-SAT] {_cpsat_result['solver_status']}  "
                    f"unassigned={len(_cpsat_result['unassigned_tasks'])}  "
                    f"gap={_cpsat_result['gap_pct']}%"
                )
        except Exception as _cpsat_exc:
            print(f"[CP-SAT] Error ({_cpsat_exc}); falling back to greedy.")

    # ── Greedy assignment (default path, or fallback when CP-SAT is off/fails) ─
    _greedy_tasks = [] if _cpsat_applied else all_tasks
    print(f"[DEBUG] Assigning {len(_greedy_tasks)} tasks (greedy)...")
    for task in _greedy_tasks:
        needed = task['staff_needed'] - len(task['assigned'])
        if needed <= 0:
            continue

        skill = task['skill']
        start = task['start_mins']
        end   = task['end_mins']

        # Optimization: Lazy evaluation of candidates.
        # Instead of pre-filtering all staff (which calls expensive available() for everyone),
        # we iterate and stop as soon as we meet the 'needed' headcount.
        assigned_count = 0

        _prim_pool, _any_pool = _candidate_pools(skill)
        candidate_pools = []
        if use_primary_first:
            candidate_pools = [_prim_pool, _any_pool]
        else:
            candidate_pools = [_any_pool]

        for pool_index, pool in enumerate(candidate_pools):
            if assigned_count >= needed:
                break
            for s in pool:
                if assigned_count >= needed:
                    break
                if s['id'] in task['assigned']:
                    continue
                # Post-coverage tasks (Mezz): skip break-conflict check so that
                # a worker's own break doesn't prevent them from being rostered
                # at the permanently-manned post.
                _skip_br = task.get('post_coverage', False)
                if not available(s, start, end, task.get('terminal', 'ALL'), task.get('skill', 'GNIB'), skip_break_check=_skip_br):
                    continue

                assignment = {
                    'task_id':   task['id'],
                    'task':      task['task'],
                    'skill':     task.get('skill', 'GNIB'),
                    'terminal':  task.get('terminal', 'ALL'),
                    'start':     task['start'],
                    'end':       task['end'],
                    'start_mins': start,
                    'end_mins':   end,
                }

                task['assigned'].append(s['id'])
                s['assignments'].append(assignment)
                busy_map[s['id']].append((start, end, task.get('terminal', 'ALL'), task.get('skill', 'GNIB')))
                assigned_count += 1

        if assigned_count < needed:
            gap = needed - assigned_count
            if not task.get('is_past'):
                task['alert'] = f'Under-staffed: need {needed}, assigned {assigned_count} (gap {gap})'

    # ── Pass 2: relax travel/skill-switch buffers for still-unassigned tasks ──
    # Retry with zero inter-task buffers so back-to-back tasks on the same staff
    # can be filled when minor buffer constraints prevented assignment above.
    def _available_no_buffer(s, task_start, task_end, extend_shift_mins=0):
        S, E = s['shift_start'], s['shift_end']
        D_shift = (E - S) % 1440
        if D_shift == 0 and E == S:
            D_shift = 1440
        task_dur = (task_end - task_start) % 1440
        if task_dur == 0 and task_end != task_start:
            task_dur = 1440
        ts_rel = (task_start - S) % 1440
        if ts_rel + task_dur > D_shift + extend_shift_mins:
            return False
        for (ws, we, _t, _sk) in busy_map[s['id']]:
            ws_rel = (ws - S) % 1440
            w_dur  = (we - ws) % 1440
            if w_dur == 0 and we != ws:
                w_dur = 1440
            if ts_rel < ws_rel + w_dur and ts_rel + task_dur > ws_rel:
                return False
        return True

    _pass2_tasks = [
        t for t in (_greedy_tasks if not _cpsat_applied else all_tasks)
        if len(t['assigned']) < t['staff_needed'] and not t.get('is_past')
    ]
    if _pass2_tasks:
        print(f"[DEBUG] Pass-2 (zero buffers): {len(_pass2_tasks)} tasks still need staff")
        for task in _pass2_tasks:
            needed = task['staff_needed'] - len(task['assigned'])
            if needed <= 0:
                continue
            skill = task['skill']
            start = task['start_mins']
            end   = task['end_mins']
            _p2_prim, _p2_any = _candidate_pools(skill)
            for pool in [_p2_prim, _p2_any]:
                if needed <= 0:
                    break
                for s in pool:
                    if needed <= 0:
                        break
                    if s['id'] in task['assigned']:
                        continue
                    if not _available_no_buffer(s, start, end):
                        continue
                    task['assigned'].append(s['id'])
                    s['assignments'].append({
                        'task_id':    task['id'],
                        'task':       task['task'],
                        'skill':      skill,
                        'terminal':   task.get('terminal', 'ALL'),
                        'start':      task['start'],
                        'end':        task['end'],
                        'start_mins': start,
                        'end_mins':   end,
                    })
                    busy_map[s['id']].append((start, end, task.get('terminal', 'ALL'), skill))
                    needed -= 1
            if len(task['assigned']) >= task['staff_needed']:
                task['alert'] = None

    # ── Pass 3: extend shift window (up to 90 min overtime) + any-skill fallback ─
    # Last-resort pass so no task is left completely unassigned. Staff with
    # zero skill overlap are marked with skill_mismatch=True.
    _pass3_tasks = [
        t for t in all_tasks
        if len(t['assigned']) < t['staff_needed'] and not t.get('is_past')
    ]
    if _pass3_tasks:
        print(f"[DEBUG] Pass-3 (extended shift + any skill): {len(_pass3_tasks)} tasks")
        _all_staff_sorted = sorted(on_duty, key=lambda s: len(busy_map[s['id']]))
        for task in _pass3_tasks:
            needed = task['staff_needed'] - len(task['assigned'])
            if needed <= 0:
                continue
            skill = task['skill']
            start = task['start_mins']
            end   = task['end_mins']
            if 'mismatch_assigned' not in task:
                task['mismatch_assigned'] = []
            _p3_prim, _p3_any = _candidate_pools(skill)
            skill_pools = list({id(s): s for s in (_p3_prim + _p3_any)}.values())
            for s in skill_pools:
                if needed <= 0:
                    break
                if s['id'] in task['assigned']:
                    continue
                is_fte = s.get('employment', '').strip().lower() in ('full-time', 'full time', 'fte', 'permanent')
                ext_mins = 0 if is_fte else 90
                if not _available_no_buffer(s, start, end, extend_shift_mins=ext_mins):
                    continue
                task['assigned'].append(s['id'])
                s['assignments'].append({
                    'task_id':        task['id'],
                    'task':           task['task'],
                    'skill':          skill,
                    'terminal':       task.get('terminal', 'ALL'),
                    'start':          task['start'],
                    'end':            task['end'],
                    'start_mins':     start,
                    'end_mins':       end,
                    'skill_mismatch': False,
                })
                busy_map[s['id']].append((start, end, task.get('terminal', 'ALL'), skill))
                needed -= 1
            if len(task['assigned']) >= task['staff_needed']:
                task['alert'] = None
            else:
                remaining = task['staff_needed'] - len(task['assigned'])
                task['alert'] = f'Under-staffed: need {task["staff_needed"]}, assigned {len(task["assigned"])} (gap {remaining})'

    # ── Pass 4: Full staff utilisation ────────────────────────────────────────
    # For every on-duty staff member, find 30-minute slots within their shift where
    # they have no assignment, and assign them to the highest-priority task in
    # that block that matches one of their skills and has spare capacity.
    # This ensures all staff are productive every hour they are on duty.
    _p4_priority = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}

    # Index tasks by 30-minute slot index
    _blk_task_index: dict = defaultdict(list)
    for _t in all_tasks:
        if _t.get('is_past'):
            continue
        _blk = _t['start_mins'] // TASK_SLOT_MINS
        _blk_task_index[_blk].append(_t)
    # Sort each bucket: Critical first, then more-assigned tasks (fill existing slots)
    for _blk in _blk_task_index:
        _blk_task_index[_blk].sort(key=lambda t: (
            _p4_priority.get(t.get('priority', 'Medium'), 2),
            -len(t['assigned']),
        ))

    _p4_assigned = 0
    for s in on_duty:
        emp_id   = s['id']
        # Normalise via TASK_SKILL so 'Gate 335' → 'GNIB', 'Dep/Trolleys' → 'Bussing', etc.
        s_skills = {_skill_norm(sk) for sk in [s.get('skill1',''), s.get('skill2',''), s.get('skill3',''), s.get('skill4','')] if sk}
        sh_start = s['shift_start']
        sh_end   = s['shift_end']

        s_dur = (sh_end - sh_start) % 1440
        if s_dur == 0 and sh_end != sh_start:
            s_dur = 1440
        first_blk = sh_start // TASK_SLOT_MINS
        num_blks = s_dur // TASK_SLOT_MINS

        for blk_offset in range(num_blks):
            blk = (first_blk + blk_offset) % 48
            blk_start = blk * TASK_SLOT_MINS
            blk_end   = min(1440, blk_start + TASK_SLOT_MINS)

            # Skip if staff already has an assignment overlapping this block
            already_busy = any(
                ws < blk_end and we > blk_start
                for (ws, we, _t, _sk) in busy_map[emp_id]
            )
            if already_busy:
                continue

            # Find best matching task in this block
            for _t in _blk_task_index.get(blk, []):
                t_skill = _t.get('skill', 'GNIB')
                if t_skill not in s_skills:
                    continue
                cap = _t['staff_needed']  # cap at requirement, not excess capacity
                if len(_t['assigned']) >= cap:
                    continue
                if emp_id in _t['assigned']:
                    continue
                if not _available_no_buffer(s, _t['start_mins'], _t['end_mins']):
                    continue

                _t['assigned'].append(emp_id)
                s['assignments'].append({
                    'task_id':    _t['id'],
                    'task':       _t['task'],
                    'skill':      t_skill,
                    'terminal':   _t.get('terminal', 'ALL'),
                    'start':      _t['start'],
                    'end':        _t['end'],
                    'start_mins': _t['start_mins'],
                    'end_mins':   _t['end_mins'],
                })
                busy_map[emp_id].append((_t['start_mins'], _t['end_mins'], _t.get('terminal', 'ALL'), t_skill))
                _p4_assigned += 1
                break  # one task per block per staff member

    print(f"[DEBUG] Pass-4 (full utilisation): {_p4_assigned} additional assignments")

    # Schedule breaks and compute utilisation
    print(f"[DEBUG] Scheduling breaks for staff...")
    for s in on_duty:
        windows = [(ws, we) for (ws, we, t, sk) in busy_map[s['id']] if t != 'BREAK']
        s['breaks'] = planned_breaks_by_id.get(str(s.get('id', s.get('name', ''))), schedule_breaks(s, windows, custom_constraints))
        total_busy = sum(e - st for (st, e) in windows)
        shift_len = s['shift_end'] - s['shift_start']
        s['utilisation_pct'] = round(min(total_busy / shift_len * 100, 100), 1) if shift_len > 0 else 0

    # ── Build flights_map from already-processed flight list ─────────────────
    # processed_flights already has overrides applied; no need to re-parse.
    flights_map = {}
    for pf in processed_flights:
        fn  = pf['flight_no']
        raw = pf['_raw']
        si  = _stand_info(pf['gate'], stands_map)
        flights_map[fn] = {
            'flight_no':    fn,
            'date':         raw.get('date', ''),
            'sta':          raw.get('sta', ''),
            'origin':       raw.get('origin', ''),
            'origin_code':  raw.get('origin_code', ''),
            'airline_name': raw.get('airline_name', ''),
            'aircraft_type':raw.get('aircraft_type', ''),
            'tail_reg':     raw.get('tail_reg', ''),
            'icao_cat':     pf['icao_cat'],
            'cbp_flag':     pf['cbp_flag'],
            'gate':         pf['gate'],
            'status':       pf['status'],
            'haul':         pf['haul'],
            'terminal':     si['terminal'],
            'pier':         si['pier'],
            'time_mins':    pf['time_mins'],
            'delayed':      fn in delay_map,
            'delay_mins':   delay_map.get(fn, 0),
            'tasks':        [],
        }

    # ── Attach tasks to flights ───────────────────────────────────────────────
    # Shared/partially-shared tasks cover multiple flights; they appear in the
    # task list of each flight they serve so the UI can show a full picture.
    # Dedicated and fixed tasks appear only on their primary flight.
    for task in all_tasks:
        task_summary = {
            'id':               task['id'],
            'task':             task['task'],
            'skill':            task['skill'],
            'priority':         task['priority'],
            'start':            task['start'],
            'end':              task['end'],
            'start_mins':       task['start_mins'],
            'end_mins':         task['end_mins'],
            'staff_needed':     task['staff_needed'],
            'assigned':         task['assigned'],
            'mismatch_assigned': task.get('mismatch_assigned', []),
            'alert':            task['alert'],
            'sharing_mode':     task.get('sharing_mode', 'dedicated'),
            'flights_covered':  task.get('flights_covered', []),
            'time_window':      task.get('time_window', ''),
            'terminal':         task.get('terminal', ''),
            'pier':             task.get('pier', ''),
            'post_coverage':    task.get('post_coverage', False),
            'passengers':       task.get('passengers', 0),
            'pax_rate':         task.get('pax_rate', 0),
        }
        for fn in task.get('flights_covered', [task.get('flight_no', '')]):
            if fn in flights_map:
                flights_map[fn]['tasks'].append(task_summary)

    # Mark flights as completed when all relevant tasks are in the past.
    if current_time_mins is not None:
        for flight in flights_map.values():
            if flight['tasks']:
                last_task_end = max(t['end_mins'] for t in flight['tasks'])
                if last_task_end <= current_time_mins:
                    flight['status'] = 'Completed'
            else:
                # No remaining tasks means the flight is complete for live intraday view.
                flight['status'] = 'Completed'

    flights_sorted = sorted(flights_map.values(), key=lambda x: x['time_mins'])

    # ── Build alerts list ────────────────────────────────────────────────────
    alerts = []
    for task in all_tasks:
        if task['alert']:
            needed         = task['staff_needed']
            assigned_count = len(task['assigned'])
            gap            = needed - assigned_count
            skill          = task['skill']
            start          = task['start_mins']
            end            = task['end_mins']
            rec_candidates = [s for s in staff_by_any.get(skill, [])
                               if s['id'] not in task['assigned'] and available(s, start, end, task.get('terminal', 'ALL'), task.get('skill', 'GNIB'))]
            rec_staff = [s['id'] for s in rec_candidates[:gap]]
            covered_flights = []
            for fn in task.get('flights_covered', []):
                flight = flights_map.get(fn)
                if not flight:
                    continue
                covered_flights.append({
                    'flight_no': flight.get('flight_no', ''),
                    'origin': flight.get('origin', ''),
                    'origin_code': flight.get('origin_code', ''),
                    'airline_name': flight.get('airline_name', ''),
                    'sta': flight.get('sta', ''),
                    'status': flight.get('status', ''),
                    'gate': flight.get('gate', ''),
                    'terminal': flight.get('terminal', ''),
                    'pier': flight.get('pier', ''),
                    'haul': flight.get('haul', ''),
                })
            alerts.append({
                'task_id':         task['id'],
                'flight_no':       task.get('flight_no', ''),
                'flights_covered': task.get('flights_covered', []),
                'covered_flights': covered_flights,
                'task':            task['task'],
                'skill':           skill,
                'priority':        task['priority'],
                'start':           task['start'],
                'end':             task['end'],
                'staff_needed':    needed,
                'assigned_count':  assigned_count,
                'assigned_staff':  task['assigned'],
                'gap':             gap,
                'message':         task['alert'],
                'rec_staff':       rec_staff,
                'sharing_mode':    task.get('sharing_mode', 'dedicated'),
                'terminal':        task.get('terminal', ''),
                'pier':            task.get('pier', ''),
                'time_window':     task.get('time_window', ''),
                'passengers':      task.get('passengers', 0),
                'pax_rate':        task.get('pax_rate', 0),
            })

    # Sort alerts: Critical first
    alerts.sort(key=lambda a: (priority_order.get(a['priority'], 2), a['start']))

    # ── KPIs ─────────────────────────────────────────────────────────────────
    tasks_covered = sum(1 for t in all_tasks if not t['alert'])
    tasks_total   = len(all_tasks)
    passengers_total = sum(int(t.get('passengers', 0) or 0) for t in all_tasks)
    # gates_active derived from processed_flights (overrides already applied)
    gates_active  = len(set(pf['gate'] for pf in processed_flights if pf['gate']))

    result = {
        'date':          iso_date_key,
        'date_label':    date_label,
        'staff_data_carried_forward': _cf,  # None or date string if fallback used
        'kpis': {
            'total_flights':  len(flights_sorted),
            'staff_on_duty':  len(on_duty),
            'absent':         len(absent_staff),
            'gates_active':   gates_active,
            'tasks_total':    tasks_total,
            'tasks_covered':  tasks_covered,
            'demand_windows_total': tasks_total,
            'demand_windows_covered': tasks_covered,
            'passengers_total': passengers_total,
            'coverage_pct':   round(tasks_covered / tasks_total * 100, 1) if tasks_total else 100.0,
        },
        'flights':       flights_sorted,
        'staff':         on_duty,
        'tasks':         all_tasks,
        'absent_staff':  absent_staff,
        'alerts':        alerts,
        'overrides':     overrides,
    }

    # Include PAX coverage skills derived from the PAX Config workbook so
    # the frontend can render rows even when no demand windows exist yet.
    try:
        pax_cfg = load_pax_config() or {}
        pax_skills = []
        for col in pax_cfg.keys():
            work = _pax_work_from_col(col)
            if not work:
                continue
            skill = PAX_WORK_SKILL_MAP.get(work, work)
            if skill not in pax_skills:
                pax_skills.append(skill)
        result['pax_coverage_skills'] = sorted(pax_skills, key=lambda s: str(s).lower())
    except Exception:
        result['pax_coverage_skills'] = []

    # Ensure assignments respect breaks and shift windows even when callers
    # call the optimizer without running the roster optimiser path. This keeps
    # the schedule shown on initial model open consistent with shift times.
    try:
        enforce_break_conflicts(result)
        refill_unassigned_tasks(result)
        enforce_fte_shift_containment(result)
        enforce_break_conflicts(result)
        _apply_manual_block_assigns_to_result(result, iso_date_key)
        _apply_manual_unassigns_to_result(result, iso_date_key)
        _normalize_reallocation_blocks(result, iso_date_key)
        result['kpis']['tasks_covered'] = sum(1 for t in result.get('tasks', []) if not t.get('alert'))
        result['kpis']['demand_windows_covered'] = result['kpis']['tasks_covered']
        total_windows = result['kpis'].get('demand_windows_total') or len(result.get('tasks', []))
        result['kpis']['coverage_pct'] = round(result['kpis']['tasks_covered'] / total_windows * 100, 1) if total_windows else 100.0
    except Exception:
        # Defensive: don't let a cleanup error break the API response.
        pass

    return result


# ---------------------------------------------------------------------------
# Short-term & Intraday Flask endpoints
# ---------------------------------------------------------------------------

@app.route('/api/short-term/dates')
def st_dates():
    """Return available D+1 to D+3 dates."""
    from datetime import date as date_type
    today = datetime.now()
    available_dates = []
    for i in [1, 2, 3]:
        d = today + timedelta(days=i)
        pax_source_date = _pax_source_date_for(d)
        available_dates.append({
            'label':    d.strftime('%A %d %b'),
            'date':     d.strftime('%Y-%m-%d'),
            'has_data': pax_source_date is not None,
        })
    return jsonify(available_dates)


def _get_short_term_schedule(date_str, preserve_manual_assigns=True):
    """Return short-term schedule using the same optimizer mode as intraday.

    Short-term does not use live overrides or a simulation clock, but it should
    share the same early-first assignment ordering so shift/task allocation is
    optimized consistently with the intraday view.
    """
    man = _manual_assigns.get(date_str, {}) if preserve_manual_assigns else {}
    custom_constraints = _st_custom_constraints_by_date.get(date_str, _st_custom_constraints)
    return optimize_day(
        date_str,
        manual_assigns=man,
        prefer_early=True,
        custom_constraints=custom_constraints,
    )


@app.route('/api/short-term/<date_str>')
def st_day(date_str):
    """Return full optimised schedule for a short-term day (D+1 to D+3)."""
    result = _get_short_term_schedule(date_str)
    if 'error' in result:
        return jsonify(result), 404
    return jsonify(result)


@app.route('/api/short-term/roster-board')
def st_roster_board():
    """Returns combined roster data for all short-term dates (D+1 to D+3)."""
    today = datetime.now()
    
    dates = []
    days_data = []
    
    # We look for D+1, D+2, D+3
    for i in [1, 2, 3]:
        d = today + timedelta(days=i)
        date_iso = d.strftime('%Y-%m-%d')
        if _pax_source_date_for(d) is not None:
            dates.append({
                'date': date_iso,
                'label': d.strftime('%a %d')
            })
            days_data.append(_get_short_term_schedule(date_iso))
            
    if not days_data:
        return jsonify({'error': 'No data available for short-term dates'}), 404
        
    # Aggregate by employee
    emp_map = {}
    day_stats = {}
    for i, day in enumerate(days_data):
        date_key = dates[i]['date']
        day_stats[date_key] = {
            'staff_count': len(day.get('staff', [])),
            'coverage_pct': day.get('kpis', {}).get('coverage_pct', 0),
            'absent': len(day.get('absent_staff', [])),
        }
        # Regular staff
        for s in day.get('staff', []):
            eid = s['id']
            if eid not in emp_map:
                emp_map[eid] = {
                    'id': eid,
                    'name': eid,
                    'skill': s.get('skill1', ''),
                    'shifts': {}
                }

            start = int(s.get('shift_start', 0) or 0) % 1440
            end = int(s.get('shift_end', 720) or 720)
            roster_templates = {
                0:    {'id': 'Early',   'color': '#f97316', 'code': 'E', 'end': 720},
                360:  {'id': 'Mid',     'color': '#3b82f6', 'code': 'M', 'end': 1080},
                720:  {'id': 'Late',    'color': '#8b5cf6', 'code': 'L', 'end': 1440},
                960:  {'id': 'Evening', 'color': '#10b981', 'code': 'V', 'end': 1680},
                1320: {'id': 'Night',   'color': '#ec4899', 'code': 'N', 'end': 2040},
            }
            tmpl = roster_templates.get(start)
            if tmpl:
                tmpl_id = tmpl['id']
                tmpl_color = tmpl['color']
                stype = tmpl_id.upper()
                code = tmpl['code']
                end = tmpl['end']
            else:
                tmpl_id = 'OTHER'
                tmpl_color = '#6b7280'
                stype, code = 'OTHER', 'O'

            timings = f"{mins_to_time(start)}-{mins_to_time(end)}"
            shift_lbl = f"{tmpl_id} {timings}" if tmpl_id != 'OTHER' else s.get('shift_label', f"{tmpl_id} {timings}")
            emp_map[eid]['shifts'][date_key] = {
                'label':          f"{code} {timings}",
                'shift_label':    shift_lbl,
                'template_id':    tmpl_id,
                'template_color': tmpl_color,
                'type':           stype,
                'timings':        timings,
                'shift_start':    start,
                'shift_end':      end,
                'is_absent':      False,
            }

        # Absent staff
        for s in day.get('absent_staff', []):
            eid = s['id']
            if eid not in emp_map:
                emp_map[eid] = {
                    'id': eid,
                    'name': eid,
                    'skill': s.get('skill1', ''),
                    'shifts': {}
                }
            emp_map[eid]['shifts'][date_key] = {
                'label':       'LEAVE',
                'shift_label': s.get('leave_type', 'Absent'),
                'template_id': 'LEAVE',
                'template_color': '#6b7280',
                'type':        'LEAVE',
                'timings':     s.get('leave_type', 'Absent'),
                'is_absent':   True,
            }

    # For employees who are "OFF" on some days, fill in the blanks
    for eid in emp_map:
        for d in dates:
            dk = d['date']
            if dk not in emp_map[eid]['shifts']:
                emp_map[eid]['shifts'][dk] = {
                    'label': 'OFF', 'shift_label': 'Off duty',
                    'template_id': 'OFF', 'template_color': '#374151',
                    'type': 'OFF', 'timings': '', 'is_absent': False,
                }

    # Convert to sorted list
    employees = sorted(emp_map.values(), key=lambda x: x['id'])

    return jsonify({
        'dates':     dates,
        'employees': employees,
        'day_stats': day_stats,
    })


@app.route('/api/short-term/apply-rec', methods=['POST'])
def st_apply_rec():
    """Accept a recommendation: add staff_ids to task_id for a date, re-optimise."""
    body = request.get_json(force=True) or {}
    date  = body.get('date', '')
    task_id = body.get('task_id', '')
    staff_ids = body.get('staff_ids', [])
    if not date or not task_id:
        return jsonify({'error': 'date and task_id required'}), 400
    if date not in _manual_assigns:
        _manual_assigns[date] = {}
    existing = _manual_assigns[date].get(task_id, [])
    for sid in staff_ids:
        if sid not in existing:
            existing.append(sid)
    _manual_assigns[date][task_id] = existing
    result = _get_short_term_schedule(date)
    if 'error' in result:
        return jsonify(result), 404
    return jsonify(result)


@app.route('/api/short-term/constraints', methods=['GET', 'POST'])
def st_constraints():
    """Get or update short-term planning constraints."""
    global _st_custom_constraints, _st_custom_constraints_by_date

    if request.method == 'POST':
        body = request.get_json(force=True) or {}
        date = body.get('date') # can be specific date or global
        next_constraints = {k: v for k, v in body.items() if k != 'date'}
        
        if date:
            if date not in _st_custom_constraints_by_date:
                _st_custom_constraints_by_date[date] = dict(_st_custom_constraints)
            _st_custom_constraints_by_date[date].update(next_constraints)
            # Re-optimise from the current constraints, not from previously pinned
            # recommendation overrides, so "Update Schedule" performs a full reallocation.
            _manual_assigns.pop(date, None)
            result = _get_short_term_schedule(date, preserve_manual_assigns=False)
            if 'error' in result:
                return jsonify(result), 404
            return jsonify(result)
        else:
            _st_custom_constraints.update(next_constraints)
            return jsonify(_st_custom_constraints)

    date = request.args.get('date')
    base = _st_custom_constraints_by_date.get(date, _st_custom_constraints) if date else _st_custom_constraints

    res = {
        'tt_t1_t2': base.get('tt_t1_t2', 15),
        'tt_skill_switch': base.get('tt_skill_switch', 10),
        'allow_overlap': base.get('allow_overlap', False),
        'allow_overlaps': base.get('allow_overlaps', base.get('allow_overlap', False)),
        'use_primary_first': base.get('use_primary_first', True),
        'shift_duration_hrs': base.get('shift_duration_hrs', 12),
        'b1_duration_mins': base.get('b1_duration_mins', 30),
        'b2_duration_mins': base.get('b2_duration_mins', 60),
        'leave_types_excluded': base.get('leave_types_excluded', ["Annual Leave", "Paternity Leave", "Jury Duty", "Sick Leave", "Training"]),
        'permitted_shifts': base.get('permitted_shifts')
    }
    return jsonify(res)


@app.route('/api/intraday')
def intraday_get():
    """Return todays intraday-optimised schedule with any live overrides."""
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    man = _manual_assigns.get(today_str, {})
    current_time_mins = now.hour * 60 + now.minute
    result = optimize_day(today_str, _intraday_overrides, man,
                          current_time_mins=current_time_mins,
                          prefer_early=True,
                          custom_constraints=_intraday_custom_constraints)
    if 'error' in result:
        return jsonify(result), 404
    return jsonify(result)


@app.route('/api/intraday/delay', methods=['POST'])
def intraday_delay():
    """Apply a delay or cancellation to a flight and re-optimise today."""
    body = request.get_json(force=True) or {}
    flight_no  = body.get('flight_no', '').strip()
    delay_mins = int(body.get('delay_mins', 0))
    cancelled  = bool(body.get('cancelled', False))
    if not flight_no:
        return jsonify({'error': 'flight_no required'}), 400
    _intraday_overrides[flight_no] = {'delay_mins': delay_mins, 'cancelled': cancelled}
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    man = _manual_assigns.get(today_str, {})
    current_time_mins = now.hour * 60 + now.minute
    result = optimize_day(today_str, _intraday_overrides, man,
                          current_time_mins=current_time_mins,
                          prefer_early=True,
                          custom_constraints=_intraday_custom_constraints)
    if 'error' in result:
        return jsonify(result), 404
    return jsonify(result)


@app.route('/api/intraday/reset', methods=['POST'])
def intraday_reset():
    """Clear live intraday overrides and reload today's schedule."""
    _intraday_overrides.clear()
    today_str = datetime.now().strftime('%Y-%m-%d')
    if today_str in _manual_assigns:
        _manual_assigns.pop(today_str, None)
    if today_str in _manual_unassigns:
        _manual_unassigns.pop(today_str, None)
    if today_str in _manual_block_assigns:
        _manual_block_assigns.pop(today_str, None)
    return intraday_get()


@app.route('/api/intraday/assign', methods=['POST'])
def intraday_assign():
    """Manually assign or unassign a staff member to/from a task for today."""
    body    = request.get_json(force=True) or {}
    task_id = body.get('task_id', '').strip()
    staff_id = body.get('staff_id', '').strip()
    action  = body.get('action', 'assign')  # 'assign', 'unassign', or legacy 'remove'
    if action == 'remove':
        action = 'unassign'
    if not task_id or not staff_id:
        return jsonify({'error': 'task_id and staff_id required'}), 400
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    if today_str not in _manual_assigns:
        _manual_assigns[today_str] = {}
    if today_str not in _manual_unassigns:
        _manual_unassigns[today_str] = {}
    if today_str not in _manual_block_assigns:
        _manual_block_assigns[today_str] = {}
    existing = _manual_assigns[today_str].get(task_id, [])
    blocked = _manual_unassigns[today_str].get(task_id, [])
    if action == 'assign':
        blocked = [x for x in blocked if x != staff_id]
        if staff_id not in existing:
            existing.append(staff_id)
    elif action == 'unassign':
        existing = [x for x in existing if x != staff_id]
        if staff_id not in blocked:
            blocked.append(staff_id)
    _manual_assigns[today_str][task_id] = existing
    _manual_unassigns[today_str][task_id] = blocked
    
    current_time_mins = now.hour * 60 + now.minute
    result = optimize_day(today_str, _intraday_overrides, _manual_assigns.get(today_str, {}),
                          current_time_mins=current_time_mins,
                          prefer_early=True,
                          custom_constraints=_intraday_custom_constraints)
    if 'error' in result:
        return jsonify(result), 404
    return jsonify(result)


@app.route('/api/intraday/assign-block', methods=['POST'])
def intraday_assign_block():
    """Assign or unassign a staff member for the selected hourly allocation block."""
    body       = request.get_json(force=True) or {}
    staff_id   = body.get('staff_id', '').strip()
    skill      = body.get('skill', '').strip()
    terminal   = body.get('terminal', 'ALL').strip()
    blk_start  = int(body.get('block_start', 0))
    blk_end    = int(body.get('block_end', blk_start + 60))
    action     = body.get('action', 'assign')   # 'assign' | 'unassign' | legacy 'remove'
    if action == 'remove':
        action = 'unassign'
    if not staff_id:
        return jsonify({'error': 'staff_id required'}), 400
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    cur_mins  = now.hour * 60 + now.minute
    if today_str not in _manual_assigns:
        _manual_assigns[today_str] = {}
    if today_str not in _manual_unassigns:
        _manual_unassigns[today_str] = {}
    if today_str not in _manual_block_assigns:
        _manual_block_assigns[today_str] = {}
    # Find all matching task IDs by running a quick schedule pass
    _tmp = optimize_day(today_str, _intraday_overrides,
                        _manual_assigns.get(today_str, {}),
                        current_time_mins=cur_mins, prefer_early=True,
                        custom_constraints=_intraday_custom_constraints)
    staff_member = next((s for s in _tmp.get('staff', []) if str(s.get('id', '')) == staff_id), None)
    if action == 'assign':
        if not staff_member:
            return jsonify({'error': 'Staff member is not on duty for this schedule.'}), 400
        if not _staff_has_task_skill(staff_member, skill):
            return jsonify({'error': 'Staff member does not have this task skill in skill1-skill4.'}), 400
        if not _staff_shift_covers_block(staff_member, blk_start, blk_end):
            return jsonify({'error': 'Staff shift does not cover this whole block.'}), 400
    matched_ids = []
    changed_ids = []
    matching_tasks = _matching_block_tasks(_tmp, skill, terminal, blk_start, blk_end)
    matched_ids = [t.get('id') for t in matching_tasks if t.get('id')]
    if not matched_ids:
        return jsonify({'error': 'No matching tasks found for this block.'}), 404
    hour_tasks = [
        t for t in _tmp.get('tasks', [])
        if (t.get('start_mins') or 0) < blk_end and (t.get('end_mins') or 0) > blk_start
    ]

    if action == 'assign':
        block_required = _block_required_fte(matching_tasks)
        block_assigned = _block_assigned_staff(matching_tasks)
        if staff_id not in block_assigned and len(block_assigned) >= block_required:
            return jsonify({'error': f'Selected block already has required FTE ({block_required}). Use extra FTE on another gap block.'}), 409
        busy_elsewhere = _staff_busy_outside_tasks(_tmp, staff_id, matched_ids, blk_start, blk_end)
        if busy_elsewhere:
            busy = busy_elsewhere[0]
            return jsonify({'error': f'Staff is still assigned to {busy.get("skill", "another task")} {busy.get("start", "")}-{busy.get("end", "")}. Remove them there first.'}), 409

    block_key = _manual_block_key(skill, terminal, blk_start, blk_end)
    block_entry = _manual_block_assigns[today_str].setdefault(block_key, {
        'skill': skill,
        'terminal': terminal,
        'block_start': blk_start,
        'block_end': blk_end,
        'staff_ids': [],
    })
    if action == 'assign' and staff_id not in block_entry['staff_ids']:
        block_entry['staff_ids'].append(staff_id)
    elif action == 'unassign':
        block_entry['staff_ids'] = [sid for sid in block_entry.get('staff_ids', []) if sid != staff_id]

    tasks_to_update = matching_tasks if action == 'assign' else hour_tasks
    for _t in tasks_to_update:
            tid = _t['id']
            existing = list(_manual_assigns[today_str].get(tid, []))
            blocked = list(_manual_unassigns[today_str].get(tid, []))
            if action == 'assign':
                before_existing = list(existing)
                before_blocked = list(blocked)
                blocked = [x for x in blocked if x != staff_id]
                if staff_id not in existing:
                    existing.append(staff_id)
                if existing != before_existing or blocked != before_blocked:
                    changed_ids.append(tid)
            elif action == 'unassign':
                before_existing = list(existing)
                before_blocked = list(blocked)
                existing = [x for x in existing if x != staff_id]
                if staff_id not in blocked:
                    blocked.append(staff_id)
                if existing != before_existing or blocked != before_blocked:
                    changed_ids.append(tid)
            _manual_assigns[today_str][tid] = existing
            _manual_unassigns[today_str][tid] = blocked
    result = optimize_day(today_str, _intraday_overrides,
                          _manual_assigns.get(today_str, {}),
                          current_time_mins=cur_mins, prefer_early=True,
                          custom_constraints=_intraday_custom_constraints)
    if 'error' in result:
        return jsonify(result), 404
    verified_count = _count_staff_in_block(result, staff_id, skill, terminal, blk_start, blk_end)
    result['move_status'] = {
        'action': action,
        'staff_id': staff_id,
        'skill': skill,
        'terminal': terminal,
        'block_start': blk_start,
        'block_end': blk_end,
        'matched_tasks': len(matched_ids),
        'changed_tasks': len(changed_ids),
        'verified_assigned_tasks': verified_count,
        'applied': verified_count > 0 if action == 'assign' else verified_count == 0,
    }
    if action == 'assign' and verified_count == 0:
        result['move_status']['error'] = 'Staff could not be assigned to this block. Check skill, shift, break, or duty eligibility.'
    return jsonify(result)


@app.route('/api/intraday/constraints', methods=['GET', 'POST'])
def intraday_constraints():
    """Get or update intraday constraints."""
    global _intraday_custom_constraints
    
    path = os.path.join(BASE_DIR, 'Roster_constraints.json')
    try:
        with open(path, 'r', encoding='utf-8') as f:
            base_config = json.load(f)
    except:
        base_config = {}

    if request.method == 'POST':
        body = request.get_json(force=True) or {}
        _intraday_custom_constraints.update(body)
        # trigger an update by fetching
        return intraday_get()

    res = {
        'tt_t1_t2': _intraday_custom_constraints.get('tt_t1_t2', 15),
        'tt_skill_switch': _intraday_custom_constraints.get('tt_skill_switch', 10),
        'allow_overlap': _intraday_custom_constraints.get('allow_overlap', False),
        'use_primary_first': _intraday_custom_constraints.get('use_primary_first', True),
        'shift_duration_hrs': _intraday_custom_constraints.get('shift_duration_hrs', 12),
        'b1_duration_mins': _intraday_custom_constraints.get('b1_duration_mins', 30),
        'b2_duration_mins': _intraday_custom_constraints.get('b2_duration_mins', 60),
        'leave_types_excluded': _intraday_custom_constraints.get('leave_types_excluded', ["Annual Leave", "Paternity Leave", "Jury Duty", "Sick Leave", "Training"]),
        'permitted_shifts': _intraday_custom_constraints.get('permitted_shifts'),
        # CP-SAT optimiser controls
        'use_cpsat': _intraday_custom_constraints.get('use_cpsat', False),
        'cpsat_available': _CPSAT_AVAILABLE,
    }
    return jsonify(res)


# ===========================================================================
# UNIFIED INTRADAY OPTIMISE ENDPOINT
# ===========================================================================

@app.route('/api/intraday/optimise', methods=['POST'])
def intraday_optimise():
    """Apply all constraints + run roster optimiser → return full updated intraday schedule.

    Body:
      use_cpsat, min_rest_hrs,
      tt_t1_t2, tt_skill_switch, use_primary_first, allow_overlaps,
      shift_duration_hrs, b1_duration_mins, b2_duration_mins,
      leave_types_excluded, permitted_shifts
    """
    global _intraday_custom_constraints

    try:
        body = request.get_json(force=True) or {}

        use_cpsat = bool(body.get('use_cpsat', False))
        min_rest_hrs = float(body.get('min_rest_hrs', 11))

        # ── 1. Persist tactical constraints ──────────────────────────────
        tactical_keys = [
            'tt_t1_t2', 'tt_skill_switch', 'use_primary_first', 'allow_overlaps',
            'allow_overlap', 'shift_duration_hrs', 'b1_duration_mins', 'b2_duration_mins',
            'leave_types_excluded', 'permitted_shifts', 'use_cpsat',
        ]
        for k in tactical_keys:
            if k in body:
                _intraday_custom_constraints[k] = body[k]

        # ── 2. Re-run intraday schedule with updated constraints ──────────
        now = datetime.now()
        today_str = now.strftime('%Y-%m-%d')
        man = _manual_assigns.get(today_str, {})
        current_time_mins = now.hour * 60 + now.minute
        result = optimize_day(today_str, _intraday_overrides, man,
                              current_time_mins=current_time_mins,
                              prefer_early=True,
                              custom_constraints=_intraday_custom_constraints)
        if 'error' in result:
            return jsonify(result), 404

        roster_info = {'roster_available': False, 'solver_used': 'none'}

        # ── 3. Run roster optimiser and merge assignments ─────────────────
        if _ROSTER_AVAILABLE:
            try:
                on_duty = result.get('staff', [])
                flights = result.get('flights', [])

                all_tasks = result.get('tasks') or [t for f in flights for t in f.get('tasks', [])]
                demand_windows = _roster_tasks_to_dw(all_tasks) if all_tasks else []

                if not demand_windows:
                    anchor_skill = on_duty[0].get('skill1', 'GNIB') if on_duty else 'GNIB'
                    demand_windows = [
                        _RosterDemandWindow(240,  960,  anchor_skill, 1, 'Standard'),
                        _RosterDemandWindow(960, 1440,  anchor_skill, 1, 'Standard'),
                    ]

                staff_norm = [
                    {
                        'id':               s.get('id', s.get('name', '')),
                        'skill1':           s.get('skill1', ''),
                        'skill2':           s.get('skill2', ''),
                        'shift_start_mins': s.get('shift_start_mins', 240),
                        'shift_end_mins':   s.get('shift_end_mins',   960),
                    }
                    for s in on_duty
                ]

                permitted_shifts = _intraday_custom_constraints.get('permitted_shifts')
                if permitted_shifts:
                    permitted_starts = [int(p[0]) for p in permitted_shifts]
                else:
                    permitted_starts = [0, 180, 420, 720]

                roster_constraints = {
                    'shift_duration_hrs': _intraday_custom_constraints.get('shift_duration_hrs', 12),
                    'min_rest_mins':      int(min_rest_hrs * 60),
                    'b1_duration_mins':   _intraday_custom_constraints.get('b1_duration_mins', 30),
                    'b2_duration_mins':   _intraday_custom_constraints.get('b2_duration_mins', 60),
                    'permitted_starts':   permitted_starts,
                }

                rr = _roster_generate(
                    demand_windows=demand_windows,
                    staff_list=staff_norm,
                    constraints=roster_constraints,
                    use_mip=False  # Intraday always uses Greedy for roster phase
                )

                roster_map = {e['id']: e for e in (rr.get('roster') or [])}
                for s in result['staff']:
                    sid = s.get('id', s.get('name', ''))
                    is_mezz = 'Mezz Operation' in [
                        s.get('skill1', ''), s.get('skill2', ''),
                        s.get('skill3', ''), s.get('skill4', ''),
                    ]
                    if is_mezz:
                        continue
                    entry = roster_map.get(sid)
                    if entry and entry.get('pattern_id') != 'unassigned':
                        s['shift_label']      = entry.get('shift_label', s.get('shift', ''))
                        s['pattern_id']       = entry.get('pattern_id', '')
                        s['skill_match']      = entry.get('skill_match', 'primary')
                        s['utilisation_pct']  = entry.get('utilisation_pct', 0)
                        entry_shift_start = entry.get('shift_start_mins', entry.get('shift_start'))
                        entry_shift_end = entry.get('shift_end_mins', entry.get('shift_end'))
                        if entry_shift_start is not None:
                            s['shift_start_mins'] = entry_shift_start
                            s['shift_start'] = entry_shift_start
                        if entry_shift_end is not None:
                            s['shift_end_mins'] = entry_shift_end
                            s['shift_end'] = entry_shift_end
                        if entry.get('breaks'):
                            s['breaks'] = entry['breaks']

                task_solver = 'CP-SAT' if (use_cpsat and _CPSAT_AVAILABLE) else 'Greedy'
                roster_info = {
                    'roster_available': True,
                    'solver_used':   rr.get('solver_used', 'greedy'),
                    'task_solver':   task_solver,
                    'cpsat_available': _CPSAT_AVAILABLE,
                    'pattern_count': rr.get('pattern_count', 0),
                    'patterns':      rr.get('patterns', []),
                    'fairness':      rr.get('fairness', {}),
                    'coverage':      rr.get('coverage', {}),
                    'flags':         rr.get('flags', []),
                    'utilisation':   rr.get('utilisation', {}),
                }
            except Exception as exc:
                roster_info = {'roster_available': False, 'error': str(exc)}

        enforce_break_conflicts(result)
        refill_unassigned_tasks(result)
        enforce_fte_shift_containment(result)
        enforce_break_conflicts(result)
        result['roster'] = roster_info
        result['constraints_applied'] = {
            k: _intraday_custom_constraints.get(k) for k in tactical_keys
        }
        return jsonify(result)

    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


# ===========================================================================
# SCENARIO PLANNING — MONTE CARLO SIMULATION ENGINE
# ===========================================================================

import math
import random
from datetime import datetime as _dt

# ── In-memory scenario store ─────────────────────────────────────────────────
_scenarios = {}          # {scenario_id: scenario_dict}
_scenario_seq = [0]      # auto-increment id counter

DEFAULT_CONSTRAINTS = {
    # ── Demand-side ─────────────────────────────────────────────────────────
    'surge_demand_factor':  1.0,    # Demand multiplier
    
    # ── Supply-side: staffing ────────────────────────────────────────────────
    'staff_utilisation':    0.80,   # target staff utilization
    'absence_rate':         0.06,   # mean fraction of staff absent
    'absence_cv':           0.02,   # CV of absence rate
    'cross_training_rate':  0.15,   # fraction of each skill pool cross-trained
    'new_hire_fraction':    0.00,   # fraction of workforce who are new hires
    'contractor_staff':     {'GNIB': 0, 'CBP Pre-clearance': 0, 'Arr Customer Service': 0,
                             'Check-in/Trolleys': 0, 'Dep / Trolleys': 0, 'T1/T2 Trolleys L/UL': 0,
                             'Transfer Corridor': 0, 'Ramp / Marshalling': 0, 'Bussing': 0,
                             'PBZ': 0, 'Mezz Operation': 0, 'Litter Picking': 0},
    'extra_staff':          {'GNIB': 0, 'CBP Pre-clearance': 0, 'Arr Customer Service': 0,
                             'Check-in/Trolleys': 0, 'Dep / Trolleys': 0, 'T1/T2 Trolleys L/UL': 0,
                             'Transfer Corridor': 0, 'Ramp / Marshalling': 0, 'Bussing': 0,
                             'PBZ': 0, 'Mezz Operation': 0, 'Litter Picking': 0},
    # ── Simulation ───────────────────────────────────────────────────────────
    'n_runs':               500,
}


def _box_muller():
    """Return one standard-normal sample using Box-Muller (no numpy needed)."""
    u1 = max(random.random(), 1e-12)
    u2 = random.random()
    return math.sqrt(-2 * math.log(u1)) * math.cos(2 * math.pi * u2)


def _rnorm(mean, std):
    return mean + std * _box_muller()


def _percentile(data, p):
    if not data:
        return 0.0
    s = sorted(data)
    k = (len(s) - 1) * p / 100.0
    lo, hi = int(k), min(int(k) + 1, len(s) - 1)
    return s[lo] + (s[hi] - s[lo]) * (k - lo)


def _histogram(data, n_bins=20):
    """Return list of {x, count} bin dicts for a histogram."""
    if not data:
        return []
    lo, hi = min(data), max(data)
    if hi == lo:
        return [{'x': round(lo, 3), 'count': len(data)}]
    width = (hi - lo) / n_bins
    counts = [0] * n_bins
    for v in data:
        idx = min(int((v - lo) / width), n_bins - 1)
        counts[idx] += 1
    return [{'x': round(lo + (i + 0.5) * width, 3), 'count': counts[i]}
            for i in range(n_bins)]


def _extract_baseline(result):
    """Extract per-skill required & available staff-minutes from optimizer result."""
    skill_req   = defaultdict(float)   # total staff-mins required per skill
    skill_avail = defaultdict(float)   # total staff-mins available per skill

    for flight in result.get('flights', []):
        for task in flight.get('tasks', []):
            dur = max(task['end_mins'] - task['start_mins'], 1)
            skill_req[task['skill']] += task['staff_needed'] * dur

    for s in result.get('staff', []):
        shift_dur = s['shift_end'] - s['shift_start']
        skill_avail[s['skill1']] += shift_dur

    return dict(skill_req), dict(skill_avail)


def run_scenario_projection(start_date, end_date, constraints, n_runs=500):
    demand, staff_req, skill_req, staff_avail, skill_avail = get_data()
    
    # Filter weeks based on date range
    try:
        s_date = _dt.strptime(start_date, '%Y-%m-%d')
        e_date = _dt.strptime(end_date, '%Y-%m-%d')
    except Exception as e:
        return {'error': f'Invalid date format: {e}'}
    
    if s_date > e_date:
        return {'error': 'Invalid date range'}

    included_weeks = []
    # Identify weeks
    for wk_key in staff_req.keys():
        try:
            year_str, w_str = wk_key.split('-W')
            wk_d = _dt.strptime(f'{year_str}-W{int(w_str):02d}-1', '%G-W%V-%u')
        except:
            continue
        if s_date <= wk_d <= e_date:
            included_weeks.append(wk_key)

    included_weeks.sort() # Ensure chronological order for month map and charts

    if not included_weeks:
        return {'error': 'No data found for the selected date range'}

    surge_factor = float(constraints.get('surge_demand_factor', 1.0))
    target_util  = float(constraints.get('staff_utilisation', 0.80))
    absence_rate = float(constraints.get('absence_rate', 0.06))
    absence_cv   = float(constraints.get('absence_cv', 0.02))
    cross_rate   = float(constraints.get('cross_training_rate', 0.15))
    new_hire_frac= float(constraints.get('new_hire_fraction', 0.00))
    contractor_staff = constraints.get('contractor_staff', {})
    extra_staff      = constraints.get('extra_staff', {})
    total_extra_staff = sum(max(0.0, float(v or 0)) for v in extra_staff.values())
    total_contractor_staff = sum(max(0.0, float(v or 0)) for v in contractor_staff.values())

    new_hire_prod = 1.0 - max(0.0, min(1.0, new_hire_frac)) * 0.30

    skills = [
        'GNIB', 'CBP Pre-clearance', 'Arr Customer Service', 'Check-in/Trolleys',
        'Dep / Trolleys', 'T1/T2 Trolleys L/UL', 'Transfer Corridor',
        'Ramp / Marshalling', 'Bussing', 'PBZ', 'Mezz Operation', 'Litter Picking',
    ]

    # Group weeks by month
    month_map = defaultdict(list)
    for wk in included_weeks:
        year_str, w_str = wk.split('-W')
        wk_d = _dt.strptime(f'{year_str}-W{int(w_str):02d}-1', '%G-W%V-%u')
        mk = wk_d.strftime('%b %Y')
        month_map[mk].append(wk)
        
    monthly_risk = {}
    monthly_fte_breakdown = {}
    overall_coverage_runs = []
    overall_util_runs = []
    
    comparison_data = {
        'months': [],
        'demand_fte': [],
        'scenario_fte_req': [],
        'scenario_fte_avail': [],
        'current_fte': []
    }

    def stats(data):
        n = len(data)
        if n == 0: return {'mean': 0, 'p10': 0, 'p50': 0, 'p90': 0}
        s = sorted(data)
        return {
            'mean': sum(s)/n,
            'p10': _percentile(s, 10),
            'p50': _percentile(s, 50),
            'p90': _percentile(s, 90)
        }

    for month_label, weeks in month_map.items():
        month_cov_runs = []
        month_skill_cov = defaultdict(list)
        
        month_base_req = 0
        month_base_avail = 0
        month_scen_req = 0
        month_scen_avail = 0

        month_sk_exp_req = {sk: 0 for sk in skills}
        month_sk_exp_avail = {sk: 0 for sk in skills}
        month_sk_base_req = {sk: 0 for sk in skills}
        month_sk_base_avail = {sk: 0 for sk in skills}

        for wk in weeks:
            req_sk = skill_req.get(wk, {})
            avail_sk = skill_avail.get(wk, {})
            
            wk_base_req = float(staff_req.get(wk, 0))
            wk_base_avail = float(staff_avail.get(wk, 0))
            month_base_req += wk_base_req
            month_base_avail += wk_base_avail

            for _ in range(n_runs):
                absence_f = max(0.0, min(0.6, _rnorm(absence_rate, absence_cv)))
                contr_f = max(0.4, min(1.0, _rnorm(0.85, 0.10)))
                
                run_scen_req = wk_base_req * surge_factor
                
                sk_perm = {}
                for sk in skills:
                    base = avail_sk.get(sk, 0) + extra_staff.get(sk, 0)
                    sk_perm[sk] = base * (1.0 - absence_f) * new_hire_prod * target_util

                base_total_avail = (wk_base_avail + total_extra_staff) * (1.0 - absence_f) * new_hire_prod * target_util
                contractor_total_avail = total_contractor_staff * contr_f * target_util
                run_scen_avail = base_total_avail + contractor_total_avail
                
                total_perm = sum(sk_perm.values())
                flex_per_skill = (total_perm * cross_rate) / len(skills) if skills else 0.0

                for sk in skills:
                    r = req_sk.get(sk, 0) * surge_factor
                    c_avail = contractor_staff.get(sk, 0) * contr_f * target_util
                    a = sk_perm[sk] + c_avail + flex_per_skill

                    cov = min(a / r, 1.0) if r > 0 else 1.0
                    month_skill_cov[sk].append(cov)

                wk_cov = min(run_scen_avail / run_scen_req, 1.0) if run_scen_req > 0 else 1.0
                wk_util = min(run_scen_req / run_scen_avail, 1.0) if run_scen_avail > 0 else 1.0
                month_cov_runs.append(wk_cov)
                overall_coverage_runs.append(wk_cov)
                overall_util_runs.append(wk_util)

            wk_exp_req = 0
            wk_exp_avail = 0
            for sk in skills:
                wk_sk_base_req = req_sk.get(sk, 0)
                month_sk_base_req[sk] += wk_sk_base_req
                
                wk_sk_req = wk_sk_base_req * surge_factor
                wk_sk_avail = ((avail_sk.get(sk, 0) + extra_staff.get(sk, 0)) * (1.0 - absence_rate) * new_hire_prod * target_util) + (contractor_staff.get(sk, 0) * 0.85 * target_util)
                month_sk_exp_req[sk] += wk_sk_req
                month_sk_exp_avail[sk] += wk_sk_avail
                wk_sk_base_avail = avail_sk.get(sk, 0)
                month_sk_base_avail[sk] += wk_sk_base_avail

            wk_exp_req = wk_base_req * surge_factor
            wk_exp_avail = (
                (wk_base_avail + total_extra_staff) * (1.0 - absence_rate) * new_hire_prod * target_util
            ) + (total_contractor_staff * 0.85 * target_util)

            month_scen_req += wk_exp_req
            month_scen_avail += wk_exp_avail
            
        wn = len(weeks)
        if wn > 0:
            comparison_data['months'].append(month_label)
            comparison_data['demand_fte'].append(month_base_req / wn)
            comparison_data['scenario_fte_req'].append(month_scen_req / wn)
            comparison_data['scenario_fte_avail'].append(month_scen_avail / wn)
            comparison_data['current_fte'].append(month_base_avail / wn)

        monthly_risk[month_label] = {}
        for sk in skills:
            st = stats(month_skill_cov[sk])
            rs = max(0, min(100, ((1 - st['p50']) * 40 + (1 - st['p10']) * 40) * 100))
            rl = 'Low' if rs < 20 else 'Medium' if rs < 45 else 'High' if rs < 65 else 'Critical'
            monthly_risk[month_label][sk] = {
                'p50': st['p50'],
                'risk_score': rs,
                'risk_level': rl
            }
            
        monthly_fte_breakdown[month_label] = {
            'req': {sk: month_sk_exp_req[sk] / wn for sk in skills},
            'avail': {sk: month_sk_exp_avail[sk] / wn for sk in skills},
            'base_req': {sk: month_sk_base_req[sk] / wn for sk in skills},
            'base_avail': {sk: month_sk_base_avail[sk] / wn for sk in skills},
            'base_total_req': month_base_req / wn,
            'base_total_avail': month_base_avail / wn,
            'scenario_total_req': month_scen_req / wn,
            'scenario_total_avail': month_scen_avail / wn,
        }

    overall_cov_st = stats(overall_coverage_runs)
    avg_util = sum(overall_util_runs) / len(overall_util_runs) if overall_util_runs else 0
    
    prob_critical = sum(1 for c in overall_coverage_runs if c < 0.50) / len(overall_coverage_runs) if overall_coverage_runs else 0
    raw_risk = ((1 - overall_cov_st['p50']) * 40 + (1 - overall_cov_st['p10']) * 40 + prob_critical * 20)
    risk_score = round(max(0, min(100, raw_risk * 100)), 1)
    if risk_score < 20: risk_level = 'Low'
    elif risk_score < 45: risk_level = 'Medium'
    elif risk_score < 65: risk_level = 'High'
    else: risk_level = 'Critical'
    
    return {
        'n_runs': n_runs,
        'coverage': overall_cov_st,
        'average_utilisation': avg_util,
        'overall_coverage': overall_cov_st['mean'],
        'median_coverage': overall_cov_st['p50'],
        'risk_score': risk_score,
        'risk_level': risk_level,
        'monthly_risk': monthly_risk,
        'monthly_fte_breakdown': monthly_fte_breakdown,
        'comparison_data': comparison_data
    }


# ── Scenario endpoints ────────────────────────────────────────────────────────

@app.route('/api/long-term/scenario', methods=['POST'])
def lt_run_scenario():
    """Deterministic long-term scenario: applies constraint overrides to baseline LT FTE."""
    body       = request.get_json(silent=True) or {}
    name       = body.get('name', f'Scenario {_scenario_seq[0]+1}').strip() or f'Scenario {_scenario_seq[0]+1}'
    start_date = body.get('start_date', '2026-01-01')
    end_date   = body.get('end_date', '2026-12-31')
    cst        = body.get('constraints', {})

    _PAX_SKILLS = ['Checkin', 'Security', 'CBP', 'Lounge', 'Boarding', 'Immigration', 'Baggage']

    absence_rate  = float(cst.get('absence_rate', 0.06))
    overtime_pct  = float(cst.get('overtime_allowance', 0.0))
    target_util   = float(cst.get('staff_utilisation', 0.80))
    peak_buffer   = float(cst.get('peak_buffer', 0.10))
    cross_rate    = float(cst.get('cross_training_rate', 0.15))
    new_hire_frac = float(cst.get('new_hire_fraction', 0.00))
    surge_factor  = float(cst.get('surge_demand_factor', 1.0))
    cov_floor     = float(cst.get('min_coverage_floor', 0.85))
    new_hire_prod = 1.0 - max(0.0, min(1.0, new_hire_frac)) * 0.30

    try:
        s_date = datetime.strptime(start_date, '%Y-%m-%d')
        e_date = datetime.strptime(end_date, '%Y-%m-%d')
    except Exception:
        return jsonify({'error': 'Invalid date format'}), 400
    if s_date > e_date:
        return jsonify({'error': 'Start date must be before end date'}), 400

    _, staff_req, skill_req, staff_avail, skill_avail = get_data()

    month_map = defaultdict(list)
    for wk_key in sorted(staff_req.keys()):
        try:
            yr, wn = wk_key.split('-W')
            wk_d = datetime.strptime(f'{yr}-W{int(wn):02d}-1', '%G-W%V-%u')
        except Exception:
            continue
        if s_date <= wk_d <= e_date:
            month_map[wk_d.strftime('%b %Y')].append(wk_key)

    if not month_map:
        return jsonify({'error': 'No data found for the selected date range'}), 400

    months_list, total_req_list, total_avail_list = [], [], []
    monthly_breakdown = {}

    for month_label, weeks in month_map.items():
        wn = len(weeks)
        sk_req_sum   = {sk: 0.0 for sk in _PAX_SKILLS}
        sk_avail_sum = {sk: 0.0 for sk in _PAX_SKILLS}

        for wk in weeks:
            req_sk   = skill_req.get(wk, {})
            avail_sk = skill_avail.get(wk, {})
            base_total = sum(avail_sk.get(sk, 0) for sk in _PAX_SKILLS)
            flex_per   = (base_total * (1 - absence_rate) * new_hire_prod
                          * target_util * cross_rate) / len(_PAX_SKILLS)

            for sk in _PAX_SKILLS:
                adj_req   = req_sk.get(sk, 0.0) * surge_factor * (1.0 + peak_buffer)
                adj_avail = (avail_sk.get(sk, 0.0)
                             * (1.0 - absence_rate) * new_hire_prod
                             * target_util * (1.0 + overtime_pct)
                             + flex_per)
                if adj_req > 0:
                    adj_avail = max(adj_avail, cov_floor * adj_req)
                sk_req_sum[sk]   += adj_req
                sk_avail_sum[sk] += adj_avail

        avg_req   = {sk: round(sk_req_sum[sk] / wn, 2)   for sk in _PAX_SKILLS}
        avg_avail = {sk: round(sk_avail_sum[sk] / wn, 2) for sk in _PAX_SKILLS}
        tot_req   = round(sum(avg_req.values()), 1)
        tot_avail = round(sum(avg_avail.values()), 1)

        monthly_breakdown[month_label] = {
            'required':        avg_req,
            'available':       avg_avail,
            'gap':             {sk: round(avg_avail[sk] - avg_req[sk], 2) for sk in _PAX_SKILLS},
            'total_required':  tot_req,
            'total_available': tot_avail,
        }
        months_list.append(month_label)
        total_req_list.append(tot_req)
        total_avail_list.append(tot_avail)

    _scenario_seq[0] += 1
    sid = f'lt_{_scenario_seq[0]:03d}'
    _scenarios[sid] = {
        'id':          sid,
        'name':        name,
        'type':        'long_term',
        'created_at':  datetime.utcnow().isoformat(),
        'start_date':  start_date,
        'end_date':    end_date,
        'constraints': cst,
        'results': {
            'months':             months_list,
            'monthly_breakdown':  monthly_breakdown,
            'total_required':     total_req_list,
            'total_available':    total_avail_list,
            'skills':             _PAX_SKILLS,
        },
    }
    return jsonify(_scenarios[sid])


@app.route('/api/scenarios', methods=['GET'])
def list_scenarios():
    sc_type = request.args.get('type')
    out = []
    for sid, sc in sorted(_scenarios.items(), key=lambda x: x[1]['created_at']):
        if sc_type and sc.get('type', 'monte_carlo') != sc_type:
            continue
        out.append(sc)
    return jsonify(out)


@app.route('/api/scenarios/run', methods=['POST'])
def run_scenario():
    body = request.get_json(force=True) or {}
    name       = body.get('name', f'Scenario {_scenario_seq[0]+1}').strip() or f'Scenario {_scenario_seq[0]+1}'
    start_date = body.get('start_date', _dt.now().strftime('%Y-%m-%d'))
    end_date   = body.get('end_date', _dt.now().strftime('%Y-%m-%d'))
    base_date  = _dt.now().strftime('%Y-%m-%d')
    constraints = {**DEFAULT_CONSTRAINTS, **body.get('constraints', {})}
    # Merge dict-valued constraints carefully so partial overrides work
    body_constraints = body.get('constraints', {})
    constraints['extra_staff'] = {
        **DEFAULT_CONSTRAINTS['extra_staff'],
        **body_constraints.get('extra_staff', {}),
    }
    constraints['contractor_staff'] = {
        **DEFAULT_CONSTRAINTS['contractor_staff'],
        **body_constraints.get('contractor_staff', {}),
    }
    n_runs = int(constraints.get('n_runs', 500))

    results = run_scenario_projection(start_date, end_date, constraints, n_runs)
    if 'error' in results:
        return jsonify(results), 400

    _scenario_seq[0] += 1
    sid = f'sc_{_scenario_seq[0]:03d}'
    _scenarios[sid] = {
        'id':          sid,
        'name':        name,
        'status':      'active',
        'created_at':  _dt.now().strftime('%Y-%m-%dT%H:%M:%S'),
        'base_date':   base_date,
        'start_date':  start_date,
        'end_date':    end_date,
        'constraints': constraints,
        'results':     results,
    }
    return jsonify({'id': sid, **_scenarios[sid]})


@app.route('/api/scenarios/<sid>', methods=['GET'])
def get_scenario(sid):
    sc = _scenarios.get(sid)
    if not sc:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(sc)


@app.route('/api/scenarios/<sid>/finalise', methods=['POST'])
def finalise_scenario(sid):
    sc = _scenarios.get(sid)
    if not sc:
        return jsonify({'error': 'Not found'}), 404
    # Unfinalise all others
    for s in _scenarios.values():
        s['status'] = 'active'
    sc['status'] = 'finalised'
    return jsonify({'id': sid, 'status': 'finalised'})


@app.route('/api/scenarios/<sid>', methods=['DELETE'])
def delete_scenario(sid):
    if sid not in _scenarios:
        return jsonify({'error': 'Not found'}), 404
    del _scenarios[sid]
    return jsonify({'deleted': sid})


def update_csv_dates_to_current():
    """Auto-update CSV dates to start from today."""
    now = datetime.now()
    
    # 1. Update Flights_schedule_4days.csv
    flights_path = os.path.join(BASE_DIR, 'data', 'Flights_schedule_4days.csv')
    if os.path.exists(flights_path):
        with open(flights_path, encoding='cp1252') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            fieldnames = reader.fieldnames
        if rows:
            raw_dates = set(r.get('date', '').strip() for r in rows if r.get('date', '').strip())
            parsed = [(parse_date(d), d) for d in raw_dates if parse_date(d)]
            parsed.sort(key=lambda x: x[0])
            
            if parsed:
                date_map = {}
                for i, (d_obj, d_str) in enumerate(parsed):
                    date_map[d_str] = (now + timedelta(days=i)).strftime('%d-%b-%y')
                for r in rows:
                    if r.get('date', '').strip() in date_map:
                        r['date'] = date_map[r.get('date', '').strip()]
                with open(flights_path, 'w', encoding='cp1252', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)

    # 2. Update Staff_schedule.csv
    staff_path = os.path.join(BASE_DIR, 'data', 'Staff_schedule.csv')
    if os.path.exists(staff_path):
        with open(staff_path, encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            fieldnames = reader.fieldnames
        if rows:
            raw_dates = set(r.get('DATE', '').strip() for r in rows if r.get('DATE', '').strip())
            parsed = [(parse_date(d), d) for d in raw_dates if parse_date(d)]
            parsed.sort(key=lambda x: x[0])
            
            if parsed:
                date_map = {}
                for i, (d_obj, d_str) in enumerate(parsed):
                    date_map[d_str] = (now + timedelta(days=i)).strftime('%d-%m-%Y')
                for r in rows:
                    if r.get('DATE', '').strip() in date_map:
                        r['DATE'] = date_map[r.get('DATE', '').strip()]
                with open(staff_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)

    # 3. Update short term PAX workbook timestamps (shift to start from today)
    pax_path = os.path.join(BASE_DIR, 'data', 'short term PAX.xlsx')
    if os.path.exists(pax_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(pax_path)
            ws = wb.active
            # Read header row and find 'Timestamp' column index (0-based)
            headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            ts_idx = None
            for i, h in enumerate(headers):
                if h.lower() == 'timestamp':
                    ts_idx = i
                    break
            if ts_idx is not None:
                # Collect unique dates in the Timestamp column
                seen_dates = []
                for vals in ws.iter_rows(min_row=2, values_only=True):
                    ts = vals[ts_idx]
                    d = None
                    if isinstance(ts, datetime):
                        d = ts.date()
                    else:
                        parsed = _parse_any_date(ts)
                        if parsed:
                            d = parsed.date()
                    if d and d not in seen_dates:
                        seen_dates.append(d)
                seen_dates.sort()
                if seen_dates:
                    date_map = {seen_dates[i]: (now.date() + timedelta(days=i)) for i in range(len(seen_dates))}
                    # Update cells preserving time-of-day when possible
                    for row in ws.iter_rows(min_row=2):
                        cell = row[ts_idx]
                        ts = cell.value
                        if isinstance(ts, datetime):
                            old = ts.date()
                            newd = date_map.get(old)
                            if newd:
                                cell.value = datetime.combine(newd, ts.time())
                        else:
                            parsed = _parse_any_date(ts)
                            if parsed:
                                old = parsed.date()
                                newd = date_map.get(old)
                                if newd:
                                    # Try to preserve time if the string contained it
                                    try:
                                        new_time = datetime.fromisoformat(str(ts)).time()
                                    except Exception:
                                        new_time = datetime.min.time()
                                    cell.value = datetime.combine(newd, new_time)
                    wb.save(pax_path)
        except Exception:
            # Fail silently — don't break the updater if openpyxl or parsing fails
            pass

    # 4. Update Weekly_flight_demand.csv status (Historical vs Forecast)
    demand_path = os.path.join(BASE_DIR, 'data', 'Weekly_flight_demand.csv')
    if os.path.exists(demand_path):
        with open(demand_path, encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            fieldnames = reader.fieldnames
        if rows:
            changed = False
            for r in rows:
                week_start_str = r.get('Week_Start', '').strip()
                if not week_start_str:
                    continue
                ws_date = parse_date(week_start_str)
                if ws_date:
                    # A week is considered completed if the current time is past its 7-day duration
                    if now >= ws_date + timedelta(days=7):
                        new_type = 'Historical'
                    else:
                        new_type = 'Forecast'
                    
                    if r.get('Data_type') != new_type:
                        r['Data_type'] = new_type
                        changed = True
            
            if changed:
                with open(demand_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)

# ===========================================================================
# UNIFIED SHORT-TERM OPTIMISE ENDPOINT
# ===========================================================================

@app.route('/api/short-term/optimise', methods=['POST'])
def st_optimise():
    """Apply all constraints + run roster optimiser → return full updated schedule.

    Body:
      date, use_mip, min_rest_hrs,
      tt_t1_t2, tt_skill_switch, use_primary_first, allow_overlaps,
      shift_duration_hrs, b1_duration_mins, b2_duration_mins,
      leave_types_excluded, permitted_shifts
    """
    global _st_custom_constraints, _st_custom_constraints_by_date

    body = request.get_json(force=True) or {}
    date = (body.get('date') or '').strip()
    if not date:
        return jsonify({'error': 'date required'}), 400

    use_mip = bool(body.get('use_mip', True))
    min_rest_hrs = float(body.get('min_rest_hrs', 11))

    # ── 1. Persist tactical constraints ──────────────────────────────
    tactical_keys = [
        'tt_t1_t2', 'tt_skill_switch', 'use_primary_first', 'allow_overlaps',
        'shift_duration_hrs', 'b1_duration_mins', 'b2_duration_mins',
        'leave_types_excluded', 'permitted_shifts',
        'min_coverage_pct', 'max_utilisation_pct', 'fairness_weight', 'allow_secondary_skills',
    ]
    if date not in _st_custom_constraints_by_date:
        _st_custom_constraints_by_date[date] = dict(_st_custom_constraints)
        
    for k in tactical_keys:
        if k in body:
            _st_custom_constraints_by_date[date][k] = body[k]

    # Get the active constraints for this date
    active_constraints = _st_custom_constraints_by_date[date]

    # ── 2. Re-run tactical schedule (clears manual overrides for fresh plan) ──
    _manual_assigns.pop(date, None)
    result = _get_short_term_schedule(date, preserve_manual_assigns=False)
    if 'error' in result:
        return jsonify(result), 404

    roster_info = {'roster_available': False, 'solver_used': 'none'}

    # ── 3. Run roster optimiser and merge assignments ─────────────────
    if _ROSTER_AVAILABLE:
        try:
            on_duty = result.get('staff', [])
            flights = result.get('flights', [])

            # Derive demand windows from today's tasks
            all_tasks = result.get('tasks') or [t for f in flights for t in f.get('tasks', [])]
            demand_windows = _roster_tasks_to_dw(all_tasks) if all_tasks else []

            # Fallback: two synthetic anchor windows covering the day
            if not demand_windows:
                anchor_skill = on_duty[0].get('skill1', 'GNIB') if on_duty else 'GNIB'
                demand_windows = [
                    _RosterDemandWindow(240,  960,  anchor_skill, 1, 'Standard'),
                    _RosterDemandWindow(960, 1440,  anchor_skill, 1, 'Standard'),
                ]

            # Normalise staff list for the optimiser
            staff_norm = [
                {
                    'id':               s.get('id', s.get('name', '')),
                    'skill1':           s.get('skill1', ''),
                    'skill2':           s.get('skill2', ''),
                    'shift_start_mins': s.get('shift_start_mins', 240),
                    'shift_end_mins':   s.get('shift_end_mins',   960),
                }
                for s in on_duty
            ]

            permitted_shifts = active_constraints.get('permitted_shifts')
            if permitted_shifts:
                permitted_starts = [int(p[0]) for p in permitted_shifts]
            else:
                permitted_starts = [0, 180, 420, 720]

            roster_constraints = {
                'shift_duration_hrs':  active_constraints.get('shift_duration_hrs', 12),
                'min_rest_mins':       int(min_rest_hrs * 60),
                'b1_duration_mins':    active_constraints.get('b1_duration_mins', 30),
                'b2_duration_mins':    active_constraints.get('b2_duration_mins', 60),
                'permitted_starts':    permitted_starts,
                'fairness_weight':     float(active_constraints.get('fairness_weight', 0.5)),
                'max_utilisation_pct': float(active_constraints.get('max_utilisation_pct', 95)),
            }

            rr = _roster_generate(
                demand_windows=demand_windows,
                staff_list=staff_norm,
                constraints=roster_constraints,
                use_mip=use_mip
            )

            # Merge optimised shift/break assignments back into the staff list.
            # Mezz Operation workers are exempt: their 24h-coverage anchors
            # (00:00-12:00 and 12:00-24:00) must not be overridden by the
            # demand-driven roster optimizer.
            roster_map = {e['id']: e for e in (rr.get('roster') or [])}
            for s in result['staff']:
                sid = s.get('id', s.get('name', ''))
                # Preserve Mezz workers' 24h-anchored shift assignments
                is_mezz = 'Mezz Operation' in [
                    s.get('skill1', ''), s.get('skill2', ''),
                    s.get('skill3', ''), s.get('skill4', ''),
                ]
                if is_mezz:
                    continue
                entry = roster_map.get(sid)
                if entry and entry.get('pattern_id') != 'unassigned':
                    s['shift_label']      = entry.get('shift_label', s.get('shift', ''))
                    s['pattern_id']       = entry.get('pattern_id', '')
                    s['skill_match']      = entry.get('skill_match', 'primary')
                    s['utilisation_pct']  = entry.get('utilisation_pct', 0)
                    entry_shift_start = entry.get('shift_start_mins', entry.get('shift_start'))
                    entry_shift_end = entry.get('shift_end_mins', entry.get('shift_end'))
                    if entry_shift_start is not None:
                        s['shift_start_mins'] = entry_shift_start
                        s['shift_start'] = entry_shift_start
                    if entry_shift_end is not None:
                        s['shift_end_mins'] = entry_shift_end
                        s['shift_end'] = entry_shift_end
                    if entry.get('breaks'):
                        s['breaks'] = entry['breaks']

            roster_info = {
                'roster_available': True,
                'solver_used':   rr.get('solver_used', 'greedy'),
                'mip_available': _ROSTER_SOLVER_AVAILABLE,
                'pattern_count': rr.get('pattern_count', 0),
                'patterns':      rr.get('patterns', []),
                'fairness':      rr.get('fairness', {}),
                'coverage':      rr.get('coverage', {}),
                'flags':         rr.get('flags', []),
                'utilisation':   rr.get('utilisation', {}),
            }
        except Exception as exc:
            roster_info = {'roster_available': False, 'error': str(exc)}

    enforce_break_conflicts(result)
    refill_unassigned_tasks(result)
    enforce_break_conflicts(result)
    _normalize_reallocation_blocks(result, date)

    # ── 4. Soft-constraint post-processing flags ──────────────────
    if roster_info.get('roster_available'):
        extra_flags = roster_info.get('flags', [])

        # Overtime flag: staff exceeding max utilisation threshold
        max_util = float(active_constraints.get('max_utilisation_pct', 95))
        ot_staff = [s['id'] for s in result.get('staff', [])
                    if (s.get('utilisation_pct') or 0) > max_util]
        if ot_staff:
            extra_flags.append({
                'flag_id':  'OVERTIME_RISK',
                'severity': 'warn',
                'staff_id': None,
                'detail':   f'{len(ot_staff)} staff exceed max utilisation ({max_util:.0f}%): '
                            + ', '.join(ot_staff[:5]) + ('…' if len(ot_staff) > 5 else ''),
            })

        # Coverage floor flag
        min_cov = float(active_constraints.get('min_coverage_pct', 80))
        overall_cov = result.get('kpis', {}).get('coverage_pct', 100)
        if overall_cov < min_cov:
            extra_flags.append({
                'flag_id':  'LOW_COVERAGE',
                'severity': 'critical',
                'staff_id': None,
                'detail':   f'Overall coverage {overall_cov:.1f}% is below minimum threshold {min_cov:.0f}%',
            })

        roster_info['flags'] = extra_flags

    result['roster'] = roster_info
    result['constraints_applied'] = {
        k: active_constraints.get(k) for k in tactical_keys
    }
    return jsonify(result)


# ===========================================================================
# ROSTER OPTIMISATION ENDPOINT
# ===========================================================================

@app.route('/api/roster/optimised', methods=['GET', 'POST'])
def roster_optimised():
    """Two-phase roster optimisation: shift-pattern generation → staff assignment.

    GET  ?date=YYYY-MM-DD[&use_mip=true]
    POST { "date": "YYYY-MM-DD",
           "use_mip": true,
           "constraints": {
             "b1_duration_mins": 30,
             "b2_duration_mins": 60,
             "min_rest_mins": 660,
             "shift_duration_hrs": 12
           }
         }

    Phase 1 — Shift-Pattern Generation
      Generates all feasible patterns on a 60-minute grid up to 12 h.
      Scores each by weighted demand coverage; prunes to ≤18 candidates.
      Always retains DAY (00:00–12:00) and NIGHT (12:00–24:00) anchors.

    Phase 2a — Greedy Assignment  (always runs)
      Assigns staff to patterns respecting 11-hour rest, skill fit, and load balance.

    Phase 2b — MIP Refinement  (optional; requires PuLP)
      Minimises skill-mismatch cost + L1 workload deviation + demand coverage gaps.

    Response
    --------
    {
      "date":           "YYYY-MM-DD",
      "solver_used":    "CBC (PuLP MIP)" | "Greedy",
      "mip_status":     "Optimal" | "greedy_only" | …,
      "roster_available": true,
      "staff_count":    42,
      "pattern_count":  8,
      "patterns":       [ {id, label, start_mins, end_mins, net_mins,
                           coverage_score, demand_profile, staff_count} … ],
      "roster":         [ {id, skill1-4, employment, pattern_id, shift_label,
                           shift_start, shift_end, shift_duration_mins,
                           net_working_mins, utilisation_pct, skill_match,
                           breaks, assignments} … ],
      "utilisation":    { staff_id: {gross_mins, net_working_mins,
                                     utilisation_pct, pattern_id, skill_match} },
      "fairness":       { gini_coefficient, mean_utilisation_pct,
                          std_utilisation_pct, min_utilisation_pct,
                          max_utilisation_pct, interpretation },
      "coverage":       { skill: {needed, covered, coverage_pct} },
      "flags":          [ {flag_id, severity, staff_id, detail} … ],
      "absent_staff":   [ {id, skill1, leave_type} … ],
      "constraints_used": { … }
    }
    """
    if not _ROSTER_AVAILABLE:
        return jsonify({
            'error':   'roster_optimizer module not available',
            'message': 'Ensure roster_optimizer.py is present in the project root.',
        }), 503

    # ── Parse parameters ────────────────────────────────────────────────────
    if request.method == 'POST':
        body = request.get_json(force=True) or {}
    else:
        body = {}

    date_str = (
        body.get('date')
        or request.args.get('date', '')
        or datetime.now().strftime('%Y-%m-%d')
    ).strip()

    use_mip_param = body.get('use_mip', request.args.get('use_mip', 'true'))
    use_mip = str(use_mip_param).lower() not in ('false', '0', 'no')

    constraints_override = body.get('constraints', {})
    constraints = {
        'b1_duration_mins':  int(constraints_override.get('b1_duration_mins', 30)),
        'b2_duration_mins':  int(constraints_override.get('b2_duration_mins', 60)),
        'min_rest_mins':     int(constraints_override.get('min_rest_mins', 660)),
        'shift_duration_hrs': int(constraints_override.get('shift_duration_hrs', 12)),
    }
    constraints['max_shift_mins'] = constraints['shift_duration_hrs'] * 60

    # ── Validate date ───────────────────────────────────────────────────────
    parsed_date = None
    for fmt in ('%Y-%m-%d', '%d-%b-%y', '%d-%m-%Y', '%d-%m-%y'):
        try:
            parsed_date = datetime.strptime(date_str, fmt)
            break
        except ValueError:
            pass
    if parsed_date is None:
        return jsonify({'error': f'Invalid date: {date_str!r}'}), 400

    # ── Load staff & absences for the date ─────────────────────────────────
    # We call the existing path (without the optimiser branch) to get the raw
    # staff list and absent staff.  The optimiser branch is applied below after
    # we have built demand windows from the day's flights.
    on_duty_raw, absent_staff, _cf = get_staff_for_date(
        date_str,
        custom_constraints={
            'shift_duration_hrs':  constraints['shift_duration_hrs'],
            'b1_duration_mins':    constraints['b1_duration_mins'],
            'b2_duration_mins':    constraints['b2_duration_mins'],
            'leave_types_excluded': constraints_override.get(
                'leave_types_excluded',
                ['Annual Leave', 'Paternity Leave', 'Jury Duty', 'Sick Leave', 'Training'],
            ),
        },
        use_roster_optimiser=False,   # raw list first; optimiser applied below
    )

    if not on_duty_raw:
        return jsonify({
            'date':           date_str,
            'error':          'No staff scheduled for this date',
            'absent_staff':   absent_staff,
            'roster_available': False,
        }), 404

    # ── Build demand windows from the day's flights ─────────────────────────
    # Re-use optimize_day task generation machinery to derive demand windows.
    demand_windows: list = []
    pax_tasks = build_pax_demand_tasks(parsed_date.strftime('%Y-%m-%d'))
    if pax_tasks:
        demand_windows = _roster_tasks_to_dw(pax_tasks)

    if not demand_windows:
        return jsonify({
            'date': date_str,
            'error': 'No passenger demand available for this date',
            'roster_available': False,
        }), 404

    # ── Normalise staff dicts for the optimiser ─────────────────────────────
    staff_list = [
        {
            'id':         s['id'],
            'skill1':     s.get('skill1', ''),
            'skill2':     s.get('skill2', ''),
            'skill3':     s.get('skill3', ''),
            'skill4':     s.get('skill4', ''),
            'employment': s.get('employment', ''),
        }
        for s in on_duty_raw
    ]

    # ── Run roster optimiser ────────────────────────────────────────────────
    roster_result = _roster_generate(
        demand_windows  = demand_windows,
        staff_list      = staff_list,
        constraints     = constraints,
        prev_shift_ends = {},   # no cross-day state at this endpoint for now
        use_mip         = use_mip,
    )

    # ── Build response ──────────────────────────────────────────────────────
    return jsonify({
        'date':              date_str,
        'roster_available':  True,
        'solver_used':       roster_result.get('solver_used', 'Greedy'),
        'mip_status':        roster_result.get('mip_status', 'greedy_only'),
        'mip_available':     _ROSTER_SOLVER_AVAILABLE,
        'staff_count':       roster_result.get('staff_count', len(staff_list)),
        'pattern_count':     roster_result.get('pattern_count', 0),
        'patterns':          roster_result.get('patterns', []),
        'roster':            roster_result.get('roster', []),
        'utilisation':       roster_result.get('utilisation', {}),
        'fairness':          roster_result.get('fairness', {}),
        'coverage':          roster_result.get('coverage', {}),
        'flags':             roster_result.get('flags', []),
        'absent_staff':      absent_staff,
        'demand_windows':    [
            {
                'start':    dw.start,
                'end':      dw.end,
                'skill':    dw.skill,
                'needed':   dw.needed,
                'priority': dw.priority,
                'task_id':  dw.task_id,
            }
            for dw in demand_windows
        ],
        'constraints_used':  constraints,
    })


# ===========================================================================
# MONTE CARLO SIMULATION ENDPOINT  (/api/simulation/run)
# ===========================================================================

@app.route('/api/simulation/run', methods=['POST'])
def simulation_run():
    """Intraday Monte Carlo stress-test for a staffing plan.

    This is a VALIDATION layer — it does not modify any optimisation logic.
    It perturbs the environment (delays, absences, surge) and measures how
    well the original plan survives across hundreds of random scenarios.

    POST body
    ---------
    {
      "date":     "YYYY-MM-DD",          // date to simulate (required)
      "num_runs": 200,                   // iterations 10-1000 (default 100)
      "params": {
        "delay_sigma_mins":  15,         // std-dev of per-flight delay draw
        "delay_max_mins":    30,         // hard cap on delay magnitude
        "delay_prob":        0.25,       // fraction of flights delayed per run
        "absence_rate_min":  0.05,       // minimum run-level absence fraction
        "absence_rate_max":  0.15,       // maximum run-level absence fraction
        "surge_mean":        1.0,        // expected passenger surge factor
        "surge_sigma":       0.10,       // log-normal sigma for surge draw
        "seed":              null        // int for reproducibility, null = random
      }
    }

    Response
    --------
    {
      "date":        "YYYY-MM-DD",
      "risk_score":  42.3,              // 0-100 composite risk index
      "risk_level":  "Medium",          // Low / Medium / High / Critical
      "summary": {
        "num_runs":             200,
        "unserved_tasks":       {mean, std, p10, p50, p90, p99, ...},
        "staff_utilisation":    {mean, std, p10, p50, p90, p99, ...},
        "critical_failure_probability": 0.12,
        "prob_any_unserved":    0.65,
        "prob_gt10pct_unserved": 0.18,
        ...
      },
      "worst_case": {
        "run_id": 47,
        "unserved_pct": 28.6,
        "absent_count": 9,
        "delayed_flights": {"EI123": 28, ...},
        "failing_tasks":  [{task_id, label, skill, priority, gap, ...}],
        ...
      },
      "bottlenecks": {
        "top_failing_tasks": [{task_id, label, skill, fail_rate_pct, avg_gap}],
        "failing_skills":    [{skill, fail_rate_pct, total_gap_headcount}],
        "tasks_never_failed": 18,
      },
      "distributions": {
        "unserved_pct":     [{x, count}, ...],
        "utilisation_mean": [{x, count}, ...],
        ...
      },
      "baseline": {
        "total_tasks": 35,
        "unserved_tasks": 2,
        "coverage_pct": 94.3,
        ...
      },
      "run_log": [{run_id, unserved_pct, absent_count, ...}],
      "params_used": {...},
      "meta": {elapsed_seconds, runs_per_second, ...}
    }
    """
    if not _SIM_AVAILABLE:
        return jsonify({
            "error":   "simulation_engine module not available",
            "message": "Ensure simulation_engine.py is in the project root.",
        }), 503

    body = request.get_json(force=True) or {}

    # ── Parse date ──────────────────────────────────────────────────────────
    date_str = body.get("date", "").strip()
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")

    parsed_date = None
    for fmt in ("%Y-%m-%d", "%d-%b-%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            parsed_date = datetime.strptime(date_str, fmt)
            break
        except ValueError:
            pass
    if parsed_date is None:
        return jsonify({"error": f"Invalid date: {date_str!r}"}), 400

    # ── Parse simulation parameters ─────────────────────────────────────────
    num_runs = int(body.get("num_runs", 100))
    num_runs = max(10, min(num_runs, 1000))

    sim_params = body.get("params", {})

    # ── Build the base plan via optimize_day (read-only) ────────────────────
    # We call the standard day optimiser to get the current staffing plan.
    # The simulation layer then stress-tests this plan without touching it.
    plan = optimize_day(
        date_str,
        overrides={},
        manual_assigns={},
        custom_constraints=_st_custom_constraints,
    )

    if "error" in plan:
        return jsonify({
            "error": f"Could not build plan for {date_str}: {plan['error']}",
            "date":  date_str,
        }), 404

    if not plan.get("tasks"):
        return jsonify({
            "error": "Plan has no tasks — no flights scheduled for this date?",
            "date":  date_str,
            "kpis":  plan.get("kpis", {}),
        }), 404

    # ── Run Monte Carlo simulation ──────────────────────────────────────────
    try:
        sim_result = _sim_run_simulation(
            plan     = plan,
            num_runs = num_runs,
            params   = sim_params or None,
        )
    except Exception as exc:
        logger.exception("Simulation failed for date %s", date_str)
        return jsonify({"error": str(exc), "date": date_str}), 500

    # ── Attach plan-level context to the response ───────────────────────────
    sim_result["date"]         = date_str
    sim_result["date_label"]   = plan.get("date_label", date_str)
    sim_result["plan_kpis"]    = plan.get("kpis", {})
    sim_result["sim_available"] = True

    return jsonify(sim_result)


@app.route('/api/simulation/status', methods=['GET'])
def simulation_status():
    """Quick availability check for the simulation engine."""
    return jsonify({
        "sim_available": _SIM_AVAILABLE,
        "message": (
            "Monte Carlo simulation engine ready."
            if _SIM_AVAILABLE
            else "simulation_engine.py not found. Ensure it is in the project root."
        ),
    })


# Auto-update CSV dates on start-up (ensures compatibility with WSGI servers like Gunicorn/Render)
update_csv_dates_to_current()

if __name__ == '__main__':
    app.run(debug=True)

