import json
import os
from collections import defaultdict
from datetime import datetime, timedelta


PAX_3DAY_FILE = 'pax_3day_cache.json'
PAX_PROFILE_FILE = 'short term PAX.xlsx'
TOUCHPOINTS = ('checkin', 'security', 'cbp', 'lounge', 'boarding', 'immigration', 'baggage')


def _data_path(base_dir, filename):
    return os.path.join(base_dir, 'data', filename)


def _load_json(base_dir, filename):
    with open(_data_path(base_dir, filename), encoding='utf-8') as f:
        return json.load(f)


def _numbers(values):
    out = []
    for value in values or []:
        try:
            out.append(float(value or 0))
        except (TypeError, ValueError):
            out.append(0.0)
    return out


def _peak_idx(values):
    if not values:
        return 0
    return max(range(len(values)), key=lambda idx: values[idx])


def _active_avg(values):
    active = [v for v in values if v > 0]
    return round(sum(active) / len(active)) if active else 0


def _pct(part, total):
    if not total:
        return None
    return round((part / total) * 100)


def _load_pax_workbook_rows(base_dir):
    path = _data_path(base_dir, PAX_PROFILE_FILE)
    if not os.path.exists(path):
        return {'dates': [], 'rows_by_date': {}}

    rows_by_date = defaultdict(list)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        for vals in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, vals))
            ts = row.get('Timestamp')
            if isinstance(ts, datetime):
                rows_by_date[ts.date()].append(row)
    except Exception:
        return {'dates': [], 'rows_by_date': {}}

    return {'dates': sorted(rows_by_date.keys()), 'rows_by_date': dict(rows_by_date)}


def _source_date_for_target(target_dt, source_dates):
    target_date = target_dt.date()
    if target_date in source_dates:
        return target_date

    offset = (target_date - datetime.now().date()).days
    if 0 <= offset < len(source_dates):
        return source_dates[offset]
    return None


def _touchpoint_value(row, key):
    total = 0.0
    suffix = f"_{key}"
    for col, value in row.items():
        if str(col or '').strip().lower().endswith(suffix):
            try:
                total += float(value or 0)
            except (TypeError, ValueError):
                pass
    return total


def _empty_series():
    data = {'labels': [], 'day_labels': [], 'days_data': []}
    for key in TOUCHPOINTS:
        data[key] = []
    return data


def _build_series_from_workbook(base_dir):
    profile = _load_pax_workbook_rows(base_dir)
    source_dates = profile.get('dates', [])
    rows_by_date = profile.get('rows_by_date', {})
    if not source_dates:
        return None

    data = _empty_series()
    today = datetime.now()
    for offset in (1, 2, 3):
        target_dt = today + timedelta(days=offset)
        source_date = _source_date_for_target(target_dt, source_dates)
        if source_date is None:
            continue

        day = {
            'date': target_dt.strftime('%a %d %b %Y'),
            'labels': [],
        }
        for key in TOUCHPOINTS:
            day[key] = []

        rows = sorted(rows_by_date.get(source_date, []), key=lambda r: r.get('Timestamp'))
        for row in rows:
            ts = row.get('Timestamp')
            if not isinstance(ts, datetime):
                continue

            label = ts.strftime('%H:%M')
            day_label = target_dt.strftime('%a %d %b')
            data['labels'].append(label)
            data['day_labels'].append(day_label)
            day['labels'].append(label)
            for key in TOUCHPOINTS:
                value = _touchpoint_value(row, key)
                data[key].append(value)
                day[key].append(value)

        data['days_data'].append(day)

    return data if data['labels'] else None


def _build_series_from_cache(base_dir):
    source = _load_json(base_dir, PAX_3DAY_FILE)
    data = {
        'labels': source.get('labels', []),
        'day_labels': source.get('day_labels', []),
        'days_data': source.get('days_data', []),
    }
    for key in TOUCHPOINTS:
        data[key] = _numbers(source.get(key))
    return data


def _peak_time_label(data, values):
    idx = _peak_idx(values)
    time_label = (data.get('labels') or ['--'])[idx] if idx < len(data.get('labels') or []) else '--'
    day_label = (data.get('day_labels') or [''])[idx] if idx < len(data.get('day_labels') or []) else ''
    day_short = ' '.join(str(day_label).split(' ')[1:])
    return f"{day_short} · {time_label}" if day_short else time_label


def _insight(data, key, label, accent, secondary_label, secondary_value, metric_label='Peak Pax / 15-min'):
    values = _numbers(data.get(key))
    return {
        'key': key,
        'label': label,
        'accent': accent,
        'peak': round(max(values) if values else 0),
        'metric_label': metric_label,
        'peak_time': _peak_time_label(data, values),
        'secondary_label': secondary_label,
        'secondary_value': secondary_value,
    }


def build_short_term_pax_outlook(base_dir):
    data = _build_series_from_workbook(base_dir) or _build_series_from_cache(base_dir)

    max_checkin = max(data['checkin'] or [0])
    max_security = max(data['security'] or [0])
    max_cbp = max(data['cbp'] or [0])
    max_lounge = max(data['lounge'] or [0])
    max_boarding = max(data['boarding'] or [0])
    max_immigration = max(data['immigration'] or [0])
    max_baggage = max(data['baggage'] or [0])
    avg_security = _active_avg(data['security'])

    insights = [
        _insight(data, 'checkin', 'Check-In', 'teal', 'Avg (all day)', f"{_active_avg(data['checkin']):,} pax"),
        _insight(data, 'security', 'Security', 'purple', 'Pressure Index', f"{(max_security / avg_security):.1f}x avg" if avg_security else '--'),
        _insight(data, 'cbp', 'CBP Preclearance', 'crit', '% of Security', f"{_pct(max_cbp, max_security) or '--'}% of security peak"),
        _insight(data, 'lounge', 'Lounge / Retail', 'warn', 'Avg Occupancy', f"{_active_avg(data['lounge']):,} pax", 'Peak Concurrent Pax'),
        _insight(data, 'boarding', 'Boarding Gates', 'info', 'Gate Pressure', f"{_pct(max_boarding, max_checkin) or '--'}% of check-in"),
        _insight(data, 'immigration', 'Immigration', 'ok', 'Intl. Arrival Share', f"{_pct(max_immigration, max_baggage) or '--'}% of arrivals"),
    ]

    return {
        'summary': {
            'title': '3-Day Strategic Outlook',
            'subtitle': 'passenger numbers engineered based on actual next 3-day flight schedules',
            'interval': '15-min',
            'lounge_peak': round(max_lounge),
        },
        'series': data,
        'insights': insights,
    }
